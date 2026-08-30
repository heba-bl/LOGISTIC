"""The shared workbook, and what the server does with what it sends.

The tests that matter here are the ones about trust. Excel enforces the
Maker-Checker rule on the shop floor; these check that the server enforces it
again, so a hand-edited cell buys nobody anything - and that no path through the
workbook can move the stock, because the stock rule lives elsewhere.
"""

from __future__ import annotations

import io
import zipfile

import pytest
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from sqlalchemy import func, select

from app.models.enums import RoleName, Zone
from app.models.flow import Lot
from app.models.organization import Role, User
from app.models.warehouse import StockMovement
from app.services import (
    excel_operations,
    excel_sync_service,
    stock_service,
    validation_token_service,
    whap_source,
)

APPROVED = excel_sync_service.STATUS_APPROVED
PENDING = excel_sync_service.STATUS_PENDING
DRAFT = excel_sync_service.STATUS_DRAFT


# --------------------------------------------------------------------- source
def test_source_nomenclature_is_read_unchanged():
    """The 2 200 supplied references arrive intact, codes and all."""
    articles = [a for a in whap_source.load_catalogue() if a.source == "WHAP"]

    assert len(articles) == 2200
    codes = [article.code for article in articles]
    assert len(set(codes)) == 2200
    assert codes[0] == "WHAP-0001"
    assert codes[-1] == "WHAP-2200"
    # Nothing invented in place of the file's own values.
    assert all(article.designation for article in articles)
    assert all(article.quantity_per_vehicle >= 1 for article in articles)


def test_catalogue_separates_vehicle_parts_from_warehouse_articles():
    """Paint and packaging are stored, but they are not part of a vehicle."""
    catalogue = whap_source.load_catalogue()
    summary = whap_source.catalogue_summary()

    assert summary["bom"] == 2200
    assert summary["total"] > summary["bom"], "l'entrepot doit contenir plus que le BOM"

    extras = [article for article in catalogue if not article.in_bom]
    assert extras, "aucun article hors nomenclature"
    assert all(article.source == "DEMO" for article in extras)
    families = {article.category for article in extras}
    assert whap_source.CATEGORY_MATERIAL in families
    assert whap_source.CATEGORY_PACKAGING in families


def test_catalogue_is_deterministic():
    whap_source.load_catalogue.cache_clear()
    first = [(a.reference, a.stock, a.location) for a in whap_source.load_catalogue()]
    whap_source.load_catalogue.cache_clear()
    second = [(a.reference, a.stock, a.location) for a in whap_source.load_catalogue()]
    assert first == second


def test_requirement_scales_with_the_number_of_vehicles():
    one = {row["reference"]: row["required"] for row in whap_source.requirement_for(1)}
    five = {row["reference"]: row["required"] for row in whap_source.requirement_for(5)}
    assert all(five[key] == one[key] * 5 for key in one)


# ------------------------------------------------------------------- workbook
@pytest.fixture(scope="module")
def workbook_bytes() -> bytes:
    return excel_operations.build_workbook()


def test_workbook_is_a_real_macro_enabled_package(workbook_bytes):
    """A genuine .xlsm: a zip carrying a VBA project and the ribbon."""
    assert zipfile.is_zipfile(io.BytesIO(workbook_bytes))

    with zipfile.ZipFile(io.BytesIO(workbook_bytes)) as archive:
        names = set(archive.namelist())
        content_types = archive.read("[Content_Types].xml").decode("utf-8")
        vba = archive.read("xl/vbaProject.bin")

    assert "xl/vbaProject.bin" in names, "aucun projet VBA"
    assert "customUI/customUI14.xml" in names, "aucun ruban"
    assert "macroEnabled.main+xml" in content_types
    # A compound file, which is what a VBA project is.
    assert vba[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"
    assert len(vba) > 10_000


def test_workbook_has_the_sixteen_sheets(workbook_bytes):
    """Fourteen an operator can see, plus the two they never should."""
    book = load_workbook(io.BytesIO(workbook_bytes), read_only=True, keep_vba=True)
    try:
        assert book.sheetnames == [
            "ACCUEIL", "UTILISATEURS", "ARTICLES", "BOM_VEHICULE", "RECEPTION",
            "INSPECTION", "QUALITE", "RED_CAGE", "WAREHOUSE", "MOUVEMENTS_STOCK",
            "PRODUCTION", "SORTIES", "EMPLACEMENTS", "HISTORIQUE", "LISTES",
            "CONFIGURATION",
        ]
    finally:
        book.close()


def test_articles_and_bom_carry_the_real_data(workbook_bytes):
    book = load_workbook(io.BytesIO(workbook_bytes), read_only=True, keep_vba=True)
    try:
        articles = book["ARTICLES"]
        assert articles.max_row - excel_operations.HEADER_ROW == 2239

        bom = book["BOM_VEHICULE"]
        assert bom.max_row - excel_operations.HEADER_ROW == 2200
        # The requirement column is a formula, so the sheet answers for itself.
        assert str(bom["I5"].value).startswith("=H5*$D$3")
    finally:
        book.close()


def test_validation_codes_never_appear_in_the_workbook(workbook_bytes):
    """The digests may ship; the codes may not."""
    with zipfile.ZipFile(io.BytesIO(workbook_bytes)) as archive:
        blob = b"".join(
            archive.read(name) for name in archive.namelist() if name.endswith(".xml")
        ).decode("utf-8", errors="ignore")

    for matricule, code in excel_operations.DEMO_CODES.items():
        assert code not in blob, f"le code de {matricule} est en clair dans le fichier"
        assert excel_operations.code_digest(matricule, code) in blob


def test_configuration_sheet_is_hidden_and_protected(workbook_bytes):
    book = load_workbook(io.BytesIO(workbook_bytes), keep_vba=True)
    try:
        config = book["CONFIGURATION"]
        assert config.sheet_state == "veryHidden"
        assert config.protection.sheet is True
        assert book["ARTICLES"].protection.sheet is True
    finally:
        book.close()


def test_entry_grid_stays_writable_under_protection(workbook_bytes):
    """Protection must stop tampering, not stop the operator working."""
    book = load_workbook(io.BytesIO(workbook_bytes), keep_vba=True)
    try:
        reception = book["RECEPTION"]
        assert reception.protection.sheet is True
        row = excel_operations.HEADER_ROW + 1
        columns = excel_operations.RECEPTION_COLUMNS + excel_operations.WORKFLOW_COLUMNS
        headers = [name for name, _, _ in columns]

        # What only the operator can know stays open.
        for name in ("REFERENCE_PIECE", "QUANTITE_RECUE", "MATRICULE_OPERATEUR"):
            cell = reception.cell(row=row, column=headers.index(name) + 1)
            assert cell.protection.locked is False, name

        # What the sheet works out for itself does not.
        for name in ("DATE", "HEURE", "DESIGNATION", "ID_RECEPTION"):
            cell = reception.cell(row=row, column=headers.index(name) + 1)
            assert cell.protection.locked is True, name

        # The vehicle count is the one cell the BOM expects to be changed.
        assert book["BOM_VEHICULE"]["D3"].protection.locked is False
    finally:
        book.close()


def _dropdown_columns(book, sheet_name: str) -> set[str]:
    """Which columns of a sheet offer a list."""
    sheet = book[sheet_name]
    headers = {
        get_column_letter(index): sheet.cell(
            row=excel_operations.HEADER_ROW, column=index
        ).value
        for index in range(1, sheet.max_column + 1)
    }
    return {
        headers.get(str(rule.sqref).split(":")[0].rstrip("0123456789"))
        for rule in sheet.data_validations.dataValidation
    }


def test_the_operator_chooses_rather_than_types(workbook_bytes):
    """Every field the file can answer by itself offers the answers.

    A typed part code is the likeliest way for a line to be rejected: the
    catalogue holds both `WHAP-0001` and `MOT-0001` for the same part, and only
    one of them resolves. A list removes the question.
    """
    book = load_workbook(io.BytesIO(workbook_bytes), keep_vba=True)
    try:
        expected = {
            "RECEPTION": {"REFERENCE_PIECE"},
            "QUALITE": {"DECISION"},
            "RED_CAGE": {"DECISION", "ORIGINE"},
            "WAREHOUSE": {"EMPLACEMENT"},
            "PRODUCTION": {"REFERENCE_PIECE", "PRIORITE"},
            "SORTIES": {"EMPLACEMENT"},
        }
        for name, wanted in expected.items():
            offered = _dropdown_columns(book, name)
            assert wanted <= offered, f"{name}: {wanted - offered} sans liste"
    finally:
        book.close()


def test_lists_that_come_from_slcc_reach_the_sheets(db, world):
    """Suppliers, matricules and stations are SLCC's to say, not Excel's.

    Built without a session the file offers only what it holds itself; given one
    it offers what the server currently knows. Neither case invents a list.
    """
    bare = load_workbook(io.BytesIO(excel_operations.build_workbook()), keep_vba=True)
    live = load_workbook(
        io.BytesIO(excel_operations.build_workbook(db=db)), keep_vba=True
    )
    try:
        assert "FOURNISSEUR" not in _dropdown_columns(bare, "RECEPTION")

        reception = _dropdown_columns(live, "RECEPTION")
        assert {"FOURNISSEUR", "MATRICULE_OPERATEUR"} <= reception
        assert "STATION" in _dropdown_columns(live, "PRODUCTION")
    finally:
        bare.close()
        live.close()


def test_the_stamp_is_written_by_the_sheet(workbook_bytes):
    """DATE and HEURE are the sheet's answer, not a question to the operator."""
    book = load_workbook(io.BytesIO(workbook_bytes), keep_vba=True)
    try:
        reception = book["RECEPTION"]
        headers = [name for name, _, _ in excel_operations.RECEPTION_COLUMNS]
        row = excel_operations.HEADER_ROW + 1
        date = reception.cell(row=row, column=headers.index("DATE") + 1)
        assert str(date.value).startswith("=IF(")
        assert "TODAY()" in str(date.value)
    finally:
        book.close()


def test_required_columns_are_marked_in_bold(workbook_bytes):
    """The macro reads "required" off the header font, so it must be set."""
    book = load_workbook(io.BytesIO(workbook_bytes), read_only=True, keep_vba=True)
    try:
        reception = book["RECEPTION"]
        # In read-only mode a cell past the used range has no font at all, so
        # the value is checked first.
        header_cells = [
            reception.cell(row=excel_operations.HEADER_ROW, column=index)
            for index in range(1, 26)
        ]
        bold = {cell.value for cell in header_cells if cell.value and cell.font and cell.font.bold}
        assert {
            "BON_LIVRAISON", "FOURNISSEUR", "REFERENCE_PIECE", "MATRICULE_OPERATEUR"
        } <= bold
        # A field the sheet works out for itself is never the operator's to fill.
        for computed in ("DESIGNATION", "ECART", "TOLERANCE_AUTORISEE", "RESULTAT_CONTROLE"):
            assert computed not in bold, computed
        assert "COMMENTAIRE" not in bold
    finally:
        book.close()


def test_digest_matches_between_workbook_and_server():
    """Excel and the API must judge a code by exactly the same rule."""
    assert excel_operations.CODE_SALT == excel_sync_service.CODE_SALT
    assert excel_operations.code_digest("RM-004", "REC2026") == excel_sync_service.code_digest(
        "RM-004", "REC2026"
    )


# ------------------------------------------------------------------- fixtures
@pytest.fixture()
def crew(db):
    """A maker, their manager, and a manager from another zone."""

    def role_for(name: RoleName, can_validate: bool) -> Role:
        role = db.execute(select(Role).where(Role.name == name)).scalar_one_or_none()
        if role is None:
            role = Role(name=name, label=name.value.title(), can_validate=can_validate)
            db.add(role)
            db.flush()
        role.can_validate = can_validate
        return role

    def user(matricule: str, name: RoleName, zone: Zone, can: bool, active: bool = True) -> User:
        person = User(
            employee_number=matricule,
            username=matricule.lower(),
            full_name=matricule,
            role_id=role_for(name, can).id,
            zone=zone,
            service=zone.value.title(),
            is_active=active,
        )
        if can:
            person.validation_code_hash = excel_sync_service.code_digest(matricule, "REC2026")
        db.add(person)
        db.flush()
        return person

    people = {
        "maker": user("OP-9001", RoleName.RECEPTIONIST, Zone.RECEPTION, False),
        "checker": user("RM-9002", RoleName.RECEPTION_MANAGER, Zone.RECEPTION, True),
        "other_zone": user("QM-9003", RoleName.QUALITY_MANAGER, Zone.QUALITY, True),
        "inactive": user("RM-9004", RoleName.RECEPTION_MANAGER, Zone.RECEPTION, True, active=False),
    }
    db.commit()
    return people


def _row(world, crew, **overrides) -> dict:
    """A line as the workbook sends it - signature included.

    Since the audit, an unsigned line is refused: the token is what proves a
    manager typed their code. Tests must therefore go through the same signing
    the workbook does, which is the point.
    """
    maker = overrides.get("matricule_operateur", crew["maker"].employee_number)
    checker = overrides.get("matricule_checker", crew["checker"].employee_number)
    sync_id = overrides.get("id_sync", "SLCC-REC-0001-20260822103500")
    sheet = overrides.pop("__sheet__", "RECEPTION")

    row = {
        "jeton_validation": validation_token_service.build_token(
            sheet=sheet, sync_id=sync_id, maker=maker, checker=checker
        ),
        "id_reception": "REC-0001",
        "date": "22/08/2026",
        "bl": "BL-2026-001",
        "fournisseur": world["supplier"].code,
        "reference_piece": world["small"].reference,
        "quantite_attendue": "200",
        "quantite_recue": "200",
        "matricule_operateur": crew["maker"].employee_number,
        "statut": APPROVED,
        "matricule_checker": crew["checker"].employee_number,
        "id_sync": "SLCC-REC-0001-20260822103500",
        "source_row": 5,
    }
    row.update(overrides)
    return row


# ------------------------------------------------------------ server re-checks
def test_a_draft_row_is_never_operational_data(db, world, crew):
    for status in (DRAFT, PENDING, "REJETE", ""):
        reason = excel_sync_service.validate_row(
            db, "RECEPTION", _row(world, crew, statut=status)
        )
        assert reason and "VALIDE" in reason


def test_maker_cannot_be_their_own_checker(db, world, crew):
    """The rule the whole workflow exists for, enforced away from Excel."""
    reason = excel_sync_service.validate_row(
        db,
        "RECEPTION",
        _row(world, crew, matricule_checker=crew["maker"].employee_number),
    )
    assert reason is not None
    assert "maker = checker" in reason


def test_checker_from_another_zone_is_refused(db, world, crew):
    reason = excel_sync_service.validate_row(
        db,
        "RECEPTION",
        _row(world, crew, matricule_checker=crew["other_zone"].employee_number),
    )
    assert reason is not None and "zone" in reason


def test_inactive_or_unknown_checker_is_refused(db, world, crew):
    inactive = excel_sync_service.validate_row(
        db, "RECEPTION", _row(world, crew, matricule_checker=crew["inactive"].employee_number)
    )
    assert inactive is not None and "inactif" in inactive

    unknown = excel_sync_service.validate_row(
        db, "RECEPTION", _row(world, crew, matricule_checker="ZZ-0000")
    )
    assert unknown is not None and "inconnu" in unknown


def test_a_correct_row_passes(db, world, crew):
    assert excel_sync_service.validate_row(db, "RECEPTION", _row(world, crew)) is None


def test_validation_code_is_checked_against_the_digest(db, world, crew):
    matricule = crew["checker"].employee_number
    assert excel_sync_service.verify_validation_code(db, matricule, "REC2026") is True
    assert excel_sync_service.verify_validation_code(db, matricule, "MAUVAIS") is False
    # An operator has no code at all, whatever they type.
    assert (
        excel_sync_service.verify_validation_code(db, crew["maker"].employee_number, "REC2026")
        is False
    )


# -------------------------------------------------------------- the stock rule
def test_synchronising_a_validated_reception_creates_a_lot_but_no_stock(db, world, crew):
    """The rule that must never bend, exercised through the Excel path."""
    lots_before = db.execute(select(func.count()).select_from(Lot)).scalar_one()
    moves_before = db.execute(select(func.count()).select_from(StockMovement)).scalar_one()
    stock_before = stock_service.get_available(db, world["small"].id)

    outcome = excel_sync_service.sync_rows(
        db,
        sheet="RECEPTION",
        file_name=excel_operations.WORKBOOK_NAME,
        rows=[_row(world, crew)],
    )

    assert outcome.accepted == 1, outcome.rows[0].reason
    assert db.execute(select(func.count()).select_from(Lot)).scalar_one() == lots_before + 1
    # The whole point: a validated reception is still not stock.
    assert (
        db.execute(select(func.count()).select_from(StockMovement)).scalar_one() == moves_before
    )
    assert stock_service.get_available(db, world["small"].id) == stock_before


def test_unvalidated_rows_create_nothing_at_all(db, world, crew):
    lots_before = db.execute(select(func.count()).select_from(Lot)).scalar_one()

    outcome = excel_sync_service.sync_rows(
        db,
        sheet="RECEPTION",
        file_name=excel_operations.WORKBOOK_NAME,
        rows=[
            _row(world, crew, statut=DRAFT, id_sync="A"),
            _row(world, crew, statut=PENDING, id_sync="B"),
        ],
    )

    assert outcome.accepted == 0
    assert outcome.rejected == 2
    assert db.execute(select(func.count()).select_from(Lot)).scalar_one() == lots_before


def test_the_same_line_is_never_taken_in_twice(db, world, crew):
    """Pressing sync twice must not duplicate the reception."""
    row = _row(world, crew)
    first = excel_sync_service.sync_rows(
        db, sheet="RECEPTION", file_name=excel_operations.WORKBOOK_NAME, rows=[row]
    )
    assert first.accepted == 1

    lots_after_first = db.execute(select(func.count()).select_from(Lot)).scalar_one()

    second = excel_sync_service.sync_rows(
        db, sheet="RECEPTION", file_name=excel_operations.WORKBOOK_NAME, rows=[row]
    )
    assert second.accepted == 0
    assert second.duplicates == 1
    assert db.execute(select(func.count()).select_from(Lot)).scalar_one() == lots_after_first


def test_one_bad_row_does_not_hold_back_the_good_ones(db, world, crew):
    outcome = excel_sync_service.sync_rows(
        db,
        sheet="RECEPTION",
        file_name=excel_operations.WORKBOOK_NAME,
        rows=[
            _row(world, crew, id_sync="OK-1"),
            _row(world, crew, id_sync="BAD-1", reference_piece="N-EXISTE-PAS"),
            _row(world, crew, id_sync="OK-2", statut=DRAFT),
        ],
    )
    assert outcome.received == 3
    assert outcome.accepted == 1
    assert outcome.rejected == 2
    reasons = [row.reason for row in outcome.rows if not row.accepted]
    assert any("inconnue" in (reason or "") for reason in reasons)


def test_sync_records_who_entered_and_who_validated(db, world, crew):
    """Traceability: the batch keeps both matricules and the source file."""
    from app.models.imports import DataImport

    excel_sync_service.sync_rows(
        db, sheet="RECEPTION", file_name=excel_operations.WORKBOOK_NAME, rows=[_row(world, crew)]
    )
    batch = db.execute(
        select(DataImport).order_by(DataImport.id.desc()).limit(1)
    ).scalar_one()

    assert batch.maker_reference == crew["maker"].employee_number
    assert batch.checker_reference == crew["checker"].employee_number
    assert batch.maker_reference != batch.checker_reference
    assert batch.source_filename == excel_operations.WORKBOOK_NAME
    assert batch.checked_at is not None


def test_unknown_sheet_is_refused(db):
    from app.core.exceptions import ValidationError

    with pytest.raises(ValidationError):
        excel_sync_service.sync_rows(db, sheet="INVENTEE", file_name="x.xlsm", rows=[])


@pytest.fixture()
def client(db, world):
    """TestClient bound to the same session the other fixtures write to."""
    from fastapi.testclient import TestClient

    from app.api.deps import get_session
    from app.main import app

    app.dependency_overrides[get_session] = lambda: db
    with TestClient(app) as test_client:
        yield test_client
    app.dependency_overrides.clear()


# ---------------------------------------------------------- supervision view
def test_status_separates_activity_from_validation(db, world, crew):
    """The page shows two different things; the payload must keep them apart."""
    excel_sync_service.sync_rows(
        db,
        sheet="RECEPTION",
        file_name=excel_operations.WORKBOOK_NAME,
        rows=[_row(world, crew)],
    )
    status = excel_sync_service.workbook_status(db)

    # Activity counts real records, whatever created them.
    assert status["activity"]["receptions"] >= 1
    assert set(status["activity"]) == {
        "receptions", "inspections", "quality", "red_cage",
        "warehouse_articles", "stock_movements", "production_requests", "issues",
    }

    # Validation counts the batches that came from the workbook.
    assert status["batches"]["approved"] == 1
    assert status["batches"]["pending"] == 0
    assert status["state"] == "SYNCED"
    assert status["last_maker"] == crew["maker"].employee_number
    assert status["last_actor"] == crew["checker"].employee_number


def test_status_says_never_synced_before_anything_arrives(db, world):
    status = excel_sync_service.workbook_status(db)
    assert status["state"] == "NEVER_SYNCED"
    assert status["last_sync_at"] is None
    assert status["batches"]["total"] == 0


def test_history_carries_both_signatures(db, world, crew):
    excel_sync_service.sync_rows(
        db,
        sheet="RECEPTION",
        file_name=excel_operations.WORKBOOK_NAME,
        rows=[_row(world, crew)],
    )
    entries = excel_sync_service.sync_history(db)

    assert len(entries) == 1
    entry = entries[0]
    assert entry["maker_reference"] == crew["maker"].employee_number
    assert entry["checker_reference"] == crew["checker"].employee_number
    assert entry["source_filename"] == excel_operations.WORKBOOK_NAME
    assert entry["result_references"], "les enregistrements crees doivent etre traces"


def test_history_filters_narrow_the_result(db, world, crew):
    excel_sync_service.sync_rows(
        db,
        sheet="RECEPTION",
        file_name=excel_operations.WORKBOOK_NAME,
        rows=[_row(world, crew)],
    )

    # A matricule matches whether the person entered it or signed it off.
    assert len(excel_sync_service.sync_history(db, matricule=crew["maker"].employee_number)) == 1
    assert len(excel_sync_service.sync_history(db, matricule=crew["checker"].employee_number)) == 1
    assert len(excel_sync_service.sync_history(db, matricule="ZZ-0000")) == 0

    assert len(excel_sync_service.sync_history(db, status="APPROVED")) == 1
    assert len(excel_sync_service.sync_history(db, status="REJECTED")) == 0
    assert len(excel_sync_service.sync_history(db, import_type="RECEPTION")) == 1
    assert len(excel_sync_service.sync_history(db, import_type="INSPECTION")) == 0


def test_status_endpoint_answers_over_http(client, world, crew):
    """The page calls this; a shape change must break a test, not the screen."""
    response = client.get("/api/excel/status")
    assert response.status_code == 200
    payload = response.json()

    for key in ("workbook", "state", "activity", "warehouse", "batches", "per_process"):
        assert key in payload, key
    assert payload["workbook"].endswith(".xlsm")
    assert payload["state"] in {"SYNCED", "PENDING", "NEVER_SYNCED"}


def test_history_endpoint_answers_over_http(client, world, crew):
    response = client.get("/api/excel/history", params={"limit": 5})
    assert response.status_code == 200
    assert isinstance(response.json(), list)


# ------------------------------------------------------- live stock in Excel
def _store_split(db, world, quantities: tuple[int, int]) -> None:
    """Receive a lot and shelve it across the two addresses of SM-100."""
    from app.services import inspection_service, quality_service, reception_service
    from app.services.warehouse_service import Allocation, confirm_storage

    total = sum(quantities)
    actor = world["user"]
    reception = reception_service.create_reception(
        db,
        part_id=world["small"].id,
        supplier_id=world["supplier"].id,
        quantity_expected=total,
        quantity_received=total,
        actor_id=actor.id,
    )
    lot_id = reception.lot_id
    inspection_service.start_inspection(db, lot_id=lot_id, actor_id=actor.id)
    inspection_service.record_inspection(
        db, lot_id=lot_id, sample_size=5, defects_found=0, actor_id=actor.id
    )
    quality_service.approve(
        db, lot_id=lot_id, justification="conforme", actor_id=actor.id
    )
    confirm_storage(
        db,
        lot_id=lot_id,
        allocations=[
            Allocation(location_id=world["primary"].id, quantity=quantities[0]),
            Allocation(location_id=world["secondary"].id, quantity=quantities[1]),
        ],
        actor_id=actor.id,
    )
    db.commit()


def test_live_stock_reports_the_balance_the_database_holds(db, world):
    """The figure comes from `Stock`, which `stock_service` owns."""
    _store_split(db, world, (350, 150))

    live = excel_operations.live_stock(db)
    entry = live[world["small"].reference]

    assert entry["available"] == 500
    # Located and available must agree; the sheet shows both so they can't
    # silently diverge.
    assert entry["located"] == 500


def test_live_stock_sees_a_lot_split_across_two_addresses(db, world):
    """A split lot keeps one address on the lot row, so the ledger is read."""
    _store_split(db, world, (350, 150))

    entry = excel_operations.live_stock(db)[world["small"].reference]
    placed = dict(entry["placed"])

    assert placed == {world["primary"].code: 350, world["secondary"].code: 150}
    assert entry["primary"] == world["primary"].code
    assert entry["secondary"] == [(world["secondary"].code, 150)]


def test_articles_sheet_carries_the_live_stock(db, world):
    _store_split(db, world, (350, 150))
    content = excel_operations.build_workbook(db=db)

    book = load_workbook(io.BytesIO(content), read_only=True, keep_vba=True)
    try:
        sheet = book["ARTICLES"]
        headers = [
            sheet.cell(row=excel_operations.HEADER_ROW, column=column).value
            for column in range(1, 20)
            if sheet.cell(row=excel_operations.HEADER_ROW, column=column).value
        ]
        for column in (
            "STOCK", "STOCK_TOTAL", "EMPLACEMENT_PRINCIPAL",
            "EMPLACEMENTS_SECONDAIRES", "DERNIERE_SYNCHRONISATION",
        ):
            assert column in headers, column

        index = {name: position for position, name in enumerate(headers)}
        rows = [
            row
            for row in sheet.iter_rows(
                min_row=excel_operations.HEADER_ROW + 1, values_only=True
            )
            if row[0]
        ]
        # The catalogue is untouched by this change.
        assert len(rows) == 2239

        # SM-100 belongs to the test world, not the catalogue, so the sheet
        # shows the catalogue references with a zero balance - which is the
        # honest answer for a reference the warehouse has never held.
        assert all(row[index["STOCK"]] == row[index["STOCK_TOTAL"]] for row in rows)
        assert all(row[index["DERNIERE_SYNCHRONISATION"]] for row in rows)
    finally:
        book.close()


def test_without_a_database_the_sheet_says_it_is_not_synchronised(db, world):
    """A workbook built with no session must not present a stock figure as live."""
    content = excel_operations.build_workbook()

    book = load_workbook(io.BytesIO(content), read_only=True, keep_vba=True)
    try:
        sheet = book["ARTICLES"]
        headers = [
            sheet.cell(row=excel_operations.HEADER_ROW, column=column).value
            for column in range(1, 20)
            if sheet.cell(row=excel_operations.HEADER_ROW, column=column).value
        ]
        index = {name: position for position, name in enumerate(headers)}
        first = next(
            row
            for row in sheet.iter_rows(
                min_row=excel_operations.HEADER_ROW + 1, values_only=True
            )
            if row[0]
        )
        assert first[index["DERNIERE_SYNCHRONISATION"]] == excel_operations.NOT_SYNCED
    finally:
        book.close()


def test_excel_never_becomes_a_second_source_of_stock(db, world):
    """Reading the sheet must not change a single balance."""
    _store_split(db, world, (350, 150))
    before = stock_service.get_available(db, world["small"].id)
    movements_before = db.execute(
        select(func.count()).select_from(StockMovement)
    ).scalar_one()

    excel_operations.build_workbook(db=db)

    assert stock_service.get_available(db, world["small"].id) == before
    assert (
        db.execute(select(func.count()).select_from(StockMovement)).scalar_one()
        == movements_before
    )


# --------------------------------------------------- the signature is the proof
def test_an_unsigned_row_is_refused(db, world, crew):
    """A status typed by hand is not a validation.

    This is the hole the audit found: the workbook said VALIDE and the server
    believed it. Now the server wants proof that it issued itself.
    """
    row = _row(world, crew)
    row.pop("jeton_validation")

    reason = excel_sync_service.validate_row(db, "RECEPTION", row)
    assert reason is not None and "jeton" in reason


def test_a_token_minted_for_someone_else_is_refused(db, world, crew):
    """The signature is bound to the pair of people it was issued for."""
    row = _row(world, crew)
    row["matricule_checker"] = crew["other_zone"].employee_number

    reason = excel_sync_service.validate_row(db, "RECEPTION", row)
    assert reason is not None


def test_a_token_from_another_sheet_is_refused(db, world, crew):
    """A signature cannot be carried from one sheet to another."""
    row = _row(world, crew)
    row["jeton_validation"] = validation_token_service.build_token(
        sheet="PRODUCTION",
        sync_id=row["id_sync"],
        maker=row["matricule_operateur"],
        checker=row["matricule_checker"],
    )

    reason = excel_sync_service.validate_row(db, "RECEPTION", row)
    assert reason is not None and "jeton" in reason


def test_a_tampered_token_is_refused(db, world, crew):
    row = _row(world, crew)
    row["jeton_validation"] = row["jeton_validation"][:-4] + "0000"

    reason = excel_sync_service.validate_row(db, "RECEPTION", row)
    assert reason is not None and "jeton" in reason


def test_the_code_is_checked_against_the_stored_digest(db, world, crew):
    """The server, not the workbook, decides whether a code is right."""
    matricule = crew["checker"].employee_number
    salt = excel_operations.CODE_SALT

    assert validation_token_service.verify_code(db, matricule, "REC2026", salt) is True
    assert validation_token_service.verify_code(db, matricule, "MAUVAIS", salt) is False
    # An operator has no code recorded, so nothing they type can work.
    assert (
        validation_token_service.verify_code(
            db, crew["maker"].employee_number, "REC2026", salt
        )
        is False
    )


def test_the_signing_secret_never_leaves_the_server(db, world, crew):
    """The token must not be derivable from anything the workbook holds."""
    secret = validation_token_service.signing_secret()
    token = validation_token_service.build_token(
        sheet="RECEPTION", sync_id="X-1", maker="OP-1", checker="RM-1"
    )
    assert secret not in token.encode()
    assert len(token) == 64


def test_sheets_open_on_real_history_that_cannot_be_resent(db, world):
    """Pre-filled lines are operations SLCC already holds.

    They exist so a sheet opens on something legible instead of sixty blank
    rows. Two things must hold: they are read-only, and the sync macro skips
    them - it sends VALIDE lines whose ETAT_SYNC is not SYNCHRONISE, so marking
    them is what stops history being pushed a second time.
    """
    from app.services import excel_sync_service, reception_service

    recorded = reception_service.create_reception(
        db,
        part_id=world["small"].id,
        supplier_id=world["supplier"].id,
        quantity_expected=100,
        quantity_received=100,
        actor_id=world["user"].id,
    )
    db.flush()

    book = load_workbook(
        io.BytesIO(excel_operations.build_workbook(db=db)), keep_vba=True
    )
    try:
        reception = book["RECEPTION"]
        columns = excel_operations.RECEPTION_COLUMNS + excel_operations.WORKFLOW_COLUMNS
        headers = [name for name, _, _ in columns]
        row = excel_operations.HEADER_ROW + 1

        def value(name):
            return reception.cell(row=row, column=headers.index(name) + 1).value

        # The sheet shows the operation that just happened, not an invention.
        assert value("ID_RECEPTION") == recorded.reference
        assert value("QUANTITE_RECUE") == 100
        assert value("STATUT") == excel_sync_service.STATUS_APPROVED
        assert value("ETAT_SYNC") == excel_operations.SYNC_DONE

        # Nothing on a recorded line is open to editing.
        for name in ("REFERENCE_PIECE", "QUANTITE_RECUE", "MATRICULE_OPERATEUR"):
            cell = reception.cell(row=row, column=headers.index(name) + 1)
            assert cell.protection.locked is True, name

        # The date is the day it happened, not a formula that would say today.
        assert not str(value("DATE")).startswith("=")

        # The blank rows underneath are still open for work.
        blank = excel_operations.HEADER_ROW + 1 + 40
        assert reception.cell(
            row=blank, column=headers.index("REFERENCE_PIECE") + 1
        ).protection.locked is False
    finally:
        book.close()
