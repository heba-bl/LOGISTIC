"""Excel import and the Maker-Checker rule.

The rule under test: imported data is never definitive until a habilitated
checker - a different person from the maker - confirms it, and no business
record or stock movement exists before that.
"""

from __future__ import annotations

import io

import pytest
from sqlalchemy import func, select

from app.core.exceptions import ValidationError, WorkflowError
from app.models.enums import (
    AuditAction,
    ImportRowStatus,
    ImportStatus,
    ImportType,
    LotStatus,
    RoleName,
    ValidationDecision,
)
from app.models.flow import Lot
from app.models.organization import Role, User
from app.models.system import AuditLog
from app.models.warehouse import StockMovement
from app.services import import_service, stock_service


# --------------------------------------------------------------------- helpers
def _user(db, world, *, matricule: str, username: str, role_name: RoleName, service: str):
    role = db.execute(select(Role).where(Role.name == role_name)).scalar_one_or_none()
    if role is None:
        role = Role(name=role_name, label=role_name.value.replace("_", " ").title())
        db.add(role)
        db.flush()
    user = User(
        employee_number=matricule,
        username=username,
        full_name=username.replace(".", " ").title(),
        role_id=role.id,
        service=service,
    )
    db.add(user)
    db.flush()
    return user


@pytest.fixture()
def operators(db, world):
    """A maker and several potential checkers, all identified."""
    return {
        "maker": _user(
            db,
            world,
            matricule="OP-1042",
            username="k.moreau",
            role_name=RoleName.RECEPTIONIST,
            service="Reception",
        ),
        "reception_manager": _user(
            db,
            world,
            matricule="RM-004",
            username="f.chaoui",
            role_name=RoleName.RECEPTION_MANAGER,
            service="Reception",
        ),
        "inspector": _user(
            db,
            world,
            matricule="QL-1045",
            username="s.haddad",
            role_name=RoleName.QUALITY_INSPECTOR,
            service="Quality",
        ),
        "quality_manager": _user(
            db,
            world,
            matricule="QM-002",
            username="n.benali",
            role_name=RoleName.QUALITY_MANAGER,
            service="Quality",
        ),
        "warehouse": _user(
            db,
            world,
            matricule="WH-008",
            username="y.tazi",
            role_name=RoleName.WAREHOUSE_OPERATOR,
            service="Warehouse",
        ),
    }


def _csv(rows: list[str]) -> bytes:
    header = "part_reference,supplier_code,quantity_expected,quantity_received,delivery_note,notes"
    return ("\n".join([header, *rows])).encode("utf-8")


def _reception_file(world, quantity: int = 100) -> bytes:
    return _csv(
        [f"{world['small'].reference},{world['supplier'].code},{quantity},{quantity},BL-1,"]
    )


def _upload(db, world, operators, *, content: bytes | None = None, maker=None):
    return import_service.create_import(
        db,
        import_type=ImportType.RECEPTION,
        filename="receptions.csv",
        content=content if content is not None else _reception_file(world),
        maker_id=(maker or operators["maker"]).id,
    )


# ------------------------------------------------------------------ identity
def test_operators_are_never_anonymous(db, operators):
    maker = operators["maker"]
    assert maker.employee_number == "OP-1042"
    assert maker.service == "Reception"
    assert maker.role_name == "RECEPTIONIST"
    # The identity string used in the audit trail carries matricule, name and role.
    assert maker.identity.startswith("OP-1042 - ")
    assert "Receptionist" in maker.identity or "RECEPTIONIST" in maker.identity


def test_employee_number_is_unique(db, world, operators):
    from sqlalchemy.exc import IntegrityError

    duplicate = User(
        employee_number="OP-1042",
        username="another",
        full_name="Another Person",
        role_id=operators["maker"].role_id,
    )
    db.add(duplicate)
    with pytest.raises(IntegrityError):
        db.flush()
    db.rollback()


# -------------------------------------------------------------------- upload
def test_upload_creates_no_business_record_and_no_stock(db, world, operators):
    before_lots = db.execute(select(func.count()).select_from(Lot)).scalar_one()
    before_stock = stock_service.get_available(db, world["small"].id)

    batch = _upload(db, world, operators)

    assert batch.status is ImportStatus.PENDING_REVIEW
    assert batch.valid_row_count == 1
    assert db.execute(select(func.count()).select_from(Lot)).scalar_one() == before_lots
    assert stock_service.get_available(db, world["small"].id) == before_stock
    assert db.execute(select(func.count()).select_from(StockMovement)).scalar_one() == 0


def test_upload_records_the_maker_and_the_file_hash(db, world, operators):
    content = _reception_file(world)
    batch = _upload(db, world, operators, content=content)

    import hashlib

    assert batch.maker_reference == "OP-1042"
    assert batch.maker_role == "RECEPTIONIST"
    assert batch.maker_service == "Reception"
    assert batch.source_filename == "receptions.csv"
    assert batch.source_hash == hashlib.sha256(content).hexdigest()
    assert batch.source_size_bytes == len(content)
    assert batch.submitted_at is not None


def test_maker_must_be_habilitated_for_the_import_type(db, world, operators):
    # A warehouse operator does not enter receptions.
    with pytest.raises(ValidationError, match="not habilitated"):
        _upload(db, world, operators, maker=operators["warehouse"])


def test_inactive_operator_cannot_submit(db, world, operators):
    operators["maker"].is_active = False
    db.flush()
    with pytest.raises(ValidationError, match="inactive"):
        _upload(db, world, operators)


def test_invalid_rows_are_flagged_not_rejected_silently(db, world, operators):
    content = _csv(
        [
            f"{world['small'].reference},{world['supplier'].code},100,100,BL-1,",
            "UNKNOWN-REF,SUP,50,50,BL-2,",
            f"{world['small'].reference},{world['supplier'].code},abc,10,BL-3,",
        ]
    )
    batch = _upload(db, world, operators, content=content)

    assert batch.row_count == 3
    assert batch.valid_row_count == 1
    assert batch.invalid_row_count == 2

    invalid = [row for row in batch.rows if row.status is ImportRowStatus.INVALID]
    assert len(invalid) == 2
    assert all(row.error_message for row in invalid)
    assert any("unknown part reference" in row.error_message for row in invalid)


# ------------------------------------------------------- segregation of duties
def test_maker_cannot_validate_their_own_import(db, world, operators):
    batch = _upload(db, world, operators)

    with pytest.raises(WorkflowError, match="cannot validate it"):
        import_service.approve_import(
            db, import_id=batch.id, checker_id=operators["maker"].id
        )
    assert batch.status is ImportStatus.PENDING_REVIEW


def test_checker_must_hold_a_habilitated_role(db, world, operators):
    batch = _upload(db, world, operators)

    # A warehouse operator is not a reception responsible.
    with pytest.raises(WorkflowError, match="not habilitated"):
        import_service.approve_import(
            db, import_id=batch.id, checker_id=operators["warehouse"].id
        )
    assert batch.status is ImportStatus.PENDING_REVIEW


def test_inspector_cannot_validate_their_own_inspection(db, world, operators):
    """A Quality Inspector is never a Quality Manager for their own entry."""
    assert (
        RoleName.QUALITY_INSPECTOR
        not in import_service.CHECKER_ROLES[ImportType.INSPECTION]
    )
    assert (
        RoleName.QUALITY_MANAGER in import_service.CHECKER_ROLES[ImportType.INSPECTION]
    )


def test_inactive_checker_cannot_validate(db, world, operators):
    batch = _upload(db, world, operators)
    operators["reception_manager"].is_active = False
    db.flush()

    with pytest.raises(ValidationError, match="inactive"):
        import_service.approve_import(
            db, import_id=batch.id, checker_id=operators["reception_manager"].id
        )


def test_eligible_checkers_exclude_the_maker(db, world, operators):
    batch = _upload(db, world, operators)
    checkers = import_service.eligible_checkers(db, batch.id)

    references = {user.employee_number for user in checkers}
    assert "OP-1042" not in references, "the maker must never be offered as checker"
    assert "RM-004" in references


# ------------------------------------------------------------------- approval
def test_approval_applies_the_rows_and_records_both_identities(db, world, operators):
    batch = _upload(db, world, operators)
    before_lots = db.execute(select(func.count()).select_from(Lot)).scalar_one()

    approved = import_service.approve_import(
        db,
        import_id=batch.id,
        checker_id=operators["reception_manager"].id,
        comment="Quantities checked against the delivery note",
    )

    assert approved.status is ImportStatus.APPROVED
    assert approved.decision is ValidationDecision.APPROVED
    assert approved.applied_row_count == 1
    assert approved.checker_reference == "RM-004"
    assert approved.checker_role == "RECEPTION_MANAGER"
    assert approved.checked_at is not None
    # Maker identity is preserved alongside the checker.
    assert approved.maker_reference == "OP-1042"

    # The lot exists only now.
    assert db.execute(select(func.count()).select_from(Lot)).scalar_one() == before_lots + 1
    row = approved.rows[0]
    assert row.status is ImportRowStatus.APPLIED
    assert row.result_reference is not None


def test_approval_still_creates_no_stock(db, world, operators):
    """A validated reception creates a lot, never stock."""
    batch = _upload(db, world, operators)
    before = stock_service.get_available(db, world["small"].id)

    import_service.approve_import(
        db, import_id=batch.id, checker_id=operators["reception_manager"].id
    )

    assert stock_service.get_available(db, world["small"].id) == before
    assert db.execute(select(func.count()).select_from(StockMovement)).scalar_one() == 0

    lot = db.execute(select(Lot).order_by(Lot.id.desc()).limit(1)).scalar_one()
    assert lot.status is LotStatus.PENDING_INSPECTION


def test_a_decided_import_cannot_be_decided_again(db, world, operators):
    batch = _upload(db, world, operators)
    import_service.approve_import(
        db, import_id=batch.id, checker_id=operators["reception_manager"].id
    )

    with pytest.raises(WorkflowError, match="already been decided"):
        import_service.approve_import(
            db, import_id=batch.id, checker_id=operators["reception_manager"].id
        )


# ------------------------------------------------------------------ rejection
def test_rejection_requires_a_comment(db, world, operators):
    batch = _upload(db, world, operators)

    with pytest.raises(ValidationError, match="requires a comment"):
        import_service.reject_import(
            db, import_id=batch.id, checker_id=operators["reception_manager"].id, comment="  "
        )


def test_rejection_applies_nothing_but_stays_traceable(db, world, operators):
    batch = _upload(db, world, operators)
    before_lots = db.execute(select(func.count()).select_from(Lot)).scalar_one()

    rejected = import_service.reject_import(
        db,
        import_id=batch.id,
        checker_id=operators["reception_manager"].id,
        comment="Delivery note does not match the physical count",
    )

    assert rejected.status is ImportStatus.REJECTED
    assert rejected.decision is ValidationDecision.REJECTED
    assert rejected.decision_comment.startswith("Delivery note")
    assert rejected.applied_row_count == 0
    assert rejected.checker_reference == "RM-004"

    # Nothing was written to the business tables.
    assert db.execute(select(func.count()).select_from(Lot)).scalar_one() == before_lots
    assert db.execute(select(func.count()).select_from(StockMovement)).scalar_one() == 0

    # But the data itself is still there, for the record.
    assert all(row.status is ImportRowStatus.REJECTED for row in rejected.rows)
    assert rejected.rows[0].payload_json


# ---------------------------------------------------------------- audit trail
def test_audit_trail_records_maker_checker_decision_and_file(db, world, operators):
    batch = _upload(db, world, operators)
    import_service.approve_import(
        db,
        import_id=batch.id,
        checker_id=operators["reception_manager"].id,
        comment="Checked",
    )

    entries = db.execute(select(AuditLog).order_by(AuditLog.id)).scalars().all()

    submitted = next(e for e in entries if e.action.value == "IMPORT_SUBMITTED")
    assert submitted.maker_reference == "OP-1042"
    assert submitted.maker_role == "RECEPTIONIST"
    assert submitted.actor_reference == "OP-1042"
    assert submitted.source_file == "receptions.csv"
    assert submitted.source_hash

    approved = next(e for e in entries if e.action.value == "IMPORT_APPROVED")
    assert approved.maker_reference == "OP-1042"
    assert approved.maker_role == "RECEPTIONIST"
    assert approved.checker_reference == "RM-004"
    assert approved.checker_role == "RECEPTION_MANAGER"
    assert approved.decision == "APPROVED"
    assert approved.source_hash == batch.source_hash
    assert approved.status_before == "PENDING_REVIEW"
    assert approved.status_after == "APPROVED"


def test_business_actions_record_the_employee_number(db, world, operators):
    """Even outside imports, the trail carries the matricule."""
    batch = _upload(db, world, operators)
    import_service.approve_import(
        db, import_id=batch.id, checker_id=operators["reception_manager"].id
    )

    received = db.execute(
        select(AuditLog).where(AuditLog.action == "LOT_RECEIVED").order_by(AuditLog.id.desc())
    ).scalars().first()
    assert received is not None
    assert received.actor_reference == "OP-1042"
    assert received.actor_role == "RECEPTIONIST"


# ------------------------------------------------------------------ templates
def test_template_is_a_readable_workbook_with_the_expected_header(db):
    from openpyxl import load_workbook

    for import_type in ImportType:
        content = import_service.build_template(import_type)
        workbook = load_workbook(io.BytesIO(content), read_only=True)
        header = [cell.value for cell in next(workbook.active.iter_rows(max_row=1))]
        expected = [name for name, _ in import_service.COLUMNS[import_type]]
        assert header == expected, import_type
        workbook.close()


# ------------------------------------------------------ provenance on the lot
def test_the_lot_audit_entry_names_who_validated_it(db, world, operators):
    """A lot born from a validated spreadsheet must say who approved it.

    The business services know nothing about imports, so they stamp only the
    operator who typed the line. That left the traceability screen showing an
    empty "who validated" column for a record that had in fact been validated -
    the information existed, on the import batch, but not where anyone looks
    for it.
    """
    from app.models.system import AuditLog

    batch = _upload(db, world, operators)
    approved = import_service.approve_import(
        db,
        import_id=batch.id,
        checker_id=operators["reception_manager"].id,
        comment="ok",
    )

    entry = db.execute(
        select(AuditLog)
        .where(AuditLog.action == AuditAction.LOT_RECEIVED)
        .order_by(AuditLog.id.desc())
        .limit(1)
    ).scalar_one()

    # Who did the work, and who let it through.
    assert entry.actor_reference == "OP-1042"
    assert entry.maker_reference == "OP-1042"
    assert entry.checker_reference == "RM-004"
    assert entry.checker_role == "RECEPTION_MANAGER"
    assert entry.decision == ValidationDecision.APPROVED.value
    # And the file it came from, so the record can be traced back to a source.
    assert entry.source_file == approved.source_filename
    assert entry.source_hash == approved.source_hash


def test_provenance_never_overwrites_what_a_service_already_set(db, world, operators):
    """Stamping fills the gaps; it does not rewrite a deliberate entry."""
    from app.models.system import AuditLog
    from app.services import audit_service

    mark = audit_service.high_water_mark(db)
    deliberate = audit_service.record(
        db,
        action=AuditAction.IMPORT_APPROVED,
        entity_type="test",
        checker=operators["quality_manager"],
        decision="REJECTED",
        source_file="deja-renseigne.xlsx",
    )

    audit_service.stamp_provenance(
        db,
        since_id=mark,
        maker=operators["maker"],
        checker=operators["reception_manager"],
        decision=ValidationDecision.APPROVED.value,
        source_file="autre.xlsx",
    )

    refreshed = db.get(AuditLog, deliberate.id)
    assert refreshed.checker_reference == operators["quality_manager"].employee_number
    assert refreshed.decision == "REJECTED"
    assert refreshed.source_file == "deja-renseigne.xlsx"
    # The empty field is the one that gets filled.
    assert refreshed.maker_reference == operators["maker"].employee_number
