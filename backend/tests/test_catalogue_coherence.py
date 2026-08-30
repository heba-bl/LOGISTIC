"""Excel and the database must describe the same articles.

A user test found them describing two disjoint catalogues - 2 239 articles in
the workbook, 40 unrelated ones in the database, nothing in common - which meant
no reference an operator could type in Excel was recognised by the API. These
tests exist so that cannot come back silently.

They compare the two sides through the code that builds each, rather than
against a hard-coded list: a list would be a third catalogue to keep in step.
"""

from __future__ import annotations

import io

import pytest
from openpyxl import load_workbook
from sqlalchemy import select

from app.models.catalog import Category, Part
from app.models.enums import PartSize
from app.services import excel_operations, whap_source


@pytest.fixture(scope="module")
def catalogue():
    return whap_source.load_catalogue()


@pytest.fixture(scope="module")
def workbook_codes() -> dict[str, str]:
    """`CODE -> DESIGNATION`, read back out of a freshly built workbook."""
    content = excel_operations.build_workbook()
    book = load_workbook(io.BytesIO(content), read_only=True, keep_vba=True)
    try:
        sheet = book["ARTICLES"]
        headers = {
            sheet.cell(row=excel_operations.HEADER_ROW, column=column).value: column
            for column in range(1, 20)
            if sheet.cell(row=excel_operations.HEADER_ROW, column=column).value
        }
        code_column = headers["CODE"]
        designation_column = headers["DESIGNATION"]
        return {
            str(row[code_column - 1]).strip(): str(row[designation_column - 1] or "").strip()
            for row in sheet.iter_rows(
                min_row=excel_operations.HEADER_ROW + 1, values_only=True
            )
            if row[code_column - 1]
        }
    finally:
        book.close()


def seed_catalogue(db) -> dict[str, Part]:
    """Load the catalogue into a session the way `seed.py` does.

    Deliberately mirrors the seed rather than importing it: the seed also builds
    users, warehouses and a whole history, none of which these tests need.
    """
    categories = {}
    for code in whap_source.CATEGORIES:
        category = Category(code=code, name=code.title())
        db.add(category)
        categories[code] = category
    db.flush()

    parts = {}
    for article in whap_source.load_catalogue():
        part = Part(
            reference=article.code,
            designation=article.designation,
            category_id=categories[article.category].id,
            size_class=PartSize(article.size_class),
            unit=article.unit[:10],
            safety_stock=article.minimum,
            average_daily_consumption=article.daily_consumption,
        )
        db.add(part)
        parts[article.code] = part
    db.flush()
    return parts


# ------------------------------------------------------------------- counts
def test_the_catalogue_has_the_expected_size(catalogue):
    """2 200 supplied references, plus the warehouse articles."""
    source_rows = [a for a in catalogue if a.source == "WHAP"]
    assert len(source_rows) == 2200
    assert len(catalogue) == 2239


def test_no_duplicate_reference_anywhere(catalogue, workbook_codes):
    codes = [article.code for article in catalogue]
    assert len(codes) == len(set(codes)), "doublon dans le catalogue"
    assert len(workbook_codes) == len(codes), "doublon dans le classeur"


# ------------------------------------------------------- excel vs database
def test_workbook_and_database_hold_the_same_references(db, workbook_codes):
    """The check the user test demanded: same count, nothing missing, nothing extra."""
    seed_catalogue(db)
    database = {
        reference: designation
        for reference, designation in db.execute(
            select(Part.reference, Part.designation)
        ).all()
    }

    missing = set(workbook_codes) - set(database)
    extra = set(database) - set(workbook_codes)

    assert len(workbook_codes) == 2239
    assert len(database) == 2239
    assert not missing, f"{len(missing)} references du classeur absentes de la base: {sorted(missing)[:5]}"
    assert not extra, f"{len(extra)} references en base absentes du classeur: {sorted(extra)[:5]}"


def test_designations_match_on_both_sides(db, workbook_codes):
    seed_catalogue(db)
    database = dict(db.execute(select(Part.reference, Part.designation)).all())

    divergent = [
        code
        for code, designation in workbook_codes.items()
        if designation and database.get(code) != designation
    ]
    assert not divergent, f"{len(divergent)} designations divergentes: {divergent[:5]}"


def test_every_supplied_reference_reaches_the_database(db):
    """WHAP-0001 through WHAP-2200, none dropped along the way."""
    seed_catalogue(db)
    stored = {
        reference
        for (reference,) in db.execute(
            select(Part.reference).where(Part.reference.like("WHAP-%"))
        ).all()
    }
    expected = {f"WHAP-{index:04d}" for index in range(1, 2201)}
    assert stored == expected


# ------------------------------------------------------------- addressing
def test_addresses_use_the_same_format_on_both_sides(catalogue):
    """The workbook writes `A-01-02`; the database must not invent its own."""
    for article in catalogue:
        zone, aisle, level = article.location.split("-")
        assert zone in whap_source.ZONES
        assert aisle.isdigit() and level.isdigit()
        assert len(aisle) == 2 and len(level) == 2


def test_a_reference_is_never_tied_to_a_single_address(catalogue):
    """Some references must carry an overflow address, or storage cannot split."""
    with_secondary = [a for a in catalogue if a.secondary_location]
    assert with_secondary, "aucune reference n'a d'adresse secondaire"
    assert all(a.secondary_location != a.location for a in with_secondary)
    # A meaningful share, not a token one.
    assert len(with_secondary) > len(catalogue) * 0.2


# ----------------------------------------------------------- derived fields
def test_size_class_is_one_the_model_accepts(catalogue):
    for article in catalogue:
        assert PartSize(article.size_class) in (PartSize.SMALL, PartSize.LARGE)


def test_consumption_is_positive_so_coverage_can_be_computed(catalogue):
    assert all(article.daily_consumption > 0 for article in catalogue)
