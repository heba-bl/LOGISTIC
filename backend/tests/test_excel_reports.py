"""Real Excel generation, the round trip, the nomenclature and the reports.

The rule under test, on top of the workbooks being genuinely openable: a
spreadsheet can never move the stock. Only a validated import creates business
records, and even then a validated reception creates a lot, not stock.
"""

from __future__ import annotations

import io
import zipfile
from datetime import date

import pytest
from openpyxl import load_workbook
from sqlalchemy import func, select

from app.models.enums import ImportType, PartSize, RoleName
from app.models.flow import Lot
from app.models.organization import Role, User
from app.models.vehicle import Vehicle, VehicleBomLine
from app.models.warehouse import StockMovement
from app.services import excel_service, import_service, report_service, stock_service


# --------------------------------------------------------------------- fixture
@pytest.fixture()
def operators(db, world):
    def make(matricule, username, role_name, zone):
        role = db.execute(select(Role).where(Role.name == role_name)).scalar_one_or_none()
        if role is None:
            role = Role(name=role_name, label=role_name.value.title())
            db.add(role)
            db.flush()
        user = User(
            employee_number=matricule,
            username=username,
            full_name=username.title(),
            first_name=username.split(".")[0].title(),
            last_name=username.split(".")[-1].title(),
            role_id=role.id,
            zone=zone,
            service=zone.value.title(),
        )
        db.add(user)
        db.flush()
        return user

    from app.models.enums import Zone

    return {
        "maker": make("OP-7001", "k.maker", RoleName.RECEPTIONIST, Zone.RECEPTION),
        "checker": make("RM-7002", "f.checker", RoleName.RECEPTION_MANAGER, Zone.RECEPTION),
    }


# ---------------------------------------------------------------- nomenclature
def test_generated_nomenclature_exceeds_two_thousand_references():
    from scripts.bom_data import bom_statistics, generate_bom

    lines = generate_bom()
    stats = bom_statistics(lines)

    assert stats["total"] > 2000, f"only {stats['total']} references"
    assert stats["systems"] == 15
    # References are unique and follow the SYS-0000 convention.
    references = [line["part_reference"] for line in lines]
    assert len(references) == len(set(references))
    assert all(len(reference.split("-")) == 2 for reference in references)


def test_nomenclature_is_deterministic():
    """The demonstration must be reproducible from one run to the next."""
    from scripts.bom_data import generate_bom

    first = generate_bom()
    second = generate_bom()
    assert first == second


def test_bom_can_be_stored_and_linked(db, world):
    vehicle = Vehicle(code="TEST-X", name="Test vehicle", model_year=2026)
    db.add(vehicle)
    db.flush()
    db.add(
        VehicleBomLine(
            vehicle_id=vehicle.id,
            part_reference=world["small"].reference,
            part_description=world["small"].designation,
            system_code="TST",
            system_label="Test",
            subsystem="Test",
            category="Test",
            size_class=PartSize.SMALL,
            quantity_per_vehicle=4,
            is_managed=True,
            part_id=world["small"].id,
        )
    )
    db.flush()

    line = db.execute(select(VehicleBomLine)).scalar_one()
    assert line.is_managed is True
    assert line.part is not None
    assert line.part.reference == world["small"].reference


# ------------------------------------------------------------------- workbooks
@pytest.mark.parametrize(
    "zone", ["RECEIVING", "INSPECTION", "QUALITY", "WAREHOUSE", "PRODUCTION"]
)
def test_zone_workbook_is_a_real_openable_xlsx(db, world, zone):
    content = excel_service.build_zone_workbook(db, zone)

    # An .xlsx is a zip container; if this holds, Excel can open it.
    assert zipfile.is_zipfile(io.BytesIO(content))

    workbook = load_workbook(io.BytesIO(content))
    assert workbook.sheetnames, zone
    for name in workbook.sheetnames:
        sheet = workbook[name]
        # Every sheet carries a title block and a frozen header.
        assert sheet["A1"].value
        assert sheet.freeze_panes is not None
    workbook.close()


def test_global_workbook_has_the_twelve_sheets(db, world):
    content = excel_service.build_global_workbook(db)
    workbook = load_workbook(io.BytesIO(content))

    assert workbook.sheetnames == [
        "README",
        "OPERATORS",
        "PARTS",
        "VEHICLE_BOM",
        "RECEIVING",
        "INSPECTION",
        "QUALITY",
        "RED_CAGE",
        "WAREHOUSE",
        "PRODUCTION",
        "STOCK_MOVEMENTS",
        "AUDIT",
    ]

    # The README must actually explain the file.
    readme = workbook["README"]
    text = " ".join(
        str(row[0].value) for row in readme.iter_rows(max_col=1) if row[0].value
    )
    assert "MAKER" in text and "CHECKER" in text
    assert "stock" in text.lower()

    operators = workbook["OPERATORS"]
    headers = [operators.cell(row=4, column=index).value for index in range(1, 7)]
    assert headers == ["MATRICULE", "NOM", "PRENOM", "ROLE", "ZONE", "STATUT"]

    workbook.close()


def test_entry_sheet_headers_match_the_import_format(db, world):
    """The generated template must be importable without being edited."""
    content = excel_service.build_zone_workbook(db, "RECEIVING")
    workbook = load_workbook(io.BytesIO(content))
    sheet = workbook["SAISIE"]

    headers = [
        str(sheet.cell(row=4, column=index).value or "").lower()
        for index in range(1, 7)
    ]
    expected = [name for name, _ in import_service.COLUMNS[ImportType.RECEPTION]]
    assert headers == expected
    workbook.close()


def test_workbook_summary_counts_every_sheet(db, world):
    summary = excel_service.workbook_summary(db)
    assert summary["sheet_count"] == 12
    assert {sheet["name"] for sheet in summary["sheets"]} >= {
        "OPERATORS",
        "PARTS",
        "RECEIVING",
        "AUDIT",
    }


def test_zone_table_is_json_serialisable(db, world):
    table = excel_service.zone_table(db, "RECEIVING")
    assert "columns" in table and "rows" in table
    assert isinstance(table["total_rows"], int)
    for row in table["rows"]:
        for value in row:
            assert not hasattr(value, "isoformat"), "datetimes must be serialised"


# ------------------------------------------------------------------ round trip
def test_excel_round_trip_never_moves_stock(db, world, operators):
    """Fill the generated sheet, import it, validate it - stock must not move."""
    content = excel_service.build_zone_workbook(db, "RECEIVING")
    workbook = load_workbook(io.BytesIO(content))
    sheet = workbook["SAISIE"]
    sheet.cell(row=5, column=1, value=world["small"].reference)
    sheet.cell(row=5, column=2, value=world["supplier"].code)
    sheet.cell(row=5, column=3, value=150)
    sheet.cell(row=5, column=4, value=150)
    buffer = io.BytesIO()
    workbook.save(buffer)
    workbook.close()

    lots_before = db.execute(select(func.count()).select_from(Lot)).scalar_one()
    moves_before = db.execute(select(func.count()).select_from(StockMovement)).scalar_one()
    stock_before = stock_service.get_available(db, world["small"].id)

    batch = import_service.create_import(
        db,
        import_type=ImportType.RECEPTION,
        filename="SLCC_Receiving.xlsx",
        content=buffer.getvalue(),
        maker_id=operators["maker"].id,
    )

    # The header sits under a title block: the parser must still find it.
    assert batch.row_count == 1, "the title block must not be read as data"
    assert batch.valid_row_count == 1
    assert db.execute(select(func.count()).select_from(Lot)).scalar_one() == lots_before

    import_service.approve_import(
        db, import_id=batch.id, checker_id=operators["checker"].id, comment="ok"
    )

    # A validated reception creates a lot - never stock.
    assert db.execute(select(func.count()).select_from(Lot)).scalar_one() == lots_before + 1
    assert (
        db.execute(select(func.count()).select_from(StockMovement)).scalar_one()
        == moves_before
    )
    assert stock_service.get_available(db, world["small"].id) == stock_before


def test_import_ignores_the_empty_grid_of_the_template(db, world, operators):
    """The template ships 40 pre-formatted empty rows.

    They carry borders but no value, so they must not be read as data: an
    untouched template is refused rather than silently importing 40 blank rows.
    """
    from app.core.exceptions import ValidationError

    content = excel_service.build_zone_workbook(db, "RECEIVING")

    with pytest.raises(ValidationError, match="no data row"):
        import_service.create_import(
            db,
            import_type=ImportType.RECEPTION,
            filename="template_non_rempli.xlsx",
            content=content,
            maker_id=operators["maker"].id,
        )


# --------------------------------------------------------------------- reports
def test_every_report_builds(db, world):
    for kind in report_service.BUILDERS:
        report = report_service.build_report(db, kind, "month", None, None)
        assert report.columns, kind
        assert isinstance(report.rows, list), kind
        assert report.period_label, kind


def test_report_periods_resolve():
    for period in ("today", "week", "month", "year"):
        start, end, label = report_service.resolve_period(period)
        assert start <= end
        assert label

    start, end, label = report_service.resolve_period(
        "custom", date(2026, 1, 1), date(2026, 1, 31)
    )
    assert start < end
    assert "01/01/2026" in label


def test_custom_period_requires_both_dates():
    from app.core.exceptions import ValidationError

    with pytest.raises(ValidationError):
        report_service.resolve_period("custom", date(2026, 1, 1), None)
    with pytest.raises(ValidationError):
        report_service.resolve_period("custom", date(2026, 2, 1), date(2026, 1, 1))


def test_report_exports_are_real_files(db, world):
    report = report_service.build_report(db, "stock", "month", None, None)

    xlsx = report_service.report_to_xlsx(report)
    assert zipfile.is_zipfile(io.BytesIO(xlsx))
    workbook = load_workbook(io.BytesIO(xlsx))
    assert workbook.active["A1"].value.startswith("Rapport")
    workbook.close()

    pdf = report_service.report_to_pdf(report)
    assert pdf.startswith(b"%PDF-"), "the export must be a real PDF"
    assert len(pdf) > 800


def test_unknown_report_is_refused(db):
    from app.core.exceptions import ValidationError

    with pytest.raises(ValidationError):
        report_service.build_report(db, "does-not-exist", "month", None, None)


# ------------------------------------------------------------- shared folder
def test_shared_folder_workbook_ships_filled_entry_rows(db, world):
    """The file in the shared folder is one an operator has already filled."""
    blank = excel_service.build_zone_workbook(db, "RECEIVING")
    filled = excel_service.build_zone_workbook(db, "RECEIVING", prefill=True)

    def first_data_row(content: bytes):
        workbook = load_workbook(io.BytesIO(content))
        sheet = workbook["SAISIE"]
        values = [sheet.cell(row=5, column=index).value for index in range(1, 5)]
        workbook.close()
        return values

    assert first_data_row(blank) == [None, None, None, None]

    row = first_data_row(filled)
    assert row[0] in {world["small"].reference, world["large"].reference}
    assert row[1] == world["supplier"].code
    assert isinstance(row[2], int) and row[2] > 0


def test_prefilled_entry_rows_are_importable(db, world, operators):
    """A file taken straight from the shared folder must import as it stands."""
    content = excel_service.build_zone_workbook(db, "RECEIVING", prefill=True)

    batch = import_service.create_import(
        db,
        import_type=ImportType.RECEPTION,
        filename="SLCC_Receiving.xlsx",
        content=content,
        maker_id=operators["maker"].id,
    )

    assert batch.row_count > 0
    assert batch.valid_row_count == batch.row_count, "every shipped line must be valid"
    # Reading a filled spreadsheet is still not a stock movement.
    assert (
        db.execute(select(func.count()).select_from(StockMovement)).scalar_one() == 0
    )
