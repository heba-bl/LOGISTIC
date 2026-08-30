"""Compare the catalogue in the shared workbook with the one in the database.

The two must describe the same articles. They came from the same source, but
"came from the same source" is a claim, and this checks it against the two
artefacts that actually exist: the `.xlsm` sitting in the shared folder and the
rows in `parts`.

    python scripts/check_catalogue.py

Exit code 0 when they agree, 1 when they do not.
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from openpyxl import load_workbook  # noqa: E402
from sqlalchemy import select  # noqa: E402

from app.db.session import SessionLocal  # noqa: E402
from app.models.catalog import Part  # noqa: E402
from app.services.excel_operations import HEADER_ROW, WORKBOOK_NAME  # noqa: E402

WORKBOOK = (
    Path(__file__).resolve().parents[2] / "shared-folder" / "00_FICHIER_PARTAGE" / WORKBOOK_NAME
)


def read_workbook(path: Path) -> dict[str, str]:
    """`CODE -> DESIGNATION` from the ARTICLES sheet of the real file."""
    if not path.exists():
        raise FileNotFoundError(
            f"{path} introuvable. Generer le fichier avec scripts/generate_operations_xlsm.py."
        )

    book = load_workbook(path, read_only=True, data_only=True, keep_vba=True)
    try:
        sheet = book["ARTICLES"]
        headers = {
            sheet.cell(row=HEADER_ROW, column=column).value: column
            for column in range(1, 20)
            if sheet.cell(row=HEADER_ROW, column=column).value
        }
        code_column = headers["CODE"]
        designation_column = headers["DESIGNATION"]

        catalogue: dict[str, str] = {}
        duplicates: list[str] = []
        for row in sheet.iter_rows(min_row=HEADER_ROW + 1, values_only=True):
            code = row[code_column - 1]
            if not code:
                continue
            code = str(code).strip()
            if code in catalogue:
                duplicates.append(code)
            catalogue[code] = str(row[designation_column - 1] or "").strip()
        if duplicates:
            raise ValueError(f"doublons dans le classeur: {duplicates[:5]}")
        return catalogue
    finally:
        book.close()


def read_database() -> dict[str, str]:
    """`reference -> designation` from `parts`."""
    session = SessionLocal()
    try:
        rows = session.execute(select(Part.reference, Part.designation)).all()
        catalogue: dict[str, str] = {}
        duplicates: list[str] = []
        for reference, designation in rows:
            if reference in catalogue:
                duplicates.append(reference)
            catalogue[reference] = designation
        if duplicates:
            raise ValueError(f"doublons en base: {duplicates[:5]}")
        return catalogue
    finally:
        session.close()


def compare(excel: dict[str, str], database: dict[str, str]) -> dict:
    missing = sorted(set(excel) - set(database))
    extra = sorted(set(database) - set(excel))
    shared = set(excel) & set(database)
    mismatched = sorted(
        code for code in shared if excel[code] and excel[code] != database[code]
    )
    return {
        "excel": len(excel),
        "database": len(database),
        "missing": missing,
        "extra": extra,
        "mismatched": mismatched,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", type=Path, default=WORKBOOK)
    arguments = parser.parse_args()

    print("Coherence du catalogue")
    print("=" * 62)
    try:
        excel = read_workbook(arguments.workbook)
        database = read_database()
    except (FileNotFoundError, ValueError) as error:
        print(f"  ECHEC: {error}")
        return 1

    result = compare(excel, database)

    print(f"  Catalogue Excel     : {result['excel']:>6}")
    print(f"  Catalogue DB        : {result['database']:>6}")
    print(f"  Manquantes en DB    : {len(result['missing']):>6}")
    print(f"  Supplementaires DB  : {len(result['extra']):>6}")
    print(f"  Doublons            : {0:>6}  (une exception serait levee sinon)")
    print(f"  Designations divergentes : {len(result['mismatched']):>1}")

    for label, values in (
        ("manquantes", result["missing"]),
        ("supplementaires", result["extra"]),
        ("designations", result["mismatched"]),
    ):
        if values:
            print(f"\n  {label} ({len(values)}), 10 premieres:")
            for value in values[:10]:
                print(f"     {value}")

    coherent = not (result["missing"] or result["extra"] or result["mismatched"])
    print("\n" + "=" * 62)
    print("  RESULTAT: catalogues coherents" if coherent else "  RESULTAT: DIVERGENCE")
    return 0 if coherent else 1


if __name__ == "__main__":
    raise SystemExit(main())
