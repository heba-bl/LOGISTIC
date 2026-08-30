"""Read the supplied 8x8 nomenclature and turn it into the article catalogue.

`WhAP_8x8_2200_pieces.xlsx` is the customer's own file: 2 200 rows, each a real
part code from `WHAP-0001` to `WHAP-2200`, grouped into 27 systems. Four of its
columns were left empty on purpose - internal reference, storage location,
current stock and minimum level - because they are what a logistics system is
supposed to fill in.

Two rules govern this module:

* **The source is never rewritten.** Code, system, subsystem, name, quantity per
  vehicle and unit are copied through exactly as they appear in the file. Where
  this module adds a value it adds it in a column the source left blank, and
  every row carries `source="WHAP"` so the origin stays visible.
* **The warehouse is wider than the bill of materials.** A plant also stores
  paint, adhesives, lubricant, packaging and fire extinguishers, none of which
  belong to a vehicle. Those are added here as clearly-marked demonstration
  articles with `source="DEMO"`, so nobody mistakes them for customer data.

Everything is derived deterministically from the part code, so two runs produce
the same catalogue and a demonstration can be replayed.
"""

from __future__ import annotations

import hashlib
import unicodedata
from dataclasses import dataclass, field
from functools import lru_cache
from pathlib import Path
from typing import Iterator

from openpyxl import load_workbook

#: The vendored copy, so the build never depends on somebody's Downloads folder.
SOURCE_FILE = Path(__file__).resolve().parents[3] / "data" / "source" / "WhAP_8x8_2200_pieces.xlsx"
SOURCE_SHEET = "Pièces_2200"

#: The six families the warehouse distinguishes.
CATEGORY_VEHICLE_PART = "PIECE_VEHICULE"
CATEGORY_CONSUMABLE = "CONSOMMABLE"
CATEGORY_MATERIAL = "MATIERE"
CATEGORY_VEHICLE_EQUIPMENT = "EQUIPEMENT_VEHICULE"
CATEGORY_ACCESSORY = "ACCESSOIRE"
CATEGORY_PACKAGING = "EMBALLAGE"

CATEGORIES = (
    CATEGORY_VEHICLE_PART,
    CATEGORY_CONSUMABLE,
    CATEGORY_MATERIAL,
    CATEGORY_VEHICLE_EQUIPMENT,
    CATEGORY_ACCESSORY,
    CATEGORY_PACKAGING,
)

#: Systems whose parts are equipment carried by the vehicle rather than built
#: into it. Everything not listed here is a vehicle part; the one consumable
#: system in the source is named explicitly.
EQUIPMENT_SYSTEMS = frozenset(
    {
        "Incendie",
        "Protection NBC",
        "Treuil et récupération",
        "Observation",
        "Navigation et communication",
    }
)
CONSUMABLE_SYSTEMS = frozenset({"Fixations et consommables"})

#: Short codes for the internal reference, one per source system.
SYSTEM_CODES = {
    "Moteur": "MOT",
    "Refroidissement": "REF",
    "Carburant": "CAR",
    "Transmission": "TRA",
    "Boîte de transfert": "BTR",
    "Arbres de transmission": "ARB",
    "Essieux et différentiels": "ESS",
    "Suspension": "SUS",
    "Direction": "DIR",
    "Freinage": "FRE",
    "Roues et pneus": "ROU",
    "Pneumatique": "PNE",
    "Hydraulique": "HYD",
    "Électricité": "ELE",
    "Électronique et capteurs": "ELC",
    "Éclairage": "ECL",
    "Carrosserie et coque": "CAR2",
    "Portes et trappes": "POR",
    "Habitacle": "HAB",
    "HVAC": "HVA",
    "Observation": "OBS",
    "Navigation et communication": "NAV",
    "Incendie": "INC",
    "Protection NBC": "NBC",
    "Treuil et récupération": "TRE",
    "Propulsion amphibie": "AMP",
    "Fixations et consommables": "FIX",
}

#: How critical a system is to a vehicle leaving the line. Drives the BOM's
#: criticality column and the ordering of shortage risk.
CRITICAL_SYSTEMS = frozenset(
    {"Moteur", "Transmission", "Freinage", "Direction", "Essieux et différentiels"}
)
IMPORTANT_SYSTEMS = frozenset(
    {
        "Suspension",
        "Roues et pneus",
        "Électricité",
        "Électronique et capteurs",
        "Boîte de transfert",
        "Arbres de transmission",
        "Hydraulique",
        "Refroidissement",
        "Carburant",
    }
)

#: The source carries no size, and size decides the reception rule: a SMALL part
#: accepts the configurable tolerance, a LARGE one must be counted exactly. The
#: split below is an assumption, stated rather than hidden - structural systems
#: are handled unit by unit, everything else moves in bulk.
LARGE_SYSTEMS = frozenset(
    {
        "Moteur",
        "Transmission",
        "Boîte de transfert",
        "Arbres de transmission",
        "Essieux et différentiels",
        "Suspension",
        "Roues et pneus",
        "Carrosserie et coque",
        "Portes et trappes",
        "Direction",
        "Propulsion amphibie",
        "Refroidissement",
        "Habitacle",
    }
)

#: Families that are always counted in bulk, whatever their system.
BULK_CATEGORIES = frozenset(
    {CATEGORY_CONSUMABLE, CATEGORY_MATERIAL, CATEGORY_PACKAGING, CATEGORY_ACCESSORY}
)


def _size_class_for(system: str, category: str) -> str:
    """"SMALL" or "LARGE" - the value `PartSize` expects."""
    if category in BULK_CATEGORIES:
        return "SMALL"
    if category == CATEGORY_VEHICLE_EQUIPMENT:
        return "LARGE"
    return "LARGE" if system in LARGE_SYSTEMS else "SMALL"


def _daily_consumption(seed: int, stock: int, minimum: int) -> float:
    """A plausible burn rate, so days-of-cover varies across the catalogue.

    Anchored on the minimum rather than picked freely: a reference held deep is
    one that moves, so its consumption scales with what the plant keeps of it.
    """
    base = max(minimum, 1) / (4 + seed % 9)
    return round(max(0.2, base), 1)


#: Warehouse zones the articles are addressed to.
ZONES = ("A", "B", "C", "D", "E", "F")
AISLES = tuple(range(1, 7))
LEVELS = tuple(range(1, 5))


@dataclass(frozen=True)
class Article:
    """One line of the article catalogue."""

    code: str
    reference: str
    designation: str
    system: str
    subsystem: str
    category: str
    quantity_per_vehicle: int
    unit: str
    location: str
    stock: int
    minimum: int
    supplier: str
    criticality: str
    #: Derived, because the source file carries neither: see `_size_class_for`
    #: and `_daily_consumption` for the rules and why they exist.
    size_class: str
    daily_consumption: float
    #: A second address, so a reference is never tied to a single one.
    secondary_location: str | None
    #: "WHAP" when it comes from the customer file, "DEMO" when added here.
    source: str
    #: Only bill-of-materials articles belong to a vehicle.
    in_bom: bool = True
    extras: dict = field(default_factory=dict)


def _seed(code: str) -> int:
    """A stable integer per article: same catalogue on every machine."""
    return int(hashlib.sha256(code.encode("utf-8")).hexdigest()[:8], 16)


def _ascii(value: str) -> str:
    """Strip accents for codes that must stay ASCII."""
    return "".join(
        character
        for character in unicodedata.normalize("NFKD", value)
        if not unicodedata.combining(character)
    )


def _category_for(system: str) -> str:
    if system in CONSUMABLE_SYSTEMS:
        return CATEGORY_CONSUMABLE
    if system in EQUIPMENT_SYSTEMS:
        return CATEGORY_VEHICLE_EQUIPMENT
    return CATEGORY_VEHICLE_PART


def _criticality_for(system: str) -> str:
    if system in CRITICAL_SYSTEMS:
        return "CRITIQUE"
    if system in IMPORTANT_SYSTEMS:
        return "IMPORTANT"
    return "STANDARD"


def _internal_reference(system: str, index: int) -> str:
    """`MOT-0001` style, derived from the system - never from the source code."""
    prefix = SYSTEM_CODES.get(system) or _ascii(system)[:3].upper()
    return f"{prefix}-{index:04d}"


def _location_for(seed: int) -> str:
    zone = ZONES[seed % len(ZONES)]
    aisle = AISLES[(seed // 7) % len(AISLES)]
    level = LEVELS[(seed // 53) % len(LEVELS)]
    return f"{zone}-{aisle:02d}-{level:02d}"


def _secondary_location(seed: int) -> str | None:
    """An overflow address for roughly one reference in three.

    A reference pinned to a single address cannot take a delivery once that
    address is full, which is exactly the situation the warehouse has to cope
    with. Giving a third of the catalogue a second address keeps split storage
    part of the normal case rather than an edge case.
    """
    if seed % 3 != 0:
        return None
    zone = ZONES[(seed // 11) % len(ZONES)]
    aisle = AISLES[(seed // 17) % len(AISLES)]
    level = LEVELS[(seed // 23) % len(LEVELS)]
    candidate = f"{zone}-{aisle:02d}-{level:02d}"
    return None if candidate == _location_for(seed) else candidate


def _stock_and_minimum(seed: int, criticality: str) -> tuple[int, int]:
    """Plausible holdings: critical parts are held deeper, and a slice is short.

    The point of the demonstration is that some references are in trouble, so
    roughly one in nine lands below its minimum - enough to exercise the
    shortage logic without making the plant look broken.
    """
    base = {"CRITIQUE": 40, "IMPORTANT": 90, "STANDARD": 160}[criticality]
    minimum = base // 4 + (seed % 11)
    if seed % 9 == 0:
        stock = max(0, minimum - (seed % 7) - 1)  # deliberately short
    else:
        stock = minimum + (seed % (base * 2)) + 5
    return stock, minimum


def _supplier_for(seed: int) -> str:
    suppliers = (
        "DEL", "YZK", "SUM", "VAL", "BOS", "CTL", "MAG", "FAU", "HEL", "SKF",
    )
    return suppliers[seed % len(suppliers)]


# ------------------------------------------------------------ source articles
def read_source(path: Path | None = None) -> list[Article]:
    """The 2 200 rows of the customer file, with the blank columns filled in."""
    source = path or SOURCE_FILE
    if not source.exists():
        raise FileNotFoundError(
            f"Nomenclature source introuvable: {source}. "
            "Copier WhAP_8x8_2200_pieces.xlsx dans data/source/."
        )

    workbook = load_workbook(source, read_only=True, data_only=True)
    try:
        sheet = workbook[SOURCE_SHEET]
        rows = sheet.iter_rows(min_row=2, values_only=True)

        articles: list[Article] = []
        for position, row in enumerate(rows, start=1):
            code = row[1]
            if not code:
                continue
            system = str(row[2] or "").strip()
            subsystem = str(row[3] or "").strip()
            designation = str(row[4] or "").strip()
            quantity = int(row[5] or 1)
            unit = str(row[6] or "unité").strip()

            seed = _seed(str(code))
            category = _category_for(system)
            criticality = _criticality_for(system)
            stock, minimum = _stock_and_minimum(seed, criticality)

            articles.append(
                Article(
                    code=str(code),
                    reference=_internal_reference(system, position),
                    designation=designation,
                    system=system,
                    subsystem=subsystem,
                    category=category,
                    quantity_per_vehicle=quantity,
                    unit=unit,
                    location=_location_for(seed),
                    stock=stock,
                    minimum=minimum,
                    supplier=_supplier_for(seed),
                    criticality=criticality,
                    size_class=_size_class_for(system, category),
                    daily_consumption=_daily_consumption(seed, stock, minimum),
                    secondary_location=_secondary_location(seed),
                    source="WHAP",
                    in_bom=True,
                )
            )
        return articles
    finally:
        workbook.close()


# -------------------------------------------------------- warehouse additions
#: Articles a plant stores that no vehicle contains. Marked DEMO throughout:
#: they are needed to exercise the warehouse, and they are not customer data.
NON_BOM_ARTICLES: tuple[tuple[str, str, str, str], ...] = (
    # (designation, category, unit, family label used as the "system")
    ("Peinture polyuréthane vert armée 20 L", CATEGORY_MATERIAL, "bidon", "Peinture"),
    ("Peinture antirouille primaire 20 L", CATEGORY_MATERIAL, "bidon", "Peinture"),
    ("Vernis de protection mat 10 L", CATEGORY_MATERIAL, "bidon", "Peinture"),
    ("Durcisseur pour peinture 5 L", CATEGORY_MATERIAL, "bidon", "Peinture"),
    ("Diluant de nettoyage pistolet 25 L", CATEGORY_MATERIAL, "bidon", "Peinture"),
    ("Mastic de carrosserie 3 kg", CATEGORY_MATERIAL, "pot", "Traitement"),
    ("Produit de phosphatation 20 L", CATEGORY_MATERIAL, "bidon", "Traitement"),
    ("Anticorrosion cavités 500 mL", CATEGORY_MATERIAL, "aérosol", "Traitement"),
    ("Adhésif structural bicomposant 400 mL", CATEGORY_CONSUMABLE, "cartouche", "Adhésifs"),
    ("Colle pare-brise polyuréthane 310 mL", CATEGORY_CONSUMABLE, "cartouche", "Adhésifs"),
    ("Frein-filet fort 50 mL", CATEGORY_CONSUMABLE, "flacon", "Adhésifs"),
    ("Ruban adhésif de masquage 50 mm", CATEGORY_CONSUMABLE, "rouleau", "Adhésifs"),
    ("Huile moteur 15W40 20 L", CATEGORY_CONSUMABLE, "bidon", "Lubrifiants"),
    ("Huile hydraulique ISO 46 20 L", CATEGORY_CONSUMABLE, "bidon", "Lubrifiants"),
    ("Graisse haute température 1 kg", CATEGORY_CONSUMABLE, "pot", "Lubrifiants"),
    ("Liquide de refroidissement -35 °C 20 L", CATEGORY_CONSUMABLE, "bidon", "Lubrifiants"),
    ("Liquide de frein DOT 4 1 L", CATEGORY_CONSUMABLE, "flacon", "Lubrifiants"),
    ("Dégraissant industriel 5 L", CATEGORY_CONSUMABLE, "bidon", "Nettoyage"),
    ("Chiffon non pelucheux (lot de 50)", CATEGORY_CONSUMABLE, "lot", "Nettoyage"),
    ("Absorbant granulé 20 kg", CATEGORY_CONSUMABLE, "sac", "Nettoyage"),
    ("Gants nitrile taille 9 (boîte de 100)", CATEGORY_CONSUMABLE, "boîte", "Protection"),
    ("Lunettes de protection", CATEGORY_CONSUMABLE, "unité", "Protection"),
    ("Extincteur poudre ABC 2 kg", CATEGORY_VEHICLE_EQUIPMENT, "unité", "Sécurité"),
    ("Extincteur poudre ABC 6 kg", CATEGORY_VEHICLE_EQUIPMENT, "unité", "Sécurité"),
    ("Trousse de premiers secours véhicule", CATEGORY_VEHICLE_EQUIPMENT, "unité", "Sécurité"),
    ("Triangle de signalisation", CATEGORY_VEHICLE_EQUIPMENT, "unité", "Sécurité"),
    ("Cale de roue caoutchouc", CATEGORY_VEHICLE_EQUIPMENT, "unité", "Sécurité"),
    ("Gilet haute visibilité", CATEGORY_VEHICLE_EQUIPMENT, "unité", "Sécurité"),
    ("Jeu de tapis de sol", CATEGORY_ACCESSORY, "jeu", "Accessoires"),
    ("Housse de siège renforcée", CATEGORY_ACCESSORY, "unité", "Accessoires"),
    ("Sangle d'arrimage 5 t", CATEGORY_ACCESSORY, "unité", "Accessoires"),
    ("Bâche de protection 6 x 4 m", CATEGORY_ACCESSORY, "unité", "Accessoires"),
    ("Kit d'outillage de bord", CATEGORY_ACCESSORY, "kit", "Accessoires"),
    ("Carton double cannelure 600x400x400", CATEGORY_PACKAGING, "unité", "Emballage"),
    ("Palette bois 1200x800 EUR", CATEGORY_PACKAGING, "unité", "Emballage"),
    ("Film étirable 500 mm", CATEGORY_PACKAGING, "rouleau", "Emballage"),
    ("Caisse plastique gerbable 60 L", CATEGORY_PACKAGING, "unité", "Emballage"),
    ("Coussin de calage gonflable", CATEGORY_PACKAGING, "unité", "Emballage"),
    ("Cerclage polyester 16 mm", CATEGORY_PACKAGING, "rouleau", "Emballage"),
)

#: Reference prefix per non-BOM family.
NON_BOM_PREFIX = {
    "Peinture": "PNT",
    "Traitement": "TRT",
    "Adhésifs": "ADH",
    "Lubrifiants": "LUB",
    "Nettoyage": "NET",
    "Protection": "EPI",
    "Sécurité": "SEC",
    "Accessoires": "ACC",
    "Emballage": "EMB",
}


def build_non_bom_articles() -> list[Article]:
    """Warehouse articles that are not part of any vehicle."""
    articles: list[Article] = []
    counters: dict[str, int] = {}

    for designation, category, unit, family in NON_BOM_ARTICLES:
        counters[family] = counters.get(family, 0) + 1
        prefix = NON_BOM_PREFIX[family]
        reference = f"{prefix}-{counters[family]:03d}"
        seed = _seed(reference)
        stock, minimum = _stock_and_minimum(seed, "STANDARD")

        articles.append(
            Article(
                code=reference,
                reference=reference,
                designation=designation,
                system=family,
                subsystem=family,
                category=category,
                quantity_per_vehicle=0,
                unit=unit,
                location=_location_for(seed),
                stock=stock,
                minimum=minimum,
                supplier=_supplier_for(seed),
                criticality="STANDARD",
                size_class=_size_class_for(family, category),
                daily_consumption=_daily_consumption(seed, stock, minimum),
                secondary_location=_secondary_location(seed),
                source="DEMO",
                in_bom=False,
            )
        )
    return articles


@lru_cache(maxsize=1)
def load_catalogue() -> tuple[Article, ...]:
    """The full article catalogue: the source file plus the warehouse extras."""
    return tuple(read_source() + build_non_bom_articles())


def bom_articles() -> Iterator[Article]:
    """Only the articles a vehicle is built from."""
    return (article for article in load_catalogue() if article.in_bom)


def catalogue_summary() -> dict:
    """Counts per category and per origin, for the README sheet and the tests."""
    catalogue = load_catalogue()
    per_category: dict[str, int] = {}
    per_source: dict[str, int] = {}
    for article in catalogue:
        per_category[article.category] = per_category.get(article.category, 0) + 1
        per_source[article.source] = per_source.get(article.source, 0) + 1
    return {
        "total": len(catalogue),
        "bom": sum(1 for article in catalogue if article.in_bom),
        "per_category": per_category,
        "per_source": per_source,
        "systems": len({article.system for article in catalogue if article.in_bom}),
    }


def requirement_for(vehicles: int) -> list[dict]:
    """What `vehicles` units of the 8x8 need, reference by reference."""
    if vehicles < 1:
        raise ValueError("Le nombre de véhicules doit être au moins 1")
    return [
        {
            "code": article.code,
            "reference": article.reference,
            "designation": article.designation,
            "system": article.system,
            "criticality": article.criticality,
            "per_vehicle": article.quantity_per_vehicle,
            "required": article.quantity_per_vehicle * vehicles,
            "stock": article.stock,
            "shortfall": max(0, article.quantity_per_vehicle * vehicles - article.stock),
        }
        for article in bom_articles()
    ]
