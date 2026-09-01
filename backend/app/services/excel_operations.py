"""Build `SLCC_Logistics_Operations.xlsm`, the shared shop-floor workbook.

One file, fifteen sheets, one per job. Reception fills in RECEPTION, quality
fills in QUALITE, the warehouse fills in WAREHOUSE - and none of them can
approve their own line, because the same Maker-Checker block sits at the end of
every operational sheet and the macros refuse to move a status without a second
matricule.

The workbook is generated, never hand-edited: `openpyxl` lays out the sheets and
the data, then the package is reopened as a zip so the compiled VBA project and
the ribbon can be dropped in. That is what makes it reproducible - the file can
be thrown away and rebuilt identically from the source nomenclature plus the
committed `vbaProject.bin`.

What the workbook is not is a source of truth. Every value it sends is checked
again by `import_service` on the way in; the sheet protection and the hashed
codes here stop honest mistakes, not determined ones.
"""

from __future__ import annotations

import hashlib
import io
import zipfile
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import case, func, select
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from app.core.timeutils import to_local
from app.services import vba_source, whap_source
from app.services.vba_source import (
    STATUS_APPROVED,
    STATUS_DRAFT,
    STATUS_PENDING,
    STATUS_REJECTED,
)

WORKBOOK_NAME = "SLCC_Logistics_Operations.xlsm"
VBA_ASSET = Path(__file__).resolve().parents[1] / "assets" / "vbaProject.bin"

#: The VBA reads its headers from this row; the two must agree.
HEADER_ROW = 4

#: Rows an entry sheet offers, seeded history included. The entry grid, the
#: status colouring and the dropdown ranges must cover the same cells: a
#: dropdown that stops at row 60 silently lets row 61 be typed free-hand.
GRID_ROWS = 120

#: How much history each sheet arrives with. Enough to look like a shift and to
#: practise a multi-line validation on, not so much that the blank rows below
#: are pushed off the screen.
SEED_ROWS = 40

#: Sober, industrial. Colour marks a state and nothing else.
INK = "1F2937"
HEADER_BG = "1F3864"
HEADER_FG = "FFFFFF"
BAND_BG = "F2F5F9"
MUTED = "6B7280"
BORDER = "D1D5DB"

#: Where the operator types, and where they may not. White against grey, because
#: two pale tints of the same lightness are a distinction nobody can see.
FILL_INPUT = "FFFFFF"
FILL_LOCKED = "DDE3EA"

#: Header of a column the operator fills, against one the sheet fills. The tab
#: has to answer "which columns are mine" before a single cell is clicked.
HEADER_BG_INPUT = "1F3864"
HEADER_BG_AUTO = "7C8BA1"

FILL_DRAFT = "FFF4CE"
FILL_PENDING = "FDE9C8"
FILL_APPROVED = "DCF2E3"
FILL_REJECTED = "FADBD8"

TITLE_FONT = Font(size=15, bold=True, color=INK)
SUBTITLE_FONT = Font(size=9, italic=True, color=MUTED)
HEADER_FONT = Font(size=10, bold=True, color=HEADER_FG)
HEADER_FONT_OPTIONAL = Font(size=10, color=HEADER_FG)
HEADER_FILL = PatternFill("solid", start_color=HEADER_BG)
INPUT_FILL = PatternFill("solid", start_color=FILL_INPUT)
LOCKED_FILL = PatternFill("solid", start_color=FILL_LOCKED)
HEADER_FILL_AUTO = PatternFill("solid", start_color=HEADER_BG_AUTO)
THIN = Side(style="thin", color=BORDER)
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

SYNTHETIC_NOTICE = (
    "Jeu de donnees de demonstration. Les 2 200 references proviennent du "
    "fichier fourni WhAP_8x8_2200_pieces.xlsx; les quantites, emplacements et "
    "identites sont ajoutes pour la demonstration."
)

#: The block every operational sheet ends with. `True` marks a column the
#: operator must fill before the line can be submitted - the macros read that
#: back from the bold header, so this list is the single definition.
WORKFLOW_COLUMNS: tuple[tuple[str, bool, int], ...] = (
    ("MATRICULE_OPERATEUR", True, 20),
    ("STATUT", False, 24),
    ("MATRICULE_CHECKER", False, 20),
    ("DATE_VALIDATION", False, 18),
    ("MOTIF_REJET", False, 30),
    ("ETAT_SYNC", False, 14),
    #: Below this line, the plumbing. Kept last so it can be folded away as one
    #: block: an operator has no use for it, and the macros read it by name.
    ("DATE_SOUMISSION", False, 18),
    ("ID_SYNC", False, 26),
    #: Written by SLCC when it accepts a manager's code. Proof, not a note.
    ("JETON_VALIDATION", False, 34),
)

#: Columns that exist for the machine. Folded away by default; the outline
#: button above the sheet opens them for anyone who needs to audit a line.
TECHNICAL_COLUMNS = ("DATE_SOUMISSION", "ID_SYNC", "JETON_VALIDATION")

#: Working figures the sheet needs but nobody reads: the tolerance feeds the
#: verdict shown next to it, and the time is already in the submission stamp.
QUIET_COLUMNS = ("HEURE", "TOLERANCE_AUTORISEE", "TAUX_ECHANTILLONNAGE")

#: (header, required, width) for the business part of each operational sheet.
#: A column the file can work out for itself is never marked required: the
#: operator types what only they know, the sheet computes the rest.
RECEPTION_COLUMNS = (
    ("ID_RECEPTION", False, 14),
    ("DATE", True, 12),
    ("HEURE", False, 10),
    ("REFERENCE_PIECE", True, 16),
    ("DESIGNATION", False, 34),
    ("FOURNISSEUR", True, 14),
    ("BON_LIVRAISON", True, 16),
    ("QUANTITE_ATTENDUE", True, 18),
    ("QUANTITE_RECUE", True, 16),
    ("ECART", False, 10),
    ("TOLERANCE_AUTORISEE", False, 20),
    ("RESULTAT_CONTROLE", False, 24),
    ("COMMENTAIRE", False, 30),
)

INSPECTION_COLUMNS = (
    ("ID_INSPECTION", False, 14),
    ("DATE", True, 12),
    ("HEURE", False, 10),
    ("ID_LOT", True, 16),
    ("REFERENCE_PIECE", True, 16),
    ("DESIGNATION", False, 30),
    ("QUANTITE_LOT", True, 14),
    ("TAUX_ECHANTILLONNAGE", False, 20),
    ("TAILLE_ECHANTILLON", True, 18),
    ("QUANTITE_CONFORME", True, 18),
    ("QUANTITE_NON_CONFORME", True, 22),
    ("RESULTAT", False, 16),
    ("COMMENTAIRE", False, 28),
)

QUALITY_COLUMNS = (
    ("ID_QUALITE", False, 14),
    ("DATE", True, 12),
    ("HEURE", False, 10),
    ("ID_LOT", True, 16),
    ("REFERENCE_PIECE", True, 16),
    ("DESIGNATION", False, 30),
    ("QUANTITE", True, 12),
    ("RESULTAT_INSPECTION", False, 20),
    ("DECISION", True, 16),
    ("QUANTITE_APPROUVEE", False, 18),
    ("COMMENTAIRE", True, 34),
)

RED_CAGE_COLUMNS = (
    ("ID_RED_CAGE", False, 14),
    ("DATE", True, 12),
    ("ID_LOT", True, 16),
    ("REFERENCE_PIECE", True, 16),
    ("DESIGNATION", False, 30),
    ("QUANTITE", True, 12),
    ("MOTIF", True, 30),
    ("ORIGINE", True, 16),
    ("DECISION", False, 16),
    ("JUSTIFICATION", False, 32),
    ("MATRICULE_RESPONSABLE", False, 20),
    ("DATE_DECISION", False, 16),
)

WAREHOUSE_COLUMNS = (
    ("ID_STOCKAGE", False, 14),
    ("DATE", True, 12),
    ("HEURE", False, 10),
    ("ID_LOT", True, 16),
    ("REFERENCE_PIECE", True, 16),
    ("DESIGNATION", False, 30),
    ("QUANTITE", True, 12),
    ("EMPLACEMENT", True, 14),
    ("EMPLACEMENT_PRINCIPAL", False, 20),
    ("TYPE_EMPLACEMENT", False, 18),
    ("COMMENTAIRE", False, 26),
)

PRODUCTION_COLUMNS = (
    ("ID_DEMANDE", False, 14),
    ("DATE", True, 12),
    ("STATION", True, 12),
    ("REFERENCE_PIECE", True, 16),
    ("DESIGNATION", False, 30),
    ("QUANTITE_DEMANDEE", True, 18),
    ("QUANTITE_DISPONIBLE", False, 20),
    ("ECART", False, 12),
    ("PRIORITE", False, 10),
    ("COMMENTAIRE", False, 26),
)

ISSUE_COLUMNS = (
    ("ID_SORTIE", False, 14),
    ("DATE", True, 12),
    ("HEURE", False, 10),
    ("ID_DEMANDE", True, 16),
    ("REFERENCE_PIECE", True, 16),
    ("DESIGNATION", False, 30),
    ("QUANTITE_PREPAREE", True, 18),
    ("QUANTITE_SORTIE", True, 16),
    ("EMPLACEMENT", True, 14),
    ("COMMENTAIRE", False, 26),
)

#: Fallback roster, used only when the workbook is generated without a database
#: session. With one, `operator_roster()` reads the real operators instead: the
#: matricule an operator types into Excel has to be the matricule the server
#: knows, and keeping a second list here is how the two drifted apart.
DEMO_USERS: tuple[tuple[str, str, str, str, str, str, bool], ...] = (
    # matricule, nom, prenom, role, zone, statut, droit de validation
    ("OP-1042", "Moreau", "Karim", "OPERATEUR_RECEPTION", "RECEPTION", "ACTIF", False),
    ("OP-1051", "Bouzid", "Hicham", "OPERATEUR_RECEPTION", "RECEPTION", "ACTIF", False),
    ("OP-1063", "Serrano", "Nora", "OPERATEUR_RECEPTION", "RECEPTION", "ACTIF", False),
    ("RM-004", "Chaoui", "Fatima", "RESPONSABLE_RECEPTION", "RECEPTION", "ACTIF", True),
    ("RM-005", "Idrissi", "Amal", "RESPONSABLE_RECEPTION", "RECEPTION", "INACTIF", True),
    ("QI-021", "Haddad", "Sara", "INSPECTEUR_QUALITE", "INSPECTION", "ACTIF", False),
    ("QI-022", "Mansouri", "Omar", "INSPECTEUR_QUALITE", "INSPECTION", "ACTIF", False),
    ("QM-003", "Benali", "Nadia", "RESPONSABLE_QUALITE", "QUALITE", "ACTIF", True),
    ("WH-012", "Tazi", "Youssef", "OPERATEUR_WAREHOUSE", "WAREHOUSE", "ACTIF", False),
    ("WH-019", "Morel", "David", "OPERATEUR_WAREHOUSE", "WAREHOUSE", "ACTIF", False),
    ("WM-002", "Alami", "Rachid", "RESPONSABLE_WAREHOUSE", "WAREHOUSE", "ACTIF", True),
    ("ST-012", "Dupont", "Marc", "DEMANDEUR_PRODUCTION", "PRODUCTION", "ACTIF", False),
    ("PM-001", "Ferrand", "Luc", "RESPONSABLE_PRODUCTION", "PRODUCTION", "ACTIF", True),
    ("LM-001", "Sahli", "Amine", "RESPONSABLE_LOGISTIQUE", "LOGISTIQUE", "ACTIF", True),
)

#: Demonstration codes. Printed by the generator, never written to a sheet.
DEMO_CODES = {
    "RM-004": "REC2026",
    "RM-005": "REC2026",
    "QM-003": "QUA2026",
    "WM-002": "WHS2026",
    "PM-001": "PRD2026",
    "LM-001": "LOG2026",
}

#: Salt for the stored digests. Changing it invalidates every stored code.
CODE_SALT = "SLCC-2026-OPS"

DEFAULT_API_BASE = "http://127.0.0.1:8001/api"


def code_digest(matricule: str, code: str, salt: str = CODE_SALT) -> str:
    """The digest the workbook stores, and the backend recomputes.

    Same construction on both sides - `matricule:code:salt`, SHA-256, lowercase
    hex - so a code typed into Excel and a code posted to the API are judged by
    exactly the same rule.
    """
    payload = f"{matricule.strip().upper()}:{code.strip()}:{salt}"
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


#: Database role -> the label the shop floor reads on the UTILISATEURS sheet.
ROLE_LABELS = {
    "RECEPTIONIST": "OPERATEUR_RECEPTION",
    "RECEPTION_MANAGER": "RESPONSABLE_RECEPTION",
    "QUALITY_INSPECTOR": "INSPECTEUR_QUALITE",
    "QUALITY_MANAGER": "RESPONSABLE_QUALITE",
    "WAREHOUSE_OPERATOR": "OPERATEUR_WAREHOUSE",
    "STATION_LEADER": "DEMANDEUR_PRODUCTION",
    "PRODUCTION_MANAGER": "RESPONSABLE_PRODUCTION",
    "LOGISTICS_MANAGER": "RESPONSABLE_LOGISTIQUE",
}


def operator_roster(db=None) -> list[tuple[str, str, str, str, str, str, bool]]:
    """The operators the workbook names, straight from the database.

    Returns the same seven-field shape as `DEMO_USERS`. Without a session the
    fallback list is used so the generator still runs standalone - but a file
    built that way can name people the server does not know, which is exactly
    the failure this function exists to prevent.
    """
    if db is None:
        return list(DEMO_USERS)

    from sqlalchemy import select

    from app.models.organization import Role, User

    rows = db.execute(
        select(User, Role).join(Role, User.role_id == Role.id).order_by(User.employee_number)
    ).all()
    if not rows:
        return list(DEMO_USERS)

    roster: list[tuple[str, str, str, str, str, str, bool]] = []
    for user, role in rows:
        role_name = role.name.value if hasattr(role.name, "value") else str(role.name)
        zone = user.zone.value if hasattr(user.zone, "value") else str(user.zone or "")
        roster.append(
            (
                user.employee_number,
                user.last_name or "",
                user.first_name or "",
                ROLE_LABELS.get(role_name, role_name),
                zone,
                "ACTIF" if user.is_active else "INACTIF",
                bool(role.can_validate),
            )
        )
    return roster


def validation_digests(db=None) -> list[tuple[str, str]]:
    """(matricule, digest) pairs written to CONFIGURATION.

    With a session the digests already stored on the users are reused verbatim,
    so a code typed into Excel is judged against the very row the API checks.
    """
    fallback = [
        (matricule, code_digest(matricule, code))
        for matricule, code in sorted(DEMO_CODES.items())
    ]
    if db is None:
        return fallback

    from sqlalchemy import select

    from app.models.organization import User

    rows = db.execute(
        select(User.employee_number, User.validation_code_hash)
        .where(User.validation_code_hash.is_not(None))
        .order_by(User.employee_number)
    ).all()
    return [(matricule, digest) for matricule, digest in rows] or fallback



# ------------------------------------------------------------------ layout
#: The ARTICLES sheet, declared once so formulas can find a column by name
#: instead of by a letter somebody has to keep in step.
ARTICLES_HEADERS: tuple[tuple[str, int], ...] = (
    ("CODE", 14), ("REFERENCE", 14), ("DESIGNATION", 42), ("SYSTEME", 24),
    ("SOUS_SYSTEME", 24), ("CATEGORIE", 20), ("UNITE", 10),
    ("CLASSE_TAILLE", 14), ("TOLERANCE_PCT", 14),
    ("EMPLACEMENT_PRINCIPAL", 20), ("EMPLACEMENTS_SECONDAIRES", 34),
    ("STOCK", 10), ("STOCK_TOTAL", 12), ("SEUIL_MINIMUM", 14),
    ("FOURNISSEUR", 12), ("CRITICITE", 12), ("ORIGINE", 10), ("DANS_BOM", 10),
    ("DERNIERE_SYNCHRONISATION", 22),
)


def _articles_index(name: str) -> int:
    """1-based position of an ARTICLES column."""
    for position, (header, _width) in enumerate(ARTICLES_HEADERS, start=1):
        if header == name:
            return position
    raise KeyError(f"colonne ARTICLES inconnue: {name}")


#: The ARTICLES sheet, as a range formulas can look into. Data starts under the
#: header and runs for the whole catalogue; both are known at build time.
def _articles_lookup(reference_cell: str, rows: int, column: str, fallback: str) -> str:
    """Look one ARTICLES column up by reference, quiet on an empty row."""
    index = _articles_index(column)
    last_letter = get_column_letter(len(ARTICLES_HEADERS))
    span = f"ARTICLES!$A${HEADER_ROW + 1}:${last_letter}${HEADER_ROW + rows}"
    return (
        f'=IF({reference_cell}="","",'
        f"IFERROR(VLOOKUP({reference_cell},{span},{index},FALSE),{fallback}))"
    )


def _title(sheet: Worksheet, title: str, subtitle: str, columns: int) -> None:
    sheet["A1"] = title
    sheet["A1"].font = TITLE_FONT
    sheet["A2"] = subtitle
    sheet["A2"].font = SUBTITLE_FONT
    if columns > 1:
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=columns)
        sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=columns)
    sheet.row_dimensions[1].height = 22
    sheet.row_dimensions[3].height = 6


#: What row 3 says before the workbook has ever reached SLCC. Rewritten in place
#: by StampFreshness, so the wording only has to survive being overwritten.
FRESH_ROW = 3
FRESH_IDLE = "SLCC - aucune synchronisation depuis l'ouverture du fichier"


def _freshness_banner(sheet: Worksheet, columns: int) -> None:
    """The synchronisation line, under the title of an entry sheet."""
    cell = sheet.cell(row=FRESH_ROW, column=1, value=FRESH_IDLE)
    cell.font = Font(size=9, italic=True, color=MUTED)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    if columns > 1:
        sheet.merge_cells(
            start_row=FRESH_ROW, start_column=1, end_row=FRESH_ROW, end_column=columns
        )
    # _title flattens this row to a 6px gap; an entry sheet needs to read it.
    sheet.row_dimensions[FRESH_ROW].height = 15


def _headers(
    sheet: Worksheet,
    columns: tuple[tuple[str, bool, int], ...],
    automatic: set[str] | None = None,
) -> None:
    automatic = automatic or set()
    for index, (name, required, width) in enumerate(columns, start=1):
        cell = sheet.cell(row=HEADER_ROW, column=index, value=name)
        # Bold means "required": the macro reads this back rather than keeping
        # a second list that could drift from the sheet.
        cell.font = HEADER_FONT if required else HEADER_FONT_OPTIONAL
        # A washed-out header marks a column nobody types in. Reading the tab
        # from left to right, the dark ones are the day's actual work.
        cell.fill = HEADER_FILL_AUTO if name in automatic else HEADER_FILL
        cell.border = CELL_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.row_dimensions[HEADER_ROW].height = 30
    sheet.freeze_panes = sheet.cell(row=HEADER_ROW + 1, column=1)


#: Where the choice lists live. Very hidden, like CONFIGURATION: it is plumbing,
#: not a sheet anybody should be reading or editing.
LISTS_SHEET = "LISTES"

#: Sets small and stable enough to spell out. Anything drawn from the database
#: goes through `_choice_lists` instead, so the file never disagrees with SLCC.
FIXED_CHOICES: dict[str, dict[str, tuple[str, ...]]] = {
    "QUALITE": {"DECISION": ("APPROUVE", "REJETE", "RED_CAGE")},
    "RED_CAGE": {
        "DECISION": ("LIBERE", "REBUT"),
        "ORIGINE": ("INSPECTION", "QUALITE", "PRODUCTION", "MAGASIN"),
    },
    "PRODUCTION": {"PRIORITE": ("1", "2", "3", "4", "5")},
}

#: Which lot a sheet may talk about. A lot awaiting inspection has no business
#: appearing in the storage sheet, so each zone sees only its own queue.
LOT_QUEUES: dict[str, tuple[str, ...]] = {
    "INSPECTION": ("PENDING_INSPECTION", "INSPECTION_IN_PROGRESS"),
    "QUALITE": ("QUALITY_PENDING", "INSPECTION_IN_PROGRESS"),
    "RED_CAGE": ("RED_CAGE",),
    "WAREHOUSE": ("APPROVED",),
}


#: What each list is for, in the words of the person choosing from it.
PROMPTS: dict[str, str] = {
    "REFERENCE_PIECE": "Choisissez le code de la piece (WHAP-...). La designation "
                       "s'affiche toute seule a cote.",
    "FOURNISSEUR": "Choisissez le fournisseur qui a livre.",
    "MATRICULE_OPERATEUR": "Choisissez votre matricule.",
    "ID_LOT": "Choisissez le lot. Seuls les lots concernant cette zone sont proposes.",
    "ID_DEMANDE": "Choisissez la demande de production a servir.",
    "EMPLACEMENT": "Choisissez l'adresse de stockage. La feuille EMPLACEMENTS "
                   "indique celles qui sont saturees.",
    "STATION": "Choisissez la station de montage qui demande la piece.",
    "DECISION": "Choisissez la decision. Elle sera rejouee par SLCC a la synchronisation.",
    "ORIGINE": "Choisissez l'etape qui a place le lot en quarantaine.",
    "PRIORITE": "1 est la plus urgente, 5 la moins urgente.",
}

#: The handful of things no list can know: what the delivery note says, and what
#: was actually counted on the dock.
TYPED_PROMPTS: dict[str, str] = {
    "BON_LIVRAISON": "Numero du bon de livraison papier du fournisseur.",
    "QUANTITE_ATTENDUE": "Quantite annoncee sur le bon de livraison.",
    "QUANTITE_RECUE": "Quantite reellement comptee. L'ecart se calcule tout seul.",
    "QUANTITE": "Quantite concernee par cette ligne.",
    "QUANTITE_LOT": "Taille totale du lot.",
    "TAILLE_ECHANTILLON": "Nombre de pieces controlees.",
    "QUANTITE_CONFORME": "Pieces conformes dans l'echantillon.",
    "QUANTITE_NON_CONFORME": "Pieces non conformes dans l'echantillon.",
    "QUANTITE_APPROUVEE": "Laissez vide pour approuver la totalite.",
    "QUANTITE_DEMANDEE": "Quantite demandee par la station.",
    "QUANTITE_PREPAREE": "Quantite preparee au magasin.",
    "QUANTITE_SORTIE": "Quantite reellement sortie du stock.",
    "MOTIF": "Pourquoi ce lot est en quarantaine.",
    "JUSTIFICATION": "Obligatoire: pourquoi le lot sort du Red Cage.",
    "COMMENTAIRE": "Facultatif.",
}

def _choice_lists(db) -> dict[str, list[str]]:
    """The lists that come from SLCC, read once at build time.

    Without a session the lists come back empty and the dropdowns are simply not
    offered - an empty list that silently refuses every entry would be worse than
    no list at all.
    """
    if db is None:
        return {}

    from sqlalchemy import select

    from app.models.catalog import Supplier
    from app.models.enums import LotStatus, ProductionRequestStatus
    from app.models.flow import Lot
    from app.models.organization import User
    from app.models.production import ProductionRequest, ProductionStation

    def scalars(statement) -> list[str]:
        return [str(value) for value in db.execute(statement).scalars().all() if value]

    lists: dict[str, list[str]] = {
        "FOURNISSEURS": scalars(select(Supplier.code).order_by(Supplier.code)),
        "STATIONS": scalars(
            select(ProductionStation.code).order_by(ProductionStation.code)
        ),
        "MATRICULES": scalars(
            select(User.employee_number)
            .where(User.is_active.is_(True))
            .order_by(User.employee_number)
        ),
        "DEMANDES": scalars(
            select(ProductionRequest.reference)
            .where(
                ProductionRequest.status.in_(
                    [
                        ProductionRequestStatus.SUBMITTED,
                        ProductionRequestStatus.APPROVED,
                        ProductionRequestStatus.PREPARING,
                        ProductionRequestStatus.READY,
                    ]
                )
            )
            .order_by(ProductionRequest.reference)
        ),
    }

    for sheet, states in LOT_QUEUES.items():
        lists[f"LOTS_{sheet}"] = scalars(
            select(Lot.lot_number)
            .where(Lot.status.in_([LotStatus[state] for state in states]))
            .order_by(Lot.lot_number)
        )
    return lists


def _list_ranges(lists: dict[str, list[str]]) -> dict[str, str]:
    """Where each list will sit, before a single cell of it is written.

    The dropdowns need these addresses while the sheets are being built, but the
    lists themselves belong at the end of the tab strip, out of the way. Only the
    lengths matter, so the addresses can be settled first.
    """
    return {
        name: (
            f"{LISTS_SHEET}!${get_column_letter(index)}$2:"
            f"${get_column_letter(index)}${len(values) + 1}"
        )
        for index, (name, values) in enumerate(sorted(lists.items()), start=1)
        if values
    }


def _lists_sheet(sheet: Worksheet, lists: dict[str, list[str]]) -> None:
    """Park the choice lists out of sight, in the order `_list_ranges` assumed."""
    for index, (name, values) in enumerate(sorted(lists.items()), start=1):
        sheet.cell(row=1, column=index, value=name).font = HEADER_FONT_OPTIONAL
        for offset, value in enumerate(values, start=2):
            sheet.cell(row=offset, column=index, value=value)


def _dropdowns(
    sheet: Worksheet,
    zone: str,
    columns: tuple[tuple[str, bool, int], ...],
    ranges: dict[str, str],
    catalogue_rows: int,
    location_rows: int,
    rows: int = GRID_ROWS,
) -> None:
    """Turn every column with a knowable set of answers into a list.

    A dropdown is not decoration here: it is what stops `MOT-0001` being typed
    where `WHAP-0001` is meant. The two look equally plausible on the shelf, and
    only one of them resolves.
    """
    names = [name for name, _, _ in columns]
    sources: dict[str, str] = {}

    if catalogue_rows:
        last = HEADER_ROW + catalogue_rows
        sources["REFERENCE_PIECE"] = f"ARTICLES!$A${HEADER_ROW + 1}:$A${last}"
    if location_rows:
        last = HEADER_ROW + location_rows
        sources["EMPLACEMENT"] = f"EMPLACEMENTS!$A${HEADER_ROW + 1}:$A${last}"

    for column, key in (
        ("FOURNISSEUR", "FOURNISSEURS"),
        ("STATION", "STATIONS"),
        ("MATRICULE_OPERATEUR", "MATRICULES"),
        ("ID_DEMANDE", "DEMANDES"),
        ("ID_LOT", f"LOTS_{zone}"),
    ):
        if key in ranges:
            sources[column] = ranges[key]

    for column, values in FIXED_CHOICES.get(zone, {}).items():
        sources[column] = '"' + ",".join(values) + '"'

    for column, formula in sources.items():
        if column not in names:
            continue
        letter = get_column_letter(names.index(column) + 1)
        validation = DataValidation(
            type="list",
            formula1=formula,
            allow_blank=True,
            #: Counter-intuitive, and worth stating: in this format the flag
            #: *hides* the arrow. False is what makes it appear.
            showDropDown=False,
            promptTitle="Choisissez dans la liste",
            prompt=PROMPTS.get(
                column, "Cliquez la fleche a droite de la cellule et choisissez."
            ),
            showInputMessage=True,
            errorTitle="Valeur hors liste",
            error=(
                "Choisissez une valeur dans la liste. Elle vient de SLCC: "
                "une valeur saisie a la main ne serait pas reconnue."
            ),
        )
        sheet.add_data_validation(validation)
        validation.add(f"{letter}{HEADER_ROW + 1}:{letter}{HEADER_ROW + rows}")

    # The remaining open columns hold what no list can know. They get a note
    # too, so every cell an operator lands on says something.
    for index, name in enumerate(names, start=1):
        if name in sources or name not in TYPED_PROMPTS:
            continue
        letter = get_column_letter(index)
        note = DataValidation(
            type=None, allow_blank=True, showInputMessage=True,
            promptTitle="A saisir", prompt=TYPED_PROMPTS[name],
        )
        sheet.add_data_validation(note)
        note.add(f"{letter}{HEADER_ROW + 1}:{letter}{HEADER_ROW + rows}")


def _status_rules(sheet: Worksheet, columns: tuple[tuple[str, bool, int], ...], rows: int) -> None:
    """Colour the status column by value, and offer the allowed words."""
    names = [name for name, _, _ in columns]
    if "STATUT" not in names:
        return
    index = names.index("STATUT") + 1
    letter = get_column_letter(index)
    span = f"{letter}{HEADER_ROW + 1}:{letter}{HEADER_ROW + rows}"

    from openpyxl.formatting.rule import CellIsRule

    for value, colour in (
        (STATUS_DRAFT, FILL_DRAFT),
        (STATUS_PENDING, FILL_PENDING),
        (STATUS_APPROVED, FILL_APPROVED),
        (STATUS_REJECTED, FILL_REJECTED),
    ):
        sheet.conditional_formatting.add(
            span,
            CellIsRule(
                operator="equal",
                formula=[f'"{value}"'],
                fill=PatternFill("solid", start_color=colour),
            ),
        )

    validation = DataValidation(
        type="list",
        formula1=f'"{STATUS_DRAFT},{STATUS_PENDING},{STATUS_APPROVED},{STATUS_REJECTED}"',
        allow_blank=True,
        showDropDown=False,
    )
    sheet.add_data_validation(validation)
    validation.add(span)


def _entry_grid(
    sheet: Worksheet,
    columns: tuple[tuple[str, bool, int], ...],
    rows: int = GRID_ROWS,
    seeded: list[list] | None = None,
    formulas: dict[str, str] | None = None,
) -> None:
    """Bordered rows ready for entry, with any seeded lines already filled.

    `formulas` maps a column name to a template carrying `{row}`. Those cells
    are written on every row and locked, because they are the file's answer, not
    the operator's: retyping a gap the sheet can subtract is how errors get in.
    """
    names = [name for name, _, _ in columns]
    status_index = names.index("STATUT") + 1 if "STATUT" in names else 0
    computed = {
        names.index(name) + 1: template
        for name, template in (formulas or {}).items()
        if name in names
    }
    # The workflow columns are written by the macros, never typed. Left open, an
    # operator can put VALIDE in the status themselves and walk straight past
    # the Maker-Checker rule the whole file exists to enforce.
    system_owned = {
        names.index(name) + 1
        for name in (
            "STATUT", "MATRICULE_CHECKER", "DATE_SOUMISSION",
            "DATE_VALIDATION", "MOTIF_REJET", "ID_SYNC", "ETAT_SYNC",
            "JETON_VALIDATION",
        )
        if name in names
    }
    # The line's own identifier belongs to SLCC, which assigns it on sync. Asking
    # an operator to make one up invites two people inventing the same one.
    system_owned |= {
        index
        for index, name in enumerate(names, start=1)
        if name.startswith("ID_") and name not in ("ID_LOT", "ID_DEMANDE")
    }

    for offset in range(rows):
        excel_row = HEADER_ROW + 1 + offset
        values = seeded[offset] if seeded and offset < len(seeded) else None
        for index in range(1, len(columns) + 1):
            cell = sheet.cell(row=excel_row, column=index)
            cell.border = CELL_BORDER
            cell.font = Font(size=10, color=INK)
            recorded = (
                values[index - 1]
                if values is not None and index <= len(values)
                else None
            )
            if index in computed:
                # A recorded value wins over the formula. `DATE` and `HEURE` are
                # live stamps for a line being filled in now; on a line that
                # already happened they would overwrite when it happened with
                # today, which is the one thing a record must never do.
                if recorded not in (None, ""):
                    cell.value = recorded
                else:
                    cell.value = computed[index].format(row=excel_row)
                cell.font = Font(size=10, color=MUTED, italic=True)
                # Computed cells stay locked: they are not the operator's to edit.
                cell.protection = Protection(locked=True)
                cell.fill = LOCKED_FILL
            elif index in system_owned:
                cell.protection = Protection(locked=True)
                cell.fill = LOCKED_FILL
            elif values is not None:
                # A seeded line is an operation SLCC already holds. It is shown
                # so the sheet opens on something legible, never to be edited.
                cell.protection = Protection(locked=True)
                cell.fill = LOCKED_FILL
            else:
                # Unlocked so the operator can type; `LockRow` in the macro locks
                # a line the moment it is submitted, and nothing else.
                cell.protection = Protection(locked=False)
                cell.fill = INPUT_FILL
            if values is not None and index <= len(values) and index not in computed:
                cell.value = values[index - 1]
        if status_index and (values is None or len(values) < status_index or not values[status_index - 1]):
            sheet.cell(row=excel_row, column=status_index).value = STATUS_DRAFT


def _column_letter(columns: tuple[tuple[str, bool, int], ...], name: str) -> str:
    """Where a column sits, so a formula can point at its neighbour."""
    names = [header for header, _, _ in columns] + [
        header for header, _, _ in WORKFLOW_COLUMNS
    ]
    return get_column_letter(names.index(name) + 1)


def _sheet_formulas(
    columns: tuple[tuple[str, bool, int], ...], catalogue_rows: int
) -> dict[str, str]:
    """The computed cells of one operational sheet.

    Only fields that follow from what the operator typed. Anything requiring a
    business decision - the sampling verdict, the quality decision - is left to
    the operator and re-decided by the server.
    """
    names = {header for header, _, _ in columns}
    formulas: dict[str, str] = {}

    def letter(name: str) -> str:
        return _column_letter(columns, name)

    # Nobody should be typing today's date sixty times a shift. The stamp appears
    # as soon as the line has a subject, and `TerminerMaTache` freezes it to a
    # fixed value on submission - a record must not drift when the file reopens.
    subject = next(
        (name for name in ("ID_LOT", "ID_DEMANDE", "REFERENCE_PIECE") if name in names),
        None,
    )
    if subject:
        anchor = f"${letter(subject)}{{row}}"
        if "DATE" in names:
            formulas["DATE"] = f'=IF({anchor}="","",TEXT(TODAY(),"dd/mm/yyyy"))'
        if "HEURE" in names:
            formulas["HEURE"] = f'=IF({anchor}="","",TEXT(NOW(),"hh:mm"))'

    if "REFERENCE_PIECE" in names and "DESIGNATION" in names:
        reference = f"${letter('REFERENCE_PIECE')}{{row}}"
        formulas["DESIGNATION"] = _articles_lookup(
            reference, catalogue_rows, "DESIGNATION", '"reference inconnue"'
        )

    if {"QUANTITE_ATTENDUE", "QUANTITE_RECUE", "ECART"} <= names:
        expected = f"{letter('QUANTITE_ATTENDUE')}{{row}}"
        received = f"{letter('QUANTITE_RECUE')}{{row}}"
        formulas["ECART"] = f'=IF({received}="","",{received}-{expected})'

    if "TOLERANCE_AUTORISEE" in names:
        reference = f"${letter('REFERENCE_PIECE')}{{row}}"
        formulas["TOLERANCE_AUTORISEE"] = _articles_lookup(
            reference, catalogue_rows, "TOLERANCE_PCT", "0"
        )

    if "RESULTAT_CONTROLE" in names:
        expected = f"{letter('QUANTITE_ATTENDUE')}{{row}}"
        gap = f"{letter('ECART')}{{row}}"
        tolerance = f"{letter('TOLERANCE_AUTORISEE')}{{row}}"
        # Indicative only: the server re-applies the same rule on the way in.
        formulas["RESULTAT_CONTROLE"] = (
            f'=IF({expected}="","",'
            f'IF({gap}=0,"{CONTROL_EXACT}",'
            f"IF(AND({tolerance}>0,ABS({gap})<={expected}*{tolerance}/100),"
            f'"{CONTROL_WITHIN}","{CONTROL_OUTSIDE}")))'
        )

    if {"TAILLE_ECHANTILLON", "QUANTITE_LOT", "TAUX_ECHANTILLONNAGE"} <= names:
        sample = f"{letter('TAILLE_ECHANTILLON')}{{row}}"
        lot = f"{letter('QUANTITE_LOT')}{{row}}"
        formulas["TAUX_ECHANTILLONNAGE"] = (
            f'=IF(OR({lot}="",{lot}=0,{sample}=""),"",ROUND({sample}/{lot}*100,1)&" %")'
        )

    if {"QUANTITE_DEMANDEE", "QUANTITE_DISPONIBLE", "ECART"} <= names:
        reference = f"${letter('REFERENCE_PIECE')}{{row}}"
        formulas["QUANTITE_DISPONIBLE"] = _articles_lookup(
            reference, catalogue_rows, "STOCK", "0"
        )
        wanted = f"{letter('QUANTITE_DEMANDEE')}{{row}}"
        available = f"{letter('QUANTITE_DISPONIBLE')}{{row}}"
        formulas["ECART"] = f'=IF({wanted}="","",{available}-{wanted})'

    if "EMPLACEMENT_PRINCIPAL" in names:
        reference = f"${letter('REFERENCE_PIECE')}{{row}}"
        formulas["EMPLACEMENT_PRINCIPAL"] = _articles_lookup(
            reference, catalogue_rows, "EMPLACEMENT_PRINCIPAL", '""'
        )

    return formulas


#: A line already in SLCC. The sync macro skips anything marked this way, so
#: history on the sheet stays history and is never pushed twice.
SYNC_DONE = "SYNCHRONISE"


def _fmt(moment) -> tuple[str, str]:
    """A timestamp as the two columns the sheets carry."""
    if moment is None:
        return "", ""
    return moment.strftime("%d/%m/%Y"), moment.strftime("%H:%M")


def _seed_rows(db, zone: str, limit: int = SEED_ROWS) -> list[dict[str, object]]:
    """The most recent operations of one zone, as sheet rows.

    Read straight from the business tables. Nothing here decides anything: it is
    a rendering of what happened, which is why every row arrives already
    validated - it was validated when it happened.
    """
    if db is None:
        return []

    from sqlalchemy import select
    from sqlalchemy.orm import selectinload

    from app.models.enums import LotStatus, MovementType
    from app.models.flow import Inspection, Lot, QualityValidation, Reception
    from app.models.production import ProductionRequest
    from app.models.warehouse import StockMovement

    def recent(statement, order, spread=None):
        """Newest first, then thinned so one busy part cannot fill the sheet.

        `spread` reads the value a row should vary on. Rows repeating a value
        already taken are held back and only used to top up, so a zone with
        little variety still fills rather than coming back half empty.
        """
        found = db.execute(
            statement.order_by(order.desc()).limit(limit * 12)
        ).scalars().all()
        if spread is None:
            return found[:limit]

        picked, spare, seen = [], [], set()
        for row in found:
            try:
                key = spread(row)
            except AttributeError:
                key = None
            if key is not None and key in seen:
                spare.append(row)
                continue
            seen.add(key)
            picked.append(row)
            if len(picked) == limit:
                return picked
        return (picked + spare)[:limit]

    rows: list[dict[str, object]] = []

    if zone == "RECEPTION":
        for item in recent(
            select(Reception).options(
                selectinload(Reception.lot).selectinload(Lot.part),
                selectinload(Reception.lot).selectinload(Lot.supplier),
                selectinload(Reception.received_by),
            ),
            Reception.received_at,
            spread=lambda row: row.lot.part_id,
        ):
            date, hour = _fmt(item.received_at)
            rows.append({
                "ID_RECEPTION": item.reference, "DATE": date, "HEURE": hour,
                "REFERENCE_PIECE": item.lot.part.reference,
                "FOURNISSEUR": item.lot.supplier.code if item.lot.supplier else "",
                "BON_LIVRAISON": item.delivery_note or "",
                "QUANTITE_ATTENDUE": item.quantity_expected,
                "QUANTITE_RECUE": item.quantity_received,
                "COMMENTAIRE": item.notes or "",
                "_maker": item.received_by.employee_number if item.received_by else "",
            })

    elif zone == "INSPECTION":
        for item in recent(
            select(Inspection).options(
                selectinload(Inspection.lot).selectinload(Lot.part),
                selectinload(Inspection.inspector),
            ),
            Inspection.inspected_at,
            spread=lambda row: row.lot.part_id,
        ):
            date, hour = _fmt(item.inspected_at or item.started_at)
            rows.append({
                "ID_INSPECTION": item.reference, "DATE": date, "HEURE": hour,
                "ID_LOT": item.lot.lot_number,
                "REFERENCE_PIECE": item.lot.part.reference,
                "QUANTITE_LOT": item.lot.quantity_received,
                "TAILLE_ECHANTILLON": item.sample_size,
                "QUANTITE_CONFORME": item.sample_size - item.defects_found,
                "QUANTITE_NON_CONFORME": item.defects_found,
                "COMMENTAIRE": item.observations or "",
                "_maker": item.inspector.employee_number if item.inspector else "",
            })

    elif zone == "QUALITE":
        for item in recent(
            select(QualityValidation).options(
                selectinload(QualityValidation.lot).selectinload(Lot.part),
                selectinload(QualityValidation.decided_by),
            ),
            QualityValidation.decided_at,
            spread=lambda row: (row.lot.part_id, row.decision),
        ):
            date, hour = _fmt(item.decided_at)
            decision = getattr(item.decision, "value", str(item.decision)).upper()
            rows.append({
                "ID_QUALITE": f"QV-{item.id:05d}", "DATE": date, "HEURE": hour,
                "ID_LOT": item.lot.lot_number,
                "REFERENCE_PIECE": item.lot.part.reference,
                "QUANTITE": item.lot.quantity_received,
                "DECISION": DECISION_WORDS.get(decision, decision),
                "QUANTITE_APPROUVEE": item.quantity_approved or "",
                "COMMENTAIRE": item.justification or "",
                "_maker": item.decided_by.employee_number if item.decided_by else "",
            })

    elif zone == "RED_CAGE":
        for item in recent(
            select(Lot)
            .where(Lot.status == LotStatus.RED_CAGE)
            .options(selectinload(Lot.part)),
            Lot.received_at,
            spread=lambda row: row.part_id,
        ):
            date, _ = _fmt(item.received_at)
            rows.append({
                "ID_RED_CAGE": f"RC-{item.id:05d}", "DATE": date,
                "ID_LOT": item.lot_number,
                "REFERENCE_PIECE": item.part.reference,
                "QUANTITE": item.quantity_received,
                "MOTIF": item.blocked_reason or "non conformite constatee",
                "ORIGINE": "INSPECTION",
            })

    elif zone in ("WAREHOUSE", "SORTIES"):
        wanted = MovementType.IN if zone == "WAREHOUSE" else MovementType.OUT
        for item in recent(
            select(StockMovement)
            .where(StockMovement.movement_type == wanted)
            .options(
                selectinload(StockMovement.part),
                selectinload(StockMovement.lot),
                selectinload(StockMovement.location),
                selectinload(StockMovement.production_request),
                selectinload(StockMovement.actor),
            ),
            StockMovement.occurred_at,
            spread=lambda row: (row.part_id, row.location_id),
        ):
            date, hour = _fmt(item.occurred_at)
            common = {
                "DATE": date, "HEURE": hour,
                "REFERENCE_PIECE": item.part.reference,
                "EMPLACEMENT": item.location.code if item.location else "",
                "COMMENTAIRE": item.reason or "",
                "_maker": item.actor.employee_number if item.actor else "",
            }
            if zone == "WAREHOUSE":
                rows.append({
                    **common, "ID_STOCKAGE": item.reference,
                    "ID_LOT": item.lot.lot_number if item.lot else "",
                    "QUANTITE": item.quantity,
                })
            else:
                request = item.production_request
                rows.append({
                    **common, "ID_SORTIE": item.reference,
                    "ID_DEMANDE": request.reference if request else "",
                    "QUANTITE_PREPAREE": item.quantity,
                    "QUANTITE_SORTIE": item.quantity,
                })

    elif zone == "PRODUCTION":
        for item in recent(
            select(ProductionRequest).options(
                selectinload(ProductionRequest.part),
                selectinload(ProductionRequest.station),
                selectinload(ProductionRequest.requested_by),
            ),
            ProductionRequest.created_on,
            spread=lambda row: (row.part_id, row.station_id),
        ):
            date, _ = _fmt(item.submitted_at or item.created_on)
            rows.append({
                "ID_DEMANDE": item.reference, "DATE": date,
                "STATION": item.station.code,
                "REFERENCE_PIECE": item.part.reference,
                "QUANTITE_DEMANDEE": item.quantity_requested,
                "PRIORITE": item.priority,
                "COMMENTAIRE": item.notes or "",
                "_maker": (
                    item.requested_by.employee_number if item.requested_by else ""
                ),
            })

    return rows


#: The decisions as the sheet words them, against how the database stores them.
DECISION_WORDS = {
    "APPROVED": "APPROUVE", "REJECTED": "REJETE", "RED_CAGE": "RED_CAGE",
    "SCRAPPED": "REBUT",
}


#: How many outstanding items a sheet opens with. Enough to practise a bulk
#: validation on; small enough that the file is not a to-do list of a hundred.
PENDING_ROWS = 10


def _pending_rows(db, zone: str, limit: int = PENDING_ROWS) -> list[dict[str, object]]:
    """Work the plant still owes, as draft rows an operator can complete.

    Read from the same tables as the history: a lot sitting in
    PENDING_INSPECTION really is waiting for somebody. Nothing is fabricated,
    and the columns the operator has to measure are deliberately left empty.
    """
    if db is None:
        return []

    from sqlalchemy import select
    from sqlalchemy.orm import selectinload

    from app.models.enums import LotStatus, ProductionRequestStatus
    from app.models.flow import Lot
    from app.models.production import ProductionRequest

    def lots(*states):
        return db.execute(
            select(Lot)
            .where(Lot.status.in_(states))
            .options(selectinload(Lot.part), selectinload(Lot.supplier))
            .order_by(Lot.received_at.desc())
            .limit(limit)
        ).scalars().all()

    rows: list[dict[str, object]] = []

    if zone == "INSPECTION":
        for lot in lots(LotStatus.PENDING_INSPECTION, LotStatus.INSPECTION_IN_PROGRESS):
            rows.append({
                "ID_LOT": lot.lot_number,
                "REFERENCE_PIECE": lot.part.reference,
                "QUANTITE_LOT": lot.quantity_received,
            })

    elif zone == "QUALITE":
        for lot in lots(LotStatus.QUALITY_PENDING):
            rows.append({
                "ID_LOT": lot.lot_number,
                "REFERENCE_PIECE": lot.part.reference,
                "QUANTITE": lot.quantity_received,
            })

    elif zone == "WAREHOUSE":
        # Approved and not yet stored: the magasin has these on the dock.
        for lot in lots(LotStatus.APPROVED):
            rows.append({
                "ID_LOT": lot.lot_number,
                "REFERENCE_PIECE": lot.part.reference,
                "QUANTITE": lot.quantity_received,
            })

    elif zone == "SORTIES":
        for request in db.execute(
            select(ProductionRequest)
            .where(
                ProductionRequest.status.in_(
                    (
                        ProductionRequestStatus.APPROVED,
                        ProductionRequestStatus.PREPARING,
                        ProductionRequestStatus.READY,
                    )
                )
            )
            .options(
                selectinload(ProductionRequest.part),
                selectinload(ProductionRequest.station),
            )
            .order_by(ProductionRequest.created_on.desc())
            .limit(limit)
        ).scalars().all():
            rows.append({
                "ID_DEMANDE": request.reference,
                "REFERENCE_PIECE": request.part.reference,
                "QUANTITE_PREPAREE": request.quantity_requested,
            })

    return rows


def _seeded_grid(
    db, zone: str, columns: tuple[tuple[str, bool, int], ...]
) -> list[list]:
    """Lay the zone's history out in column order, workflow columns included."""
    names = [name for name, _, _ in columns]
    grid: list[list] = []
    for record in _seed_rows(db, zone):
        maker = str(record.pop("_maker", "") or "")
        record["MATRICULE_OPERATEUR"] = maker
        record["STATUT"] = STATUS_APPROVED
        record["ETAT_SYNC"] = SYNC_DONE
        record["ID_SYNC"] = f"{zone}-HISTORIQUE-{len(grid) + 1:03d}"
        grid.append([record.get(name, "") for name in names])

    # Outstanding work, below the history and left for the operator. BROUILLON
    # and not EN ATTENTE: the line is not finished, so it is not yet anybody's
    # to validate - and no ID_SYNC, because nothing has been sent.
    for record in _pending_rows(db, zone):
        record["STATUT"] = STATUS_DRAFT
        grid.append([record.get(name, "") for name in names])

    return grid


def _fold_machinery(
    sheet: Worksheet, columns: tuple[tuple[str, bool, int], ...]
) -> None:
    """Hide what the operator has no use for, without removing it.

    Hidden is not gone: `ColumnIndex` walks the header row by name and finds a
    folded column exactly as it finds any other, and the sync sends the whole
    row. This changes what the sheet looks like, and nothing about what it does.
    """
    names = [name for name, _, _ in columns]

    for name in QUIET_COLUMNS:
        if name in names:
            letter = get_column_letter(names.index(name) + 1)
            sheet.column_dimensions[letter].hidden = True

    present = [name for name in TECHNICAL_COLUMNS if name in names]
    if not present:
        return
    positions = sorted(names.index(name) + 1 for name in present)
    # They were laid out last and contiguous, so one outline covers them.
    if positions == list(range(positions[0], positions[0] + len(positions))):
        sheet.column_dimensions.group(
            get_column_letter(positions[0]),
            get_column_letter(positions[-1]),
            outline_level=1,
            hidden=True,
        )
    else:
        for index in positions:
            sheet.column_dimensions[get_column_letter(index)].hidden = True


def _automatic_columns(
    columns: tuple[tuple[str, bool, int], ...], formulas: dict[str, str] | None
) -> set[str]:
    """Columns the operator never fills: computed, or owned by the macros."""
    names = [name for name, _, _ in columns]
    automatic = set(formulas or {})
    automatic |= {
        name
        for name in names
        if name.startswith("ID_") and name not in ("ID_LOT", "ID_DEMANDE")
    }
    automatic |= {
        "STATUT", "MATRICULE_CHECKER", "DATE_SOUMISSION", "DATE_VALIDATION",
        "MOTIF_REJET", "ID_SYNC", "ETAT_SYNC", "JETON_VALIDATION",
    }
    return automatic & set(names)


def _operational_sheet(
    sheet: Worksheet,
    title: str,
    subtitle: str,
    business: tuple[tuple[str, bool, int], ...],
    seeded: list[list] | None = None,
    formulas: dict[str, str] | None = None,
    ranges: dict[str, str] | None = None,
    catalogue_rows: int = 0,
    location_rows: int = 0,
    db=None,
) -> tuple[tuple[str, bool, int], ...]:
    columns = business + WORKFLOW_COLUMNS
    if seeded is None and db is not None:
        seeded = _seeded_grid(db, title, columns)
    _title(sheet, title, subtitle, len(columns))
    _freshness_banner(sheet, len(columns))
    _headers(sheet, columns, automatic=_automatic_columns(columns, formulas))
    _entry_grid(sheet, columns, seeded=seeded, formulas=formulas)
    _fold_machinery(sheet, columns)
    _status_rules(sheet, columns, rows=GRID_ROWS)
    _dropdowns(
        sheet,
        title,
        columns,
        ranges or {},
        catalogue_rows=catalogue_rows,
        location_rows=location_rows,
    )
    return columns


def _reference_sheet(
    sheet: Worksheet,
    title: str,
    subtitle: str,
    headers: tuple[tuple[str, int], ...],
    rows: list[list],
    banded: bool = True,
) -> None:
    """A read-only reference table: catalogue, BOM, locations, history."""
    _title(sheet, title, subtitle, len(headers))
    for index, (name, width) in enumerate(headers, start=1):
        cell = sheet.cell(row=HEADER_ROW, column=index, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = CELL_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.row_dimensions[HEADER_ROW].height = 28
    sheet.freeze_panes = sheet.cell(row=HEADER_ROW + 1, column=1)

    band = PatternFill("solid", start_color=BAND_BG)
    for offset, row in enumerate(rows):
        excel_row = HEADER_ROW + 1 + offset
        for index, value in enumerate(row, start=1):
            cell = sheet.cell(row=excel_row, column=index, value=value)
            cell.font = Font(size=9, color=INK)
            cell.border = CELL_BORDER
            if banded and offset % 2:
                cell.fill = band
            if isinstance(value, int) and index > 1:
                cell.alignment = Alignment(horizontal="right")

    if rows:
        sheet.auto_filter.ref = (
            f"A{HEADER_ROW}:{get_column_letter(len(headers))}{HEADER_ROW + len(rows)}"
        )


# ------------------------------------------------------------------- sheets
def _home_sheet(sheet: Worksheet, summary: dict, generated_at: datetime) -> None:
    """What to do, in the order an operator does it."""
    lines: list[tuple[str, str]] = [
        ("SLCC - FICHIER LOGISTIQUE OPERATIONNEL", "title"),
        (SYNTHETIC_NOTICE, "muted"),
        ("", ""),
        ("A QUOI SERT CE FICHIER", "section"),
        ("Chaque zone travaille dans sa propre feuille. Les donnees saisies ici", ""),
        ("ne deviennent operationnelles dans SLCC qu'une fois validees par un", ""),
        ("responsable different de la personne qui a saisi.", ""),
        ("", ""),
        ("VOTRE JOURNEE, EN QUATRE GESTES", "section"),
        ("1. Ouvrez la feuille de votre zone (RECEPTION, INSPECTION, ...).", ""),
        ("2. Remplissez une ligne. Les en-tetes en gras sont obligatoires.", ""),
        ("3. Placez le curseur sur votre ligne, puis onglet SLCC >", ""),
        ("   TERMINER MA TACHE. La ligne passe EN ATTENTE DE VALIDATION", ""),
        ("   et se verrouille: vous ne pouvez plus la modifier.", ""),
        ("4. Un responsable de votre zone la valide ou la rejette.", ""),
        ("", ""),
        ("POUR LE RESPONSABLE", "section"),
        ("Selectionnez une ligne EN ATTENTE DE VALIDATION, puis onglet SLCC >", ""),
        ("VALIDER (ou REJETER). Le fichier demande votre matricule et votre", ""),
        ("code. Il refuse la validation si vous etes la personne qui a saisi,", ""),
        ("si votre zone ne correspond pas, ou si le code est faux.", ""),
        ("Un rejet exige un motif et renvoie la ligne a l'operateur.", ""),
        ("", ""),
        ("ENVOYER VERS SLCC", "section"),
        ("Onglet SLCC > ENREGISTRER & SYNCHRONISER.", ""),
        ("Seules les lignes VALIDE partent. Une ligne BROUILLON ou EN ATTENTE", ""),
        ("DE VALIDATION n'est jamais consideree comme une donnee operationnelle.", ""),
        ("", ""),
        ("LA REGLE DU STOCK - NON NEGOCIABLE", "section"),
        ("Une reception, une inspection ou une validation qualite ne modifient", ""),
        ("JAMAIS le stock. Le stock augmente uniquement a la confirmation de", ""),
        ("stockage, et diminue uniquement a la sortie confirmee.", ""),
        ("Ce fichier ne peut pas contourner cette regle: c'est le serveur qui", ""),
        ("l'applique.", ""),
        ("", ""),
        ("CE QUE CONTIENT LE FICHIER", "section"),
        (f"Articles ................ {summary['total']:>6}", "mono"),
        (f"  dont nomenclature ..... {summary['bom']:>6}  (fichier WhAP fourni)", "mono"),
        (
            f"  dont hors nomenclature  {summary['total'] - summary['bom']:>6}"
            "  (peinture, EPI, emballage...)",
            "mono",
        ),
        (f"Systemes ................ {summary['systems']:>6}", "mono"),
        ("", ""),
        (f"Genere le {generated_at.astimezone().strftime('%d/%m/%Y a %H:%M')}", "muted"),
    ]

    sheet.column_dimensions["A"].width = 78
    for index, (text, kind) in enumerate(lines, start=1):
        cell = sheet.cell(row=index, column=1, value=text)
        if kind == "title":
            cell.font = TITLE_FONT
        elif kind == "section":
            cell.font = Font(size=11, bold=True, color=HEADER_BG)
        elif kind == "muted":
            cell.font = SUBTITLE_FONT
        elif kind == "mono":
            cell.font = Font(size=10, name="Consolas", color=INK)
        else:
            cell.font = Font(size=10, color=INK)
    sheet.sheet_view.showGridLines = False


def _users_sheet(sheet: Worksheet, db=None) -> None:
    headers = (
        ("MATRICULE", 14), ("NOM", 16), ("PRENOM", 14), ("ROLE", 22),
        ("ZONE", 14), ("STATUT", 12), ("DROIT_VALIDATION", 18),
    )
    rows = [
        [matricule, nom, prenom, role, zone, statut, "OUI" if can else "NON"]
        for matricule, nom, prenom, role, zone, statut, can in operator_roster(db)
    ]
    _reference_sheet(
        sheet,
        "UTILISATEURS",
        "Un operateur n'est jamais anonyme: le matricule identifie chaque action. "
        "Les codes de validation ne figurent pas sur cette feuille.",
        headers,
        rows,
    )


#: What the ARTICLES sheet says when it was built without a database.
NOT_SYNCED = "non synchronise"

#: The three answers the bill of materials can give.
COVERAGE_OK = "\u2713 COUVERT"
COVERAGE_LOW = "\u26a0 STOCK FAIBLE"
COVERAGE_SHORT = "\u2715 INSUFFISANT"

#: What the reception sheet says about a counted quantity. Indicative: the
#: server applies the same rule again and its answer is the one that counts.
CONTROL_EXACT = "\u2713 CONFORME"
CONTROL_WITHIN = "\u2713 DANS LA TOLERANCE"
CONTROL_OUTSIDE = "\u2715 HORS TOLERANCE"


def live_stock(db) -> dict[str, dict]:
    """The real balance per reference, plus where it sits.

    Reads, never computes. `Stock.quantity_available` is owned by
    `stock_service` and is the only figure that counts; the per-address
    breakdown comes from the stored lots, which is where the quantity physically
    is. Nothing here adds a second way of working out what the stock is.

    Returns `reference -> {available, located, primary, secondary}` where
    `located` is the sum across addresses. The two totals must agree; the sheet
    shows both so a divergence is visible rather than silent.
    """
    from app.models.catalog import Part
    from app.models.enums import LocationRole, MovementType
    from app.models.warehouse import PartLocation, Stock, StockMovement, WarehouseLocation

    balances = {
        reference: int(quantity or 0)
        for reference, quantity in db.execute(
            select(Part.reference, Stock.quantity_available).join(
                Stock, Stock.part_id == Part.id
            )
        ).all()
    }

    # Where the quantity actually lies, address by address.
    #
    # Read from the movement ledger, not from `Lot.location_id`: a lot split
    # across two shelves keeps one "main" address on the lot row, so the lot
    # cannot tell you the split. The ledger records one movement per address,
    # and its sum per address matches `WarehouseLocation.occupied` exactly.
    placement: dict[str, list[tuple[str, int]]] = {}
    signed = func.sum(
        case(
            (StockMovement.movement_type == MovementType.IN, StockMovement.quantity),
            else_=-StockMovement.quantity,
        )
    )
    rows = db.execute(
        select(Part.reference, WarehouseLocation.code, signed)
        .select_from(StockMovement)
        .join(Part, Part.id == StockMovement.part_id)
        .join(WarehouseLocation, WarehouseLocation.id == StockMovement.location_id)
        .group_by(Part.reference, WarehouseLocation.code)
        .having(signed > 0)
        .order_by(Part.reference, WarehouseLocation.code)
    ).all()
    for reference, code, quantity in rows:
        placement.setdefault(reference, []).append((code, int(quantity or 0)))

    # The declared primary address, which exists whether or not stock sits there.
    primaries = {
        reference: code
        for reference, code in db.execute(
            select(Part.reference, WarehouseLocation.code)
            .select_from(PartLocation)
            .join(Part, Part.id == PartLocation.part_id)
            .join(WarehouseLocation, WarehouseLocation.id == PartLocation.location_id)
            .where(PartLocation.role == LocationRole.PRIMARY)
        ).all()
    }

    catalogue: dict[str, dict] = {}
    for reference in set(balances) | set(placement) | set(primaries):
        addresses = placement.get(reference, [])
        primary = primaries.get(reference)
        catalogue[reference] = {
            "available": balances.get(reference, 0),
            "located": sum(quantity for _, quantity in addresses),
            "primary": primary,
            #: Every address holding this reference, with its quantity.
            "placed": addresses,
            # Everywhere the stock is, other than the declared primary.
            "secondary": [
                (code, quantity) for code, quantity in addresses if code != primary
            ],
        }
    return catalogue


def article_tolerances(db) -> dict[str, float]:
    """The reception tolerance that applies to each reference.

    Resolved by `reception_service`, which owns the rule: per-part override
    first, then the setting for the size class. Writing the resolved number into
    ARTICLES lets the sheet show it without holding a second copy of the rule.
    """
    from app.models.catalog import Part
    from app.services import reception_service

    tolerances: dict[str, float] = {}
    for part in db.execute(select(Part)).scalars():
        rule = reception_service.resolve_tolerance(db, part, 100)
        tolerances[part.reference] = round(float(rule.percent), 2)
    return tolerances


def _articles_sheet(
    sheet: Worksheet,
    catalogue,
    stock: dict[str, dict] | None,
    generated_at: datetime,
    tolerances: dict[str, float] | None = None,
) -> None:
    """The catalogue, with the stock read back from SLCC.

    STOCK is not a figure the workbook owns. It is the balance the database
    holds at the moment the file was built, and it is stamped with that moment
    so nobody reads a week-old number as today's. Excel remains where the work
    is entered; the database remains the only thing that says what is in stock.
    """
    headers = ARTICLES_HEADERS

    stamp = to_local(generated_at).strftime("%d/%m/%Y %H:%M") if stock is not None else NOT_SYNCED
    tolerances = tolerances or {}

    rows = []
    for article in catalogue:
        live = stock.get(article.code) if stock is not None else None
        if live is None:
            # No database to read, or a reference the warehouse has never held.
            available = 0 if stock is not None else article.stock
            located = available
            primary = article.location
            secondary = ""
        else:
            available = live["available"]
            located = live["located"]
            primary_code = live["primary"] or article.location
            # Show the quantity on the primary too, so the addresses visibly
            # add up to STOCK rather than leaving the reader to infer it.
            primary_quantity = next(
                (quantity for code, quantity in live["placed"] if code == primary_code),
                None,
            )
            primary = (
                f"{primary_code} : {primary_quantity}"
                if primary_quantity is not None
                else primary_code
            )
            secondary = " · ".join(
                f"{code} : {quantity}" for code, quantity in live["secondary"]
            )

        rows.append([
            article.code, article.reference, article.designation, article.system,
            article.subsystem, article.category, article.unit,
            article.size_class, tolerances.get(article.code, 0.0),
            primary, secondary,
            available, located, article.minimum,
            article.supplier, article.criticality,
            article.source, "OUI" if article.in_bom else "NON",
            stamp,
        ])

    _reference_sheet(
        sheet,
        "ARTICLES",
        "STOCK et STOCK_TOTAL sont synchronises depuis SLCC: la base est la seule "
        "source de verite du stock, ce classeur en affiche une photo datee. "
        "STOCK = solde disponible; STOCK_TOTAL = somme reellement localisee sur "
        "les emplacements. ORIGINE=WHAP: reference du fichier fourni, inchangee.",
        headers,
        rows,
    )


def _bom_sheet(
    sheet: Worksheet, catalogue, stock: dict[str, dict] | None, generated_at: datetime
) -> None:
    """The bill of materials, answering "can we build them?".

    Three columns do the work and two of them are formulas, so changing the
    vehicle count in D3 re-answers the question without regenerating anything:

        QUANTITE_REQUISE   = quantite par vehicule x D3
        ECART              = stock actuel - quantite requise
        STATUT_COUVERTURE  = the verdict, in words

    STOCK_ACTUEL is the same live balance the ARTICLES sheet shows - one file
    cannot hold two different answers to "how much is there". It is read, never
    written: the bill of materials is a control, and controls do not move stock.
    """
    bom = [article for article in catalogue if article.in_bom]

    sheet["A1"] = "BOM_VEHICULE"
    sheet["A1"].font = TITLE_FONT
    sheet["A2"] = (
        "Nomenclature du 8x8. Changez le nombre de vehicules en D3: quantite "
        "requise, ecart et statut se recalculent. STOCK_ACTUEL est synchronise "
        "depuis SLCC, comme sur la feuille ARTICLES."
    )
    sheet["A2"].font = SUBTITLE_FONT

    sheet["C3"] = "VEHICULES A PRODUIRE"
    sheet["C3"].font = Font(size=10, bold=True, color=HEADER_BG)
    sheet["D3"] = 5
    sheet["D3"].font = Font(size=12, bold=True, color=INK)
    sheet["D3"].fill = PatternFill("solid", start_color="FFF4CE")
    sheet["D3"].border = CELL_BORDER
    sheet["D3"].alignment = Alignment(horizontal="center")
    sheet["D3"].protection = Protection(locked=False)

    headers = (
        ("VEHICULE", 12), ("CODE_PIECE", 14), ("REFERENCE", 14), ("DESIGNATION", 40),
        ("SYSTEME", 24), ("SOUS_SYSTEME", 22), ("CATEGORIE", 18),
        ("QUANTITE_PAR_VEHICULE", 20), ("QUANTITE_REQUISE", 18), ("UNITE", 10),
        ("STOCK_ACTUEL", 14), ("ECART", 12), ("STATUT_COUVERTURE", 22),
        ("FOURNISSEUR", 12), ("CRITICITE", 12), ("DERNIERE_SYNCHRONISATION", 22),
    )
    for index, (name, width) in enumerate(headers, start=1):
        cell = sheet.cell(row=HEADER_ROW, column=index, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = CELL_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.row_dimensions[HEADER_ROW].height = 30
    sheet.freeze_panes = sheet.cell(row=HEADER_ROW + 1, column=1)

    stamp = to_local(generated_at).strftime("%d/%m/%Y %H:%M") if stock is not None else NOT_SYNCED

    band = PatternFill("solid", start_color=BAND_BG)
    for offset, article in enumerate(bom):
        row = HEADER_ROW + 1 + offset
        live = stock.get(article.code) if stock is not None else None
        # The whole balance, however many addresses hold it - never the primary
        # address alone.
        available = live["available"] if live is not None else (
            0 if stock is not None else article.stock
        )

        values = [
            "WHAP 8x8", article.code, article.reference, article.designation,
            article.system, article.subsystem, article.category,
            article.quantity_per_vehicle,
            # Formulas, so the sheet answers "and for twelve vehicles?" itself.
            f"=H{row}*$D$3",
            article.unit,
            available,
            f"=K{row}-I{row}",
            # Below the requirement is short; within a fifth of it is thin.
            f'=IF(K{row}<I{row},"{COVERAGE_SHORT}",'
            f'IF(K{row}<I{row}*1.2,"{COVERAGE_LOW}","{COVERAGE_OK}"))',
            article.supplier, article.criticality, stamp,
        ]
        for index, value in enumerate(values, start=1):
            cell = sheet.cell(row=row, column=index, value=value)
            cell.font = Font(size=9, color=INK)
            cell.border = CELL_BORDER
            if offset % 2:
                cell.fill = band

    last_row = HEADER_ROW + len(bom)
    sheet.auto_filter.ref = f"A{HEADER_ROW}:P{last_row}"

    # A quiet wash on the verdict column, so a short reference is visible while
    # scrolling without the sheet turning into a traffic light.
    from openpyxl.formatting.rule import CellIsRule

    span = f"M{HEADER_ROW + 1}:M{last_row}"
    for verdict, colour in (
        (COVERAGE_OK, FILL_APPROVED),
        (COVERAGE_LOW, FILL_PENDING),
        (COVERAGE_SHORT, FILL_REJECTED),
    ):
        sheet.conditional_formatting.add(
            span,
            CellIsRule(
                operator="equal",
                formula=[f'"{verdict}"'],
                fill=PatternFill("solid", start_color=colour),
            ),
        )


def _locations_sheet(sheet: Worksheet, catalogue) -> int:
    """Addresses and their occupancy - a reference may sit in several."""
    per_location: dict[str, list] = {}
    for article in catalogue:
        per_location.setdefault(article.location, []).append(article)

    headers = (
        ("EMPLACEMENT", 14), ("ZONE", 8), ("ALLEE", 8), ("NIVEAU", 8),
        ("TYPE_STOCKAGE", 16), ("CAPACITE", 12), ("OCCUPATION", 12),
        ("DISPONIBLE", 12), ("REFERENCES", 12), ("STATUT", 14),
    )
    rows = []
    for code in sorted(per_location):
        articles = per_location[code]
        zone, aisle, level = code.split("-")
        occupied = sum(article.stock for article in articles)
        capacity = max(2000, ((occupied // 500) + 2) * 500)
        storage = "Rack palette" if zone in ("C", "E", "F") else "Casier"
        ratio = occupied / capacity * 100 if capacity else 0
        status = "SATURE" if ratio >= 90 else "TENDU" if ratio >= 75 else "DISPONIBLE"
        rows.append(
            [code, zone, int(aisle), int(level), storage, capacity, occupied,
             capacity - occupied, len(articles), status]
        )

    _reference_sheet(
        sheet,
        "EMPLACEMENTS",
        "Une reference peut occuper plusieurs adresses et un lot peut etre eclate. "
        "Si une adresse est saturee, choisissez une adresse DISPONIBLE.",
        headers,
        rows,
    )
    return len(rows)


def _history_sheet(sheet: Worksheet) -> None:
    headers = (
        ("ID_ENREGISTREMENT", 18), ("ACTION", 22), ("MATRICULE_OPERATEUR", 20),
        ("MATRICULE_CHECKER", 20), ("DATE", 12), ("HEURE", 10),
        ("STATUT_AVANT", 24), ("STATUT_APRES", 24), ("MOTIF", 34),
        ("FEUILLE_SOURCE", 18), ("SESSION_WINDOWS", 18),
    )
    _reference_sheet(
        sheet,
        "HISTORIQUE",
        "Journal des actions, ecrit par les macros et jamais efface automatiquement.",
        headers,
        [],
    )


def _movements_sheet(sheet: Worksheet) -> None:
    headers = (
        ("REFERENCE_MOUVEMENT", 20), ("DATE", 12), ("TYPE", 10),
        ("REFERENCE_PIECE", 16), ("EMPLACEMENT", 14), ("QUANTITE", 12),
        ("STOCK_AVANT", 12), ("STOCK_APRES", 12), ("MATRICULE", 14),
        ("MOTIF", 30), ("SOURCE", 14),
    )
    _reference_sheet(
        sheet,
        "MOUVEMENTS_STOCK",
        "Lecture seule. Le stock augmente a la confirmation de stockage et diminue "
        "a la sortie confirmee: cette regle est appliquee par le serveur.",
        headers,
        [],
    )


def _red_cage_sheet(
    sheet: Worksheet,
    catalogue_rows: int,
    ranges: dict[str, str] | None = None,
    location_rows: int = 0,
    db=None,
) -> None:
    _operational_sheet(
        sheet,
        "RED_CAGE",
        "Lots en quarantaine. Aucune sortie du Red Cage sans decision ni "
        "justification, et une decision non validee ne libere rien dans SLCC.",
        RED_CAGE_COLUMNS,
        formulas=_sheet_formulas(RED_CAGE_COLUMNS, catalogue_rows),
        ranges=ranges,
        catalogue_rows=catalogue_rows,
        location_rows=location_rows,
        db=db,
    )


def _config_sheet(sheet: Worksheet, api_base: str, db=None) -> None:
    """Salt, endpoint and the code digests. Hidden and protected."""
    sheet["A1"] = "CONFIGURATION - NE PAS MODIFIER"
    sheet["A1"].font = Font(size=12, bold=True, color="9B1C1C")
    sheet["A2"] = "SEL"
    sheet["B2"] = CODE_SALT
    sheet["A3"] = "API_SLCC"
    sheet["B3"] = api_base
    for row in (2, 3):
        sheet.cell(row=row, column=1).font = Font(size=10, bold=True, color=INK)

    sheet["A4"] = "MATRICULE"
    sheet["B4"] = "EMPREINTE_CODE"
    for reference in ("A4", "B4"):
        sheet[reference].font = HEADER_FONT
        sheet[reference].fill = HEADER_FILL

    for offset, (matricule, digest) in enumerate(validation_digests(db)):
        row = 5 + offset
        sheet.cell(row=row, column=1, value=matricule).font = Font(size=10)
        # Only the digest ever reaches the file.
        sheet.cell(row=row, column=2, value=digest).font = Font(size=9, name="Consolas")

    sheet.column_dimensions["A"].width = 16
    sheet.column_dimensions["B"].width = 70


# -------------------------------------------------------------------- ribbon
#: A ribbon tab is plain XML, unlike form-control buttons which need drawing
#: parts openpyxl cannot write. It also puts the three actions in the same place
#: on every sheet, which is what an operator needs.
RIBBON_XML = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<customUI xmlns="http://schemas.microsoft.com/office/2009/07/customui">
  <ribbon>
    <tabs>
      <tab id="tabSLCC" label="SLCC" insertBeforeMso="TabHome">
        <group id="grpMaker" label="Operateur">
          <button id="btnFinish" label="Terminer ma tache" size="large"
                  imageMso="AcceptTask" onAction="TerminerMaTache"/>
          <button id="btnFix" label="Corriger une ligne rejetee" size="large"
                  imageMso="EditTask" onAction="CorrigerEtResoumettre"/>
        </group>
        <group id="grpChecker" label="Responsable">
          <button id="btnApprove" label="Valider" size="large"
                  imageMso="AcceptInvitation" onAction="Valider"/>
          <button id="btnReject" label="Rejeter" size="large"
                  imageMso="DeclineInvitation" onAction="Rejeter"/>
        </group>
        <group id="grpSync" label="SLCC">
          <button id="btnRefresh" label="Verifier la connexion SLCC" size="large"
                  imageMso="RefreshAll" onAction="VerifierConnexionSLCC"/>
          <button id="btnSync" label="Enregistrer &amp; synchroniser" size="large"
                  imageMso="ServerPublish" onAction="EnregistrerEtSynchroniser"/>
        </group>
      </tab>
    </tabs>
  </ribbon>
</customUI>
"""


# ------------------------------------------------------------------ assembly
def _protect(sheet: Worksheet, password: str | None = None) -> None:
    """Lock the sheet but let the macros keep writing to it."""
    sheet.protection.sheet = True
    sheet.protection.enableFormatCells = False
    #: In this schema a flag means "forbidden", and both are left open. Blocking
    #: the selection of locked cells would also block a checker from clicking the
    #: submitted line they have to validate. Colour says where to type instead.
    sheet.protection.selectLockedCells = False
    sheet.protection.selectUnlockedCells = False
    if password:
        sheet.protection.password = password


def build_workbook(api_base: str = DEFAULT_API_BASE, db=None) -> bytes:
    """The complete `.xlsm`, ready to be dropped in the shared folder.

    Pass a session and ARTICLES carries the live stock; without one the sheet
    says so rather than showing a figure it cannot vouch for.
    """
    catalogue = whap_source.load_catalogue()
    stock = live_stock(db) if db is not None else None
    tolerances = article_tolerances(db) if db is not None else {}
    summary = whap_source.catalogue_summary()
    generated_at = datetime.now(timezone.utc)

    book = Workbook()
    book.remove(book.active)

    names = vba_source.SHEETS

    # Addresses now, contents last: the sheets below aim at these ranges, but the
    # lists themselves belong at the end of the tab strip.
    choices = _choice_lists(db)
    ranges = _list_ranges(choices)

    _home_sheet(book.create_sheet(names["home"]), summary, generated_at)
    _users_sheet(book.create_sheet(names["users"]), db)
    _articles_sheet(
        book.create_sheet(names["articles"]), catalogue, stock, generated_at, tolerances
    )
    _bom_sheet(book.create_sheet(names["bom"]), catalogue, stock, generated_at)

    #: How many addresses EMPLACEMENTS will hold. Counted here because the
    #: storage dropdown has to name a range before that sheet is written.
    location_count = len({article.location for article in catalogue})

    _operational_sheet(
        book.create_sheet(names["reception"]),
        "RECEPTION",
        "Livraisons fournisseur. Une reception ne modifie jamais le stock.",
        RECEPTION_COLUMNS,
        formulas=_sheet_formulas(RECEPTION_COLUMNS, len(catalogue)),
        ranges=ranges,
        catalogue_rows=len(catalogue),
        location_rows=location_count,
        db=db,
    )
    _operational_sheet(
        book.create_sheet(names["inspection"]),
        "INSPECTION",
        "Controle par echantillonnage. Une inspection ne modifie jamais le stock.",
        INSPECTION_COLUMNS,
        formulas=_sheet_formulas(INSPECTION_COLUMNS, len(catalogue)),
        ranges=ranges,
        catalogue_rows=len(catalogue),
        location_rows=location_count,
        db=db,
    )
    _operational_sheet(
        book.create_sheet(names["quality"]),
        "QUALITE",
        "Decision qualite. Une validation qualite ne modifie jamais le stock: "
        "elle autorise le stockage.",
        QUALITY_COLUMNS,
        formulas=_sheet_formulas(QUALITY_COLUMNS, len(catalogue)),
        ranges=ranges,
        catalogue_rows=len(catalogue),
        location_rows=location_count,
        db=db,
    )
    _red_cage_sheet(
        book.create_sheet(names["red_cage"]),
        len(catalogue),
        ranges=ranges,
        location_rows=location_count,
        db=db,
    )
    _operational_sheet(
        book.create_sheet(names["warehouse"]),
        "WAREHOUSE",
        "Confirmation de stockage. C'est la seule operation qui augmente le stock.",
        WAREHOUSE_COLUMNS,
        formulas=_sheet_formulas(WAREHOUSE_COLUMNS, len(catalogue)),
        ranges=ranges,
        catalogue_rows=len(catalogue),
        location_rows=location_count,
        db=db,
    )
    _movements_sheet(book.create_sheet(names["movements"]))
    _operational_sheet(
        book.create_sheet(names["production"]),
        "PRODUCTION",
        "Demandes des lignes. Une demande ne diminue jamais le stock.",
        PRODUCTION_COLUMNS,
        formulas=_sheet_formulas(PRODUCTION_COLUMNS, len(catalogue)),
        ranges=ranges,
        catalogue_rows=len(catalogue),
        location_rows=location_count,
        db=db,
    )
    _operational_sheet(
        book.create_sheet(names["issues"]),
        "SORTIES",
        "Sortie confirmee. C'est la seule operation qui diminue le stock.",
        ISSUE_COLUMNS,
        formulas=_sheet_formulas(ISSUE_COLUMNS, len(catalogue)),
        ranges=ranges,
        catalogue_rows=len(catalogue),
        location_rows=location_count,
        db=db,
    )
    _locations_sheet(book.create_sheet(names["locations"]), catalogue)
    _history_sheet(book.create_sheet(names["history"]))

    lists = book.create_sheet(LISTS_SHEET)
    _lists_sheet(lists, choices)
    lists.sheet_state = "veryHidden"
    _protect(lists, password="SLCC-LISTES")

    config = book.create_sheet(names["config"])
    _config_sheet(config, api_base, db)
    # Very hidden: not in the sheet list, and not in the unhide dialog either.
    config.sheet_state = "veryHidden"
    _protect(config, password="SLCC-CONFIG")

    for key in ("users", "articles", "bom", "locations", "movements", "history"):
        _protect(book[names[key]])

    # Operational sheets are protected too, but their entry grid is unlocked:
    # an operator types freely, and only a submitted line becomes read-only.
    for key in ("reception", "inspection", "quality", "red_cage", "warehouse",
                "production", "issues"):
        _protect(book[names[key]])

    buffer = io.BytesIO()
    book.save(buffer)
    book.close()

    return _repackage_as_xlsm(buffer.getvalue())


def _repackage_as_xlsm(xlsx: bytes) -> bytes:
    """Add the VBA project and the ribbon, and retype the package as macro-enabled."""
    if not VBA_ASSET.exists():
        raise FileNotFoundError(
            f"{VBA_ASSET} manquant. Executer scripts/build_vba_project.py une fois."
        )

    with zipfile.ZipFile(io.BytesIO(xlsx)) as source:
        parts = {name: source.read(name) for name in source.namelist()}

    content_types = parts["[Content_Types].xml"].decode("utf-8")
    content_types = content_types.replace(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml",
        "application/vnd.ms-excel.sheet.macroEnabled.main+xml",
    ).replace(
        "</Types>",
        '<Override PartName="/xl/vbaProject.bin"'
        ' ContentType="application/vnd.ms-office.vbaProject"/>'
        '<Override PartName="/customUI/customUI14.xml"'
        ' ContentType="application/xml"/></Types>',
    )
    parts["[Content_Types].xml"] = content_types.encode("utf-8")

    workbook_rels = parts["xl/_rels/workbook.xml.rels"].decode("utf-8")
    workbook_rels = workbook_rels.replace(
        "</Relationships>",
        '<Relationship Id="rIdVbaProject"'
        ' Type="http://schemas.microsoft.com/office/2006/relationships/vbaProject"'
        ' Target="vbaProject.bin"/></Relationships>',
    )
    parts["xl/_rels/workbook.xml.rels"] = workbook_rels.encode("utf-8")

    root_rels = parts["_rels/.rels"].decode("utf-8")
    root_rels = root_rels.replace(
        "</Relationships>",
        '<Relationship Id="rIdCustomUI"'
        ' Type="http://schemas.microsoft.com/office/2007/relationships/ui/extensibility"'
        ' Target="customUI/customUI14.xml"/></Relationships>',
    )
    parts["_rels/.rels"] = root_rels.encode("utf-8")

    parts["xl/vbaProject.bin"] = VBA_ASSET.read_bytes()
    parts["customUI/customUI14.xml"] = RIBBON_XML.encode("utf-8")

    out = io.BytesIO()
    with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as archive:
        for name, payload in parts.items():
            archive.writestr(name, payload)
    return out.getvalue()


def workbook_summary(content: bytes | None = None) -> dict:
    """Sheet names and row counts, for the web page and the tests."""
    payload = content if content is not None else build_workbook()
    book = load_workbook(io.BytesIO(payload), read_only=True, keep_vba=True)
    try:
        sheets = []
        for sheet in book.worksheets:
            rows = max(0, sheet.max_row - HEADER_ROW)
            sheets.append({"name": sheet.title, "rows": rows})
        return {
            "workbook": WORKBOOK_NAME,
            "sheet_count": len(sheets),
            "sheets": sheets,
            "size_bytes": len(payload),
        }
    finally:
        book.close()
