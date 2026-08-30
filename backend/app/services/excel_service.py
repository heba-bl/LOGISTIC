"""Real .xlsx generation with openpyxl.

The plant works with spreadsheets. SLCC does not replace them: it produces them,
reads them back and adds the control layer on top. Everything generated here
opens in Microsoft Excel as a normal workbook - named tables, autofilters, frozen
panes, data validation, date and number formats.

Design is deliberately sober: one dark header, one accent, and colour used only
to carry a status. No decoration.

Two families of files:

* one workbook per zone (`SLCC_Receiving.xlsx`, ...), each with a SAISIE sheet in
  the exact import format plus the zone history;
* one shared workbook (`SLCC_Logistics_Flow.xlsx`) consolidating every zone.
"""

from __future__ import annotations

import io
from datetime import datetime, timezone
from typing import Any, Iterable, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.timeutils import to_local
from app.models.enums import ImportType, LotStatus, PartSize
from app.models.flow import Inspection, Lot, QualityValidation, Reception
from app.models.organization import User
from app.models.production import ProductionRequest
from app.models.system import AuditLog
from app.models.vehicle import Vehicle, VehicleBomLine
from app.models.warehouse import Stock, StockMovement, WarehouseLocation
from app.repositories import LotRepository, PartRepository
from app.services import import_service

# --------------------------------------------------------------------- styling
INK = "1F2937"
HEADER_BG = "374151"
HEADER_FG = "FFFFFF"
MUTED = "6B7280"
BORDER = "D1D5DB"

#: Status colours - the only colour in the file, and only on status cells.
STATUS_FILL = {
    "OK": "E7F6EC",
    "WARN": "FDF3E2",
    "CRIT": "FCEBEA",
}
STATUS_FONT = {
    "OK": "1B7F3B",
    "WARN": "9A6209",
    "CRIT": "B42318",
}

TITLE_FONT = Font(size=14, bold=True, color=INK)
SUBTITLE_FONT = Font(size=9, color=MUTED, italic=True)
HEADER_FONT = Font(size=10, bold=True, color=HEADER_FG)
HEADER_FILL = PatternFill("solid", start_color=HEADER_BG)
THIN = Side(style="thin", color=BORDER)
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

DATE_FORMAT = "DD/MM/YYYY HH:MM"
INT_FORMAT = "#,##0"
PCT_FORMAT = "0.0"

SYNTHETIC_NOTICE = (
    "Jeu de donnees synthetique - demonstration SLCC. Ces donnees sont generees "
    "et ne proviennent pas de l'entreprise."
)


def _status_kind(value: Any) -> str | None:
    """Map a status string onto one of the three functional colours."""
    text = str(value or "").upper()
    if text in {
        "ACCEPTED",
        "APPROVED",
        "CONFORM",
        "STORED",
        "ISSUED",
        "OK",
        "VALIDE",
        "APPLIED",
    }:
        return "OK"
    if text in {
        "ACCEPTED_WITH_TOLERANCE",
        "QUALITY_PENDING",
        "PENDING_INSPECTION",
        "INSPECTION_IN_PROGRESS",
        "PENDING_REVIEW",
        "SUBMITTED",
        "PREPARING",
        "READY",
        "PENDING",
        "A VERIFIER",
    }:
        return "WARN"
    if text in {
        "QUANTITY_MISMATCH",
        "RED_CAGE",
        "REJECTED",
        "NON_CONFORM",
        "CANCELLED",
        "INVALID",
        "FAILED",
    }:
        return "CRIT"
    return None


def _sheet_title(sheet: Worksheet, title: str, subtitle: str, columns: int) -> None:
    sheet["A1"] = title
    sheet["A1"].font = TITLE_FONT
    sheet["A2"] = subtitle
    sheet["A2"].font = SUBTITLE_FONT
    if columns > 1:
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=columns)
        sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=columns)
    sheet.row_dimensions[1].height = 22
    sheet.row_dimensions[3].height = 6


def write_table(
    sheet: Worksheet,
    *,
    title: str,
    subtitle: str,
    headers: Sequence[str],
    rows: Iterable[Sequence[Any]],
    table_name: str,
    status_column: str | None = None,
    number_columns: Sequence[str] = (),
    date_columns: Sequence[str] = (),
    widths: dict[str, int] | None = None,
) -> int:
    """Write a professional data block and return the number of data rows."""
    _sheet_title(sheet, title, subtitle, len(headers))

    header_row = 4
    for index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=header_row, column=index, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = CELL_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[header_row].height = 26

    status_index = headers.index(status_column) + 1 if status_column in headers else None
    number_indexes = {headers.index(name) + 1 for name in number_columns if name in headers}
    date_indexes = {headers.index(name) + 1 for name in date_columns if name in headers}

    count = 0
    for offset, row in enumerate(rows, start=0):
        excel_row = header_row + 1 + offset
        count += 1
        for index, value in enumerate(row, start=1):
            cell = sheet.cell(row=excel_row, column=index, value=value)
            cell.border = CELL_BORDER
            cell.font = Font(size=10, color=INK)

            if index in number_indexes:
                cell.number_format = INT_FORMAT
                cell.alignment = Alignment(horizontal="right")
            elif index in date_indexes:
                cell.number_format = DATE_FORMAT
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.alignment = Alignment(vertical="center")

            if status_index and index == status_index:
                kind = _status_kind(value)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if kind:
                    cell.fill = PatternFill("solid", start_color=STATUS_FILL[kind])
                    cell.font = Font(size=10, bold=True, color=STATUS_FONT[kind])

    last_row = header_row + max(count, 1)

    # Named Excel table: filters and banded rows come with it.
    reference = f"A{header_row}:{get_column_letter(len(headers))}{last_row}"
    table = Table(displayName=table_name, ref=reference)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleLight8", showRowStripes=True, showColumnStripes=False
    )
    sheet.add_table(table)

    sheet.freeze_panes = sheet.cell(row=header_row + 1, column=1)

    # Column widths: from the header and a sample of the values.
    for index, header in enumerate(headers, start=1):
        letter = get_column_letter(index)
        if widths and header in widths:
            sheet.column_dimensions[letter].width = widths[header]
            continue
        longest = len(str(header))
        for excel_row in range(header_row + 1, min(last_row, header_row + 40) + 1):
            value = sheet.cell(row=excel_row, column=index).value
            longest = max(longest, len(str(value)) if value is not None else 0)
        sheet.column_dimensions[letter].width = min(max(longest + 3, 10), 46)

    return count


def add_dropdown(
    sheet: Worksheet, column: str, values: Sequence[str], *, first_row: int = 5, last_row: int = 500
) -> None:
    """Attach a closed list to a column, as an operational file would have."""
    formula = '"' + ",".join(values) + '"'
    validation = DataValidation(type="list", formula1=formula, allow_blank=True, showDropDown=True)
    validation.error = "Valeur non autorisee pour cette colonne."
    validation.errorTitle = "Saisie invalide"
    sheet.add_data_validation(validation)
    validation.add(f"{column}{first_row}:{column}{last_row}")


def _local(value: datetime | None) -> datetime | None:
    """Excel has no timezone: store the local wall time."""
    if value is None:
        return None
    return to_local(value).replace(tzinfo=None)


# ----------------------------------------------------------------- data blocks
def operators_rows(db: Session) -> tuple[list[str], list[list[Any]]]:
    headers = ["MATRICULE", "NOM", "PRENOM", "ROLE", "ZONE", "STATUT"]
    users = (
        db.execute(select(User).order_by(User.employee_number)).scalars().all()
    )
    rows = [
        [
            user.employee_number,
            (user.last_name or user.full_name.split(" ")[-1]).upper(),
            user.first_name or user.full_name.split(" ")[0],
            user.role.label if user.role else "",
            user.zone.value if user.zone else (user.service or ""),
            "ACTIF" if user.is_active else "INACTIF",
        ]
        for user in users
    ]
    return headers, rows


def parts_rows(db: Session) -> tuple[list[str], list[list[Any]]]:
    headers = [
        "REFERENCE",
        "DESIGNATION",
        "CATEGORIE",
        "TAILLE",
        "UNITE",
        "STOCK_SECURITE",
        "CONSO_JOUR",
        "TOLERANCE_%",
        "STOCK_DISPO",
    ]
    rows = []
    for part in PartRepository(db).all_active():
        rows.append(
            [
                part.reference,
                part.designation,
                part.category.name if part.category else "",
                part.size_class.value,
                part.unit,
                part.safety_stock,
                part.average_daily_consumption,
                part.reception_tolerance_percent,
                part.stock.quantity_available if part.stock else 0,
            ]
        )
    return headers, rows


def bom_rows(db: Session) -> tuple[list[str], list[list[Any]]]:
    headers = [
        "VEHICULE",
        "REFERENCE",
        "DESIGNATION",
        "SYSTEME",
        "SOUS_SYSTEME",
        "CATEGORIE",
        "TAILLE",
        "QTE_PAR_VEHICULE",
        "UNITE",
        "FOURNISSEUR",
        "GERE_EN_STOCK",
    ]
    rows = []
    query = (
        select(VehicleBomLine, Vehicle)
        .join(Vehicle, Vehicle.id == VehicleBomLine.vehicle_id)
        .order_by(VehicleBomLine.system_code, VehicleBomLine.part_reference)
    )
    for line, vehicle in db.execute(query).all():
        rows.append(
            [
                vehicle.code,
                line.part_reference,
                line.part_description,
                line.system_label,
                line.subsystem,
                line.category,
                line.size_class.value,
                line.quantity_per_vehicle,
                line.unit,
                line.supplier_code,
                "OUI" if line.is_managed else "NON",
            ]
        )
    return headers, rows


def receiving_rows(db: Session) -> tuple[list[str], list[list[Any]]]:
    headers = [
        "DATE",
        "RECEPTION",
        "LOT",
        "REFERENCE",
        "FOURNISSEUR",
        "QTE_ATTENDUE",
        "QTE_RECUE",
        "ECART",
        "TOLERANCE_%",
        "BON_LIVRAISON",
        "MATRICULE",
        "STATUT",
    ]
    rows = []
    query = (
        select(Reception)
        .join(Lot, Lot.id == Reception.lot_id)
        .order_by(Reception.id.desc())
    )
    for reception in db.execute(query).scalars().all():
        lot = reception.lot
        rows.append(
            [
                _local(reception.received_at),
                reception.reference,
                lot.lot_number,
                lot.part.reference,
                lot.supplier.name,
                reception.quantity_expected,
                reception.quantity_received,
                reception.quantity_gap,
                reception.tolerance_percent_applied,
                reception.delivery_note or "",
                reception.received_by.employee_number if reception.received_by else "",
                reception.status.value,
            ]
        )
    return headers, rows


def inspection_rows(db: Session) -> tuple[list[str], list[list[Any]]]:
    headers = [
        "DATE",
        "INSPECTION",
        "LOT",
        "REFERENCE",
        "QTE_LOT",
        "ECHANTILLON",
        "CONFORMES",
        "NON_CONFORMES",
        "TAUX_%",
        "SEUIL_%",
        "MATRICULE",
        "RESULTAT",
        "OBSERVATION",
    ]
    rows = []
    for inspection in db.execute(select(Inspection).order_by(Inspection.id.desc())).scalars():
        lot = inspection.lot
        rows.append(
            [
                _local(inspection.inspected_at),
                inspection.reference,
                lot.lot_number,
                lot.part.reference,
                lot.quantity_received,
                inspection.sample_size,
                inspection.sample_size - inspection.defects_found,
                inspection.defects_found,
                inspection.defect_rate_percent,
                inspection.defect_threshold_percent,
                inspection.inspector.employee_number if inspection.inspector else "",
                inspection.result.value,
                inspection.observations or "",
            ]
        )
    return headers, rows


def quality_rows(db: Session) -> tuple[list[str], list[list[Any]]]:
    headers = [
        "DATE",
        "LOT",
        "REFERENCE",
        "DECISION",
        "QTE_APPROUVEE",
        "MATRICULE",
        "JUSTIFICATION",
    ]
    rows = []
    for validation in db.execute(
        select(QualityValidation).order_by(QualityValidation.id.desc())
    ).scalars():
        lot = validation.lot
        rows.append(
            [
                _local(validation.decided_at),
                lot.lot_number,
                lot.part.reference,
                validation.decision.value,
                validation.quantity_approved,
                validation.decided_by.employee_number if validation.decided_by else "",
                validation.justification,
            ]
        )
    return headers, rows


def red_cage_rows(db: Session) -> tuple[list[str], list[list[Any]]]:
    headers = [
        "LOT",
        "REFERENCE",
        "FOURNISSEUR",
        "QUANTITE",
        "DATE_RECEPTION",
        "MOTIF_BLOCAGE",
        "STATUT",
    ]
    rows = []
    for lot in LotRepository(db).in_stage([LotStatus.RED_CAGE]):
        rows.append(
            [
                lot.lot_number,
                lot.part.reference,
                lot.supplier.name,
                lot.quantity_received,
                _local(lot.received_at),
                lot.blocked_reason or "",
                lot.status.value,
            ]
        )
    return headers, rows


def warehouse_rows(db: Session) -> tuple[list[str], list[list[Any]]]:
    headers = [
        "EMPLACEMENT",
        "ZONE",
        "CAPACITE",
        "OCCUPE",
        "LIBRE",
        "OCCUPATION_%",
        "REFERENCES",
        "STATUT",
    ]
    rows = []
    locations = (
        db.execute(
            select(WarehouseLocation).order_by(
                WarehouseLocation.zone, WarehouseLocation.position
            )
        )
        .scalars()
        .all()
    )
    for location in locations:
        stored = [
            lot.part.reference
            for lot in location.lots
            if lot.quantity_available > 0
        ]
        ratio = location.occupancy_percent
        status = (
            "CRIT"
            if ratio >= location.critical_threshold_percent
            else "WARN"
            if ratio >= location.warning_threshold_percent
            else "OK"
        )
        rows.append(
            [
                location.code,
                location.zone,
                location.capacity,
                location.occupied,
                location.free_capacity,
                ratio,
                ", ".join(sorted(set(stored))),
                {"OK": "OK", "WARN": "A VERIFIER", "CRIT": "SATURE"}[status],
            ]
        )
    return headers, rows


def stock_rows(db: Session) -> tuple[list[str], list[list[Any]]]:
    headers = [
        "REFERENCE",
        "DESIGNATION",
        "CATEGORIE",
        "DISPONIBLE",
        "RESERVE",
        "STOCK_SECURITE",
        "DERNIER_MOUVEMENT",
    ]
    rows = []
    for stock in db.execute(select(Stock)).scalars():
        part = stock.part
        rows.append(
            [
                part.reference,
                part.designation,
                part.category.name if part.category else "",
                stock.quantity_available,
                stock.quantity_reserved,
                part.safety_stock,
                _local(stock.last_movement_at),
            ]
        )
    return headers, rows


def production_rows(db: Session) -> tuple[list[str], list[list[Any]]]:
    headers = [
        "DATE",
        "DEMANDE",
        "STATION",
        "LIGNE",
        "REFERENCE",
        "QTE_DEMANDEE",
        "QTE_SORTIE",
        "PRIORITE",
        "MATRICULE_LEADER",
        "MATRICULE_VALIDEUR",
        "STATUT",
    ]
    rows = []
    for request in db.execute(
        select(ProductionRequest).order_by(ProductionRequest.id.desc())
    ).scalars():
        rows.append(
            [
                _local(request.created_on),
                request.reference,
                request.station.code,
                request.station.production_line or "",
                request.part.reference,
                request.quantity_requested,
                request.quantity_issued,
                request.priority,
                request.requested_by.employee_number if request.requested_by else "",
                request.approved_by.employee_number if request.approved_by else "",
                request.status.value,
            ]
        )
    return headers, rows


def movement_rows(db: Session) -> tuple[list[str], list[list[Any]]]:
    headers = [
        "DATE",
        "MOUVEMENT",
        "TYPE",
        "REFERENCE",
        "QUANTITE",
        "AVANT",
        "APRES",
        "LOT",
        "EMPLACEMENT",
        "MATRICULE",
        "MOTIF",
    ]
    rows = []
    for movement in db.execute(
        select(StockMovement).order_by(StockMovement.id.desc())
    ).scalars():
        rows.append(
            [
                _local(movement.occurred_at),
                movement.reference,
                movement.movement_type.value,
                movement.part.reference,
                movement.quantity,
                movement.quantity_before,
                movement.quantity_after,
                movement.lot.lot_number if movement.lot else "",
                movement.location.code if movement.location else "",
                movement.actor.employee_number if movement.actor else "",
                movement.reason or "",
            ]
        )
    return headers, rows


def audit_rows(db: Session, limit: int = 1000) -> tuple[list[str], list[list[Any]]]:
    headers = [
        "DATE",
        "ACTION",
        "OBJET",
        "REFERENCE",
        "QUANTITE",
        "AVANT",
        "APRES",
        "MATRICULE_SAISIE",
        "MATRICULE_VALIDATION",
        "DECISION",
        "FICHIER_SOURCE",
        "MOTIF",
    ]
    rows = []
    entries = (
        db.execute(select(AuditLog).order_by(AuditLog.id.desc()).limit(limit)).scalars().all()
    )
    for entry in entries:
        rows.append(
            [
                _local(entry.occurred_at),
                entry.action.value,
                entry.entity_type,
                entry.entity_reference or "",
                entry.quantity,
                entry.status_before or "",
                entry.status_after or "",
                entry.maker_reference or entry.actor_reference or "",
                entry.checker_reference or "",
                entry.decision or "",
                entry.source_file or "",
                entry.reason or "",
            ]
        )
    return headers, rows


# ------------------------------------------------------------------ workbooks
ZONE_SHEETS = {
    "RECEIVING": ("RECEPTION", receiving_rows, ImportType.RECEPTION),
    "INSPECTION": ("INSPECTION", inspection_rows, ImportType.INSPECTION),
    "QUALITY": ("QUALITE", quality_rows, None),
    "WAREHOUSE": ("WAREHOUSE", warehouse_rows, None),
    "PRODUCTION": ("PRODUCTION", production_rows, ImportType.PRODUCTION_REQUEST),
}

NUMBER_COLUMNS = (
    "QTE_ATTENDUE", "QTE_RECUE", "ECART", "QUANTITE", "QTE_LOT", "ECHANTILLON",
    "CONFORMES", "NON_CONFORMES", "CAPACITE", "OCCUPE", "LIBRE", "DISPONIBLE",
    "RESERVE", "STOCK_SECURITE", "QTE_DEMANDEE", "QTE_SORTIE", "QTE_APPROUVEE",
    "AVANT", "APRES", "QTE_PAR_VEHICULE", "STOCK_DISPO",
)
DATE_COLUMNS = ("DATE", "DATE_RECEPTION", "DERNIER_MOUVEMENT")


def _write_block(
    sheet: Worksheet, title: str, subtitle: str, headers, rows, table_name: str
) -> int:
    return write_table(
        sheet,
        title=title,
        subtitle=subtitle,
        headers=headers,
        rows=rows,
        table_name=table_name,
        status_column="STATUT" if "STATUT" in headers else ("RESULTAT" if "RESULTAT" in headers else ("DECISION" if "DECISION" in headers else None)),
        number_columns=NUMBER_COLUMNS,
        date_columns=DATE_COLUMNS,
    )


def entry_examples(db: Session, import_type: ImportType, limit: int = 12) -> list[list[Any]]:
    """Rows for the SAISIE sheet, drawn from the real catalogue.

    The file that sits in the shared folder is one an operator has already
    filled in - an empty grid would say nothing about how the exchange works.
    Every value points at a record that exists, so the file can be imported and
    validated as-is during the demonstration.
    """
    from app.models.catalog import Part, Supplier
    from app.models.production import ProductionStation

    rows: list[list[Any]] = []

    if import_type is ImportType.RECEPTION:
        parts = list(db.execute(select(Part).order_by(Part.reference)).scalars())
        suppliers = list(db.execute(select(Supplier).order_by(Supplier.code)).scalars())
        if not parts or not suppliers:
            return rows
        for index in range(limit):
            part = parts[index % len(parts)]
            supplier = suppliers[index % len(suppliers)]
            expected = 480 if part.size_class is PartSize.SMALL else 40
            # One line in six carries a gap, so the tolerance rule is visible.
            received = expected - 5 if index == 5 else expected
            rows.append([
                part.reference,
                supplier.code,
                expected,
                received,
                f"BL-{supplier.code}-{4100 + index}",
                "Ecart constate au comptage" if received != expected else None,
            ])
        return rows

    if import_type is ImportType.INSPECTION:
        lots = list(
            db.execute(
                select(Lot)
                .where(Lot.status == LotStatus.PENDING_INSPECTION)
                .order_by(Lot.id.desc())
                .limit(limit)
            ).scalars()
        )
        for index, lot in enumerate(lots):
            sample = max(3, round(lot.quantity_received * 0.05))
            defects = 1 if index == 2 else 0
            rows.append([
                lot.lot_number,
                sample,
                defects,
                "Bavure sur le bord de coupe" if defects else "Echantillon conforme",
            ])
        return rows

    stations = list(db.execute(select(ProductionStation).order_by(ProductionStation.code)).scalars())
    parts = list(db.execute(select(Part).order_by(Part.reference)).scalars())
    if not stations or not parts:
        return rows
    for index in range(limit):
        part = parts[(index * 3) % len(parts)]
        station = stations[index % len(stations)]
        rows.append([
            station.code,
            part.reference,
            60 if part.size_class is PartSize.SMALL else 6,
            (index % 3) + 1,
            "Besoin pour le lancement de serie" if index % 4 == 0 else None,
        ])
    return rows


def _entry_sheet(
    sheet: Worksheet, import_type: ImportType, examples: list[list[Any]] | None = None
) -> None:
    """The SAISIE sheet: exactly the columns the import expects."""
    columns = [name.upper() for name, _ in import_service.COLUMNS[import_type]]
    required = [required for _, required in import_service.COLUMNS[import_type]]

    _sheet_title(
        sheet,
        f"SAISIE - {import_type.value}",
        "Remplir une ligne par enregistrement, puis importer ce fichier dans SLCC. "
        "Les colonnes en gras sont obligatoires. Ne pas renommer les en-tetes. "
        "Aucune saisie ne modifie le stock: elle doit etre validee dans SLCC par "
        "un responsable different de la personne qui a saisi. "
        + SYNTHETIC_NOTICE,
        len(columns),
    )

    header_row = 4
    for index, (name, is_required) in enumerate(
        zip(columns, required), start=1
    ):
        # The header MUST be the exact column name: the import parser matches on
        # it. Required columns are signalled by the font, not by an asterisk.
        cell = sheet.cell(row=header_row, column=index, value=name)
        cell.font = Font(size=10, bold=True, color=HEADER_FG) if is_required else Font(
            size=10, color=HEADER_FG
        )
        cell.fill = HEADER_FILL
        cell.border = CELL_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.comment = None
        sheet.column_dimensions[get_column_letter(index)].width = max(len(name) + 6, 16)

    sheet.row_dimensions[header_row].height = 26
    sheet.freeze_panes = sheet.cell(row=header_row + 1, column=1)

    # Lines already entered by an operator, then an empty grid for the next ones.
    filled = examples or []
    first_data_row = header_row + 1
    for offset, values in enumerate(filled):
        for index, value in enumerate(values, start=1):
            cell = sheet.cell(row=first_data_row + offset, column=index, value=value)
            cell.font = Font(size=10, color=INK)
            cell.border = CELL_BORDER
            if isinstance(value, int):
                cell.number_format = INT_FORMAT
                cell.alignment = Alignment(horizontal="right")

    blank_start = first_data_row + len(filled)
    for excel_row in range(blank_start, blank_start + 30):
        for index in range(1, len(columns) + 1):
            sheet.cell(row=excel_row, column=index).border = CELL_BORDER


def build_zone_workbook(db: Session, zone: str, *, prefill: bool = False) -> bytes:
    """One workbook for one zone: entry sheet plus history.

    `prefill` is what separates the two uses of the same file. Downloaded from
    the application it is a blank template an operator is about to fill in; laid
    down in the shared folder it ships with lines already entered, which is what
    the zone actually exchanges today.
    """
    zone = zone.upper()
    if zone not in ZONE_SHEETS:
        from app.core.exceptions import ValidationError

        raise ValidationError(f"Zone inconnue: {zone}")

    label, builder, import_type = ZONE_SHEETS[zone]
    workbook = Workbook()
    workbook.remove(workbook.active)

    if import_type is not None:
        entry = workbook.create_sheet("SAISIE")
        _entry_sheet(
            entry, import_type, entry_examples(db, import_type) if prefill else None
        )

    history = workbook.create_sheet(label[:31])
    headers, rows = builder(db)
    _write_block(
        history,
        f"{label} - donnees consolidees",
        SYNTHETIC_NOTICE,
        headers,
        rows,
        f"T_{zone}",
    )

    if zone == "QUALITY":
        red = workbook.create_sheet("RED_CAGE")
        rc_headers, rc_rows = red_cage_rows(db)
        _write_block(red, "RED CAGE - lots bloques", SYNTHETIC_NOTICE, rc_headers, rc_rows, "T_REDCAGE")

    if zone == "WAREHOUSE":
        stock_sheet = workbook.create_sheet("STOCK")
        s_headers, s_rows = stock_rows(db)
        _write_block(stock_sheet, "STOCK DISPONIBLE", SYNTHETIC_NOTICE, s_headers, s_rows, "T_STOCK")

        moves = workbook.create_sheet("MOUVEMENTS")
        m_headers, m_rows = movement_rows(db)
        _write_block(moves, "MOUVEMENTS DE STOCK", SYNTHETIC_NOTICE, m_headers, m_rows, "T_MOVES")

    buffer = io.BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


README_LINES = [
    ("SLCC - FICHIER LOGISTIQUE PARTAGE", "title"),
    (SYNTHETIC_NOTICE, "muted"),
    ("", None),
    ("A QUOI SERT CE FICHIER", "section"),
    ("Ce classeur consolide les donnees des cinq zones de l'usine.", None),
    ("Il est genere par SLCC et reflete l'etat de la base a l'instant de l'export.", None),
    ("", None),
    ("COMMENT L'UTILISER", "section"),
    ("1. Chaque zone dispose de son propre fichier (SLCC_Receiving.xlsx, ...).", None),
    ("2. L'operateur remplit la feuille SAISIE de son fichier de zone.", None),
    ("3. Le fichier est importe dans SLCC (page Donnees operationnelles).", None),
    ("4. Les donnees restent EN ATTENTE DE VALIDATION.", None),
    ("5. Le responsable de la zone valide ou rejette dans SLCC.", None),
    ("6. Ce n'est qu'apres validation que les donnees entrent dans le systeme.", None),
    ("", None),
    ("REGLE DU STOCK", "section"),
    ("Le stock n'augmente QU'APRES confirmation du stockage.", None),
    ("Le stock ne diminue QU'APRES confirmation de la sortie.", None),
    ("Excel ne modifie jamais directement le stock.", None),
    ("", None),
    ("MAKER / CHECKER", "section"),
    ("MAKER   : l'operateur qui saisit les donnees.", None),
    ("CHECKER : le responsable de zone qui verifie et valide.", None),
    ("Le maker ne peut jamais valider sa propre saisie.", None),
    ("Aucun mot de passe ne figure dans ce fichier: la validation se fait dans SLCC.", None),
    ("", None),
    ("FEUILLES DU CLASSEUR", "section"),
    ("README            Ce mode d'emploi", None),
    ("OPERATORS         Matricules, roles et zones des operateurs", None),
    ("PARTS             References gerees en stock", None),
    ("VEHICLE_BOM       Nomenclature du vehicule (donnees synthetiques)", None),
    ("RECEIVING         Receptions et controle des quantites", None),
    ("INSPECTION        Echantillonnages et resultats", None),
    ("QUALITY           Decisions qualite", None),
    ("RED_CAGE          Lots bloques en quarantaine", None),
    ("WAREHOUSE         Emplacements et occupation", None),
    ("PRODUCTION        Demandes de production", None),
    ("STOCK_MOVEMENTS   Journal des mouvements de stock", None),
    ("AUDIT             Journal d'audit (qui, quoi, quand, pourquoi)", None),
]


def _readme_sheet(sheet: Worksheet, db: Session, generated_at: datetime) -> None:
    sheet.column_dimensions["A"].width = 96
    row = 1
    for text, kind in README_LINES:
        cell = sheet.cell(row=row, column=1, value=text)
        if kind == "title":
            cell.font = Font(size=15, bold=True, color=INK)
            sheet.row_dimensions[row].height = 24
        elif kind == "muted":
            cell.font = SUBTITLE_FONT
        elif kind == "section":
            cell.font = Font(size=11, bold=True, color=HEADER_BG)
            sheet.row_dimensions[row].height = 20
        else:
            cell.font = Font(size=10, color=INK)
        row += 1

    row += 1
    sheet.cell(row=row, column=1, value="GENERATION").font = Font(
        size=11, bold=True, color=HEADER_BG
    )
    row += 1
    stamp = to_local(generated_at).strftime("%d/%m/%Y %H:%M")
    sheet.cell(row=row, column=1, value=f"Genere le {stamp} par SLCC.").font = Font(
        size=10, color=INK
    )

    vehicle = db.execute(select(Vehicle).limit(1)).scalar_one_or_none()
    if vehicle is not None:
        bom_count = len(vehicle.bom_lines)
        row += 1
        sheet.cell(
            row=row,
            column=1,
            value=f"Nomenclature {vehicle.code}: {bom_count} references synthetiques.",
        ).font = Font(size=10, color=INK)


def build_global_workbook(db: Session) -> bytes:
    """The shared workbook: twelve sheets consolidating every zone."""
    generated_at = datetime.now(timezone.utc)
    workbook = Workbook()
    workbook.remove(workbook.active)

    _readme_sheet(workbook.create_sheet("README"), db, generated_at)

    blocks = [
        ("OPERATORS", "OPERATEURS", operators_rows, "T_OPERATORS"),
        ("PARTS", "REFERENCES GEREES", parts_rows, "T_PARTS"),
        ("VEHICLE_BOM", "NOMENCLATURE VEHICULE", bom_rows, "T_BOM"),
        ("RECEIVING", "RECEPTIONS", receiving_rows, "T_RECEIVING"),
        ("INSPECTION", "INSPECTIONS", inspection_rows, "T_INSPECTION"),
        ("QUALITY", "DECISIONS QUALITE", quality_rows, "T_QUALITY"),
        ("RED_CAGE", "RED CAGE - LOTS BLOQUES", red_cage_rows, "T_RED_CAGE"),
        ("WAREHOUSE", "EMPLACEMENTS", warehouse_rows, "T_WAREHOUSE"),
        ("PRODUCTION", "DEMANDES DE PRODUCTION", production_rows, "T_PRODUCTION"),
        ("STOCK_MOVEMENTS", "MOUVEMENTS DE STOCK", movement_rows, "T_MOVEMENTS"),
        ("AUDIT", "JOURNAL D AUDIT", audit_rows, "T_AUDIT"),
    ]

    for sheet_name, title, builder, table_name in blocks:
        sheet = workbook.create_sheet(sheet_name)
        headers, rows = builder(db)
        _write_block(sheet, title, SYNTHETIC_NOTICE, headers, rows, table_name)

    # Closed lists where a value must belong to a known set.
    operators = workbook["OPERATORS"]
    add_dropdown(operators, "F", ["ACTIF", "INACTIF"], last_row=200)
    parts = workbook["PARTS"]
    add_dropdown(parts, "D", ["SMALL", "LARGE"], last_row=400)

    buffer = io.BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


def workbook_summary(db: Session) -> dict:
    """Row counts per sheet, for the Data page."""
    counts = {
        "OPERATORS": len(operators_rows(db)[1]),
        "PARTS": len(parts_rows(db)[1]),
        "VEHICLE_BOM": len(bom_rows(db)[1]),
        "RECEIVING": len(receiving_rows(db)[1]),
        "INSPECTION": len(inspection_rows(db)[1]),
        "QUALITY": len(quality_rows(db)[1]),
        "RED_CAGE": len(red_cage_rows(db)[1]),
        "WAREHOUSE": len(warehouse_rows(db)[1]),
        "PRODUCTION": len(production_rows(db)[1]),
        "STOCK_MOVEMENTS": len(movement_rows(db)[1]),
        "AUDIT": len(audit_rows(db)[1]),
    }
    return {
        "workbook": "SLCC_Logistics_Flow.xlsx",
        "generated_at": datetime.now(timezone.utc),
        "sheet_count": len(counts) + 1,  # + README
        "row_count": sum(counts.values()),
        "sheets": [{"name": name, "rows": rows} for name, rows in counts.items()],
    }


ZONE_BUILDERS = {
    "RECEIVING": receiving_rows,
    "INSPECTION": inspection_rows,
    "QUALITY": quality_rows,
    "WAREHOUSE": warehouse_rows,
    "PRODUCTION": production_rows,
}


def zone_table(db: Session, zone: str, limit: int = 500) -> dict:
    """The zone data as JSON, for the in-app viewer."""
    zone = zone.upper()
    builder = ZONE_BUILDERS.get(zone)
    if builder is None:
        from app.core.exceptions import ValidationError

        raise ValidationError(f"Zone inconnue: {zone}")

    headers, rows = builder(db)
    trimmed = rows[:limit]
    serialised = [
        [
            value.isoformat() if isinstance(value, datetime) else value
            for value in row
        ]
        for row in trimmed
    ]
    return {
        "zone": zone,
        "columns": headers,
        "rows": serialised,
        "total_rows": len(rows),
        "returned_rows": len(serialised),
        "status_column": (
            "STATUT" if "STATUT" in headers
            else "RESULTAT" if "RESULTAT" in headers
            else "DECISION" if "DECISION" in headers
            else None
        ),
    }
