"""Excel import and the Maker-Checker validation workflow.

The rule this module exists to enforce:

    Data imported from a spreadsheet is NEVER definitive until a habilitated
    checker - a different person from the maker - has confirmed it in SLCC.

Consequences, all implemented below:

* uploading a file creates NO business record and moves NO stock;
* the file is parsed, every row validated, and the batch stored as PENDING_REVIEW;
* approval is refused if the checker is the maker, is inactive, or does not hold
  a role habilitated for that import type;
* only on approval are the rows applied, and they are applied through the very
  same services the UI uses, so every business rule still applies;
* rejection requires a comment, applies nothing, and keeps everything traceable;
* the maker, the checker, both roles, both timestamps, the decision, the comment,
  the file name and its SHA-256 hash are all persisted.
"""

from __future__ import annotations

import csv
import hashlib
import io
import json
from dataclasses import dataclass
from datetime import datetime, timezone
from typing import Any, Callable

from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.core.exceptions import NotFoundError, ValidationError, WorkflowError
from app.models.enums import (
    AuditAction,
    ImportRowStatus,
    ImportStatus,
    ImportType,
    RoleName,
    ValidationDecision,
)
from app.models.imports import DataImport, ImportRow
from app.models.organization import User
from app.repositories import (
    LotRepository,
    PartRepository,
    ProductionRepository,
    SupplierRepository,
    UserRepository,
)
from app.services import audit_service, inspection_service, production_service, reception_service

# ---------------------------------------------------------------------------
# Habilitation matrix
# ---------------------------------------------------------------------------

#: Roles allowed to ENTER data for a given import type.
MAKER_ROLES: dict[ImportType, tuple[RoleName, ...]] = {
    ImportType.RECEPTION: (RoleName.RECEPTIONIST, RoleName.RECEPTION_MANAGER),
    ImportType.INSPECTION: (RoleName.QUALITY_INSPECTOR, RoleName.QUALITY_MANAGER),
    ImportType.PRODUCTION_REQUEST: (RoleName.STATION_LEADER, RoleName.PRODUCTION_MANAGER),
}

#: Roles allowed to VALIDATE data for a given import type.
#: A Quality Inspector is deliberately absent from the INSPECTION checkers: an
#: inspector can never validate their own inspection as a Quality Manager.
CHECKER_ROLES: dict[ImportType, tuple[RoleName, ...]] = {
    ImportType.RECEPTION: (RoleName.RECEPTION_MANAGER, RoleName.LOGISTICS_MANAGER),
    ImportType.INSPECTION: (RoleName.QUALITY_MANAGER, RoleName.LOGISTICS_MANAGER),
    ImportType.PRODUCTION_REQUEST: (
        RoleName.PRODUCTION_MANAGER,
        RoleName.LOGISTICS_MANAGER,
    ),
}

#: Columns expected in the spreadsheet, per import type.
COLUMNS: dict[ImportType, tuple[tuple[str, bool], ...]] = {
    # (column name, required)
    ImportType.RECEPTION: (
        ("part_reference", True),
        ("supplier_code", True),
        ("quantity_expected", True),
        ("quantity_received", True),
        ("delivery_note", False),
        ("notes", False),
    ),
    ImportType.INSPECTION: (
        ("lot_number", True),
        ("sample_size", True),
        ("defects_found", True),
        ("observations", False),
    ),
    ImportType.PRODUCTION_REQUEST: (
        ("station_code", True),
        ("part_reference", True),
        ("quantity", True),
        ("priority", False),
        ("notes", False),
    ),
}

TEMPLATE_EXAMPLES: dict[ImportType, list[dict[str, Any]]] = {
    ImportType.RECEPTION: [
        {
            "part_reference": "BR-145",
            "supplier_code": "DEL",
            "quantity_expected": 500,
            "quantity_received": 500,
            "delivery_note": "BL-2026-0042",
            "notes": "",
        }
    ],
    ImportType.INSPECTION: [
        {
            "lot_number": "LOT-2026-009",
            "sample_size": 20,
            "defects_found": 0,
            "observations": "",
        }
    ],
    ImportType.PRODUCTION_REQUEST: [
        {
            "station_code": "ST-02",
            "part_reference": "BR-145",
            "quantity": 20,
            "priority": 2,
            "notes": "",
        }
    ],
}


@dataclass
class ParsedRow:
    row_number: int
    payload: dict[str, Any]
    error: str | None = None


# ---------------------------------------------------------------------------
# Parsing
# ---------------------------------------------------------------------------
def _normalise_header(value: Any) -> str:
    """Normalise a column header.

    Tolerant of what a hand-edited spreadsheet carries: case, surrounding spaces,
    a trailing required-marker, and spaces used instead of underscores.
    """
    text = str(value or "").strip().rstrip("*").strip()
    return text.lower().replace(" ", "_")


#: How far down a sheet the header row may sit. Operational files carry a title
#: block above the table, so the header is rarely the first row.
HEADER_SEARCH_DEPTH = 15

#: Name of the entry sheet in the workbooks SLCC generates.
ENTRY_SHEET_NAMES = ("saisie", "entry", "data")


def _pick_sheet(workbook, import_type: ImportType):
    """Prefer the entry sheet of a generated workbook, else the active one."""
    for name in workbook.sheetnames:
        if name.strip().lower() in ENTRY_SHEET_NAMES:
            return workbook[name]
    return workbook.active


def _find_header(rows: list[tuple], import_type: ImportType) -> int:
    """Locate the header row.

    A file produced by SLCC has a title, a notice and a spacer above the table,
    so the header cannot be assumed to be row 1. The header is the first row that
    carries the expected column names.
    """
    expected = {name for name, _ in COLUMNS[import_type]}
    required = {name for name, is_required in COLUMNS[import_type] if is_required}

    best_index, best_score = -1, 0
    for index, values in enumerate(rows[:HEADER_SEARCH_DEPTH]):
        if values is None:
            continue
        normalised = {_normalise_header(value) for value in values if value is not None}
        score = len(expected & normalised)
        if required <= normalised:
            return index
        if score > best_score:
            best_index, best_score = index, score

    if best_score >= 2:
        return best_index

    raise ValidationError(
        "No header row found. Expected the columns: "
        + ", ".join(sorted(expected))
    )


def _read_table(
    content: bytes, filename: str, import_type: ImportType
) -> list[dict[str, Any]]:
    """Read an .xlsx or .csv file into a list of dicts keyed by column name."""
    lowered = filename.lower()

    if lowered.endswith(".csv"):
        text = content.decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        if reader.fieldnames is None:
            raise ValidationError("The CSV file has no header row")
        return [
            {_normalise_header(key): value for key, value in row.items()} for row in reader
        ]

    if not lowered.endswith((".xlsx", ".xlsm")):
        raise ValidationError(
            f"Unsupported file type '{filename}'. Provide a .xlsx or .csv file."
        )

    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependency is declared
        raise ValidationError("openpyxl is required to read Excel files") from exc

    try:
        workbook = load_workbook(io.BytesIO(content), data_only=True)
    except Exception as exc:  # noqa: BLE001 - any corrupt file must be reported
        raise ValidationError(f"Unreadable Excel file: {exc}") from exc

    sheet = _pick_sheet(workbook, import_type)
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        workbook.close()
        raise ValidationError("The spreadsheet is empty")

    header_index = _find_header(rows, import_type)
    columns = [_normalise_header(cell) for cell in rows[header_index]]

    records: list[dict[str, Any]] = []
    for values in rows[header_index + 1 :]:
        if values is None or all(value is None or str(value).strip() == "" for value in values):
            continue  # skip blank lines, including the pre-formatted empty grid
        records.append(dict(zip(columns, values)))

    workbook.close()
    return records


def _as_int(value: Any, field: str) -> int:
    if value is None or str(value).strip() == "":
        raise ValueError(f"'{field}' is required")
    try:
        return int(float(str(value).strip()))
    except (TypeError, ValueError):
        raise ValueError(f"'{field}' must be a whole number (got '{value}')") from None


def _as_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _validate_row(
    db: Session, import_type: ImportType, row_number: int, raw: dict[str, Any]
) -> ParsedRow:
    """Structural and referential validation of one line. No write occurs here."""
    payload: dict[str, Any] = {}

    try:
        for column, required in COLUMNS[import_type]:
            value = raw.get(column)
            if required and (value is None or str(value).strip() == ""):
                raise ValueError(f"'{column}' is required")

        if import_type is ImportType.RECEPTION:
            reference = _as_text(raw.get("part_reference"))
            part = PartRepository(db).by_reference(str(reference).upper())
            if part is None:
                raise ValueError(f"unknown part reference '{reference}'")

            supplier_code = str(_as_text(raw.get("supplier_code")) or "").upper()
            supplier = next(
                (s for s in SupplierRepository(db).all_active() if s.code.upper() == supplier_code),
                None,
            )
            if supplier is None:
                raise ValueError(f"unknown supplier code '{supplier_code}'")

            expected = _as_int(raw.get("quantity_expected"), "quantity_expected")
            received = _as_int(raw.get("quantity_received"), "quantity_received")
            if expected <= 0:
                raise ValueError("'quantity_expected' must be strictly positive")
            if received < 0:
                raise ValueError("'quantity_received' cannot be negative")

            payload = {
                "part_id": part.id,
                "part_reference": part.reference,
                "supplier_id": supplier.id,
                "supplier_code": supplier.code,
                "quantity_expected": expected,
                "quantity_received": received,
                "delivery_note": _as_text(raw.get("delivery_note")),
                "notes": _as_text(raw.get("notes")),
            }

        elif import_type is ImportType.INSPECTION:
            lot_number = str(_as_text(raw.get("lot_number")) or "").upper()
            lot = LotRepository(db).by_number(lot_number)
            if lot is None:
                raise ValueError(f"unknown lot '{lot_number}'")

            sample = _as_int(raw.get("sample_size"), "sample_size")
            defects = _as_int(raw.get("defects_found"), "defects_found")
            if sample <= 0:
                raise ValueError("'sample_size' must be strictly positive")
            if sample > lot.quantity_received:
                raise ValueError(
                    f"'sample_size' {sample} exceeds the {lot.quantity_received} units of {lot_number}"
                )
            if defects < 0 or defects > sample:
                raise ValueError("'defects_found' must be between 0 and the sample size")

            payload = {
                "lot_id": lot.id,
                "lot_number": lot.lot_number,
                "sample_size": sample,
                "defects_found": defects,
                "observations": _as_text(raw.get("observations")),
            }

        else:  # PRODUCTION_REQUEST
            station_code = str(_as_text(raw.get("station_code")) or "").upper()
            station = next(
                (s for s in ProductionRepository(db).stations() if s.code.upper() == station_code),
                None,
            )
            if station is None:
                raise ValueError(f"unknown station '{station_code}'")

            reference = str(_as_text(raw.get("part_reference")) or "").upper()
            part = PartRepository(db).by_reference(reference)
            if part is None:
                raise ValueError(f"unknown part reference '{reference}'")

            quantity = _as_int(raw.get("quantity"), "quantity")
            if quantity <= 0:
                raise ValueError("'quantity' must be strictly positive")

            priority_raw = raw.get("priority")
            priority = 3 if priority_raw in (None, "") else _as_int(priority_raw, "priority")
            if priority not in (1, 2, 3):
                raise ValueError("'priority' must be 1, 2 or 3")

            payload = {
                "station_id": station.id,
                "station_code": station.code,
                "part_id": part.id,
                "part_reference": part.reference,
                "quantity": quantity,
                "priority": priority,
                "notes": _as_text(raw.get("notes")),
            }

    except ValueError as exc:
        return ParsedRow(row_number=row_number, payload=dict(raw), error=str(exc))

    return ParsedRow(row_number=row_number, payload=payload)


def _next_reference(db: Session) -> str:
    year = datetime.now(timezone.utc).year
    count = db.execute(
        select(func.count())
        .select_from(DataImport)
        .where(DataImport.reference.like(f"IMP-{year}-%"))
    ).scalar_one()
    return f"IMP-{year}-{count + 1:03d}"


# ---------------------------------------------------------------------------
# 1. Upload (MAKER)
# ---------------------------------------------------------------------------
def create_import(
    db: Session,
    *,
    import_type: ImportType,
    filename: str,
    content: bytes,
    maker_id: int,
    notes: str | None = None,
) -> DataImport:
    """Register an uploaded spreadsheet. Creates NO business record.

    The batch lands in PENDING_REVIEW: it is data waiting for a decision, not
    data in the system.
    """
    maker = UserRepository(db).optional(maker_id)
    if maker is None:
        raise NotFoundError(f"User {maker_id} not found")
    if not maker.is_active:
        raise ValidationError(
            f"{maker.employee_number} is inactive and cannot submit data"
        )

    allowed = MAKER_ROLES[import_type]
    if maker.role.name not in allowed:
        raise ValidationError(
            f"{maker.identity} is not habilitated to enter {import_type.value} data. "
            f"Allowed roles: {', '.join(role.value for role in allowed)}."
        )

    if not content:
        raise ValidationError("The uploaded file is empty")

    records = _read_table(content, filename, import_type)
    if not records:
        raise ValidationError("The file contains no data row")

    parsed = [
        _validate_row(db, import_type, index, raw)
        for index, raw in enumerate(records, start=2)  # row 1 is the header
    ]
    valid = [row for row in parsed if row.error is None]

    digest = hashlib.sha256(content).hexdigest()

    batch = DataImport(
        reference=_next_reference(db),
        import_type=import_type,
        status=ImportStatus.PENDING_REVIEW,
        source_filename=filename,
        source_hash=digest,
        source_size_bytes=len(content),
        row_count=len(parsed),
        valid_row_count=len(valid),
        invalid_row_count=len(parsed) - len(valid),
        maker_id=maker.id,
        maker_reference=maker.employee_number,
        maker_role=maker.role_name,
        maker_service=maker.service,
        notes=notes,
    )
    db.add(batch)
    db.flush()

    for row in parsed:
        db.add(
            ImportRow(
                import_id=batch.id,
                row_number=row.row_number,
                status=ImportRowStatus.INVALID if row.error else ImportRowStatus.PENDING,
                payload_json=json.dumps(row.payload, default=str),
                error_message=row.error,
            )
        )
    db.flush()

    audit_service.record(
        db,
        action=AuditAction.IMPORT_SUBMITTED,
        entity_type="data_import",
        entity_id=batch.id,
        entity_reference=batch.reference,
        actor=maker,
        maker=maker,
        source_file=filename,
        source_hash=digest,
        status_after=batch.status.value,
        reason=(
            f"{maker.identity} submitted {filename} "
            f"({batch.valid_row_count} valid / {batch.row_count} rows) "
            f"for {import_type.value} - awaiting validation"
        ),
    )
    return batch


# ---------------------------------------------------------------------------
# 2. Decision (CHECKER)
# ---------------------------------------------------------------------------
def _assert_checker(batch: DataImport, checker: User) -> None:
    """Segregation of duties.

    The checker must be a different, active person holding a role habilitated to
    validate this kind of data.
    """
    if not checker.is_active:
        raise ValidationError(f"{checker.employee_number} is inactive and cannot validate")

    if checker.id == batch.maker_id:
        raise WorkflowError(
            f"{checker.identity} entered this data and cannot validate it. "
            "The checker must be a different person from the maker."
        )

    allowed = CHECKER_ROLES[batch.import_type]
    if checker.role.name not in allowed:
        raise WorkflowError(
            f"{checker.identity} is not habilitated to validate {batch.import_type.value} data. "
            f"Allowed roles: {', '.join(role.value for role in allowed)}."
        )


def _require_pending(db: Session, import_id: int) -> DataImport:
    batch = db.get(DataImport, import_id)
    if batch is None:
        raise NotFoundError(f"Import {import_id} not found")
    if not batch.is_pending:
        raise WorkflowError(
            f"Import {batch.reference} has already been decided ({batch.status.value})"
        )
    return batch


def _apply_reception(db: Session, payload: dict, maker: User) -> str:
    reception = reception_service.create_reception(
        db,
        part_id=payload["part_id"],
        supplier_id=payload["supplier_id"],
        quantity_expected=payload["quantity_expected"],
        quantity_received=payload["quantity_received"],
        delivery_note=payload.get("delivery_note"),
        notes=payload.get("notes"),
        actor_id=maker.id,
    )
    return reception.reference


def _apply_inspection(db: Session, payload: dict, maker: User) -> str:
    from app.models.enums import LotStatus

    lot = LotRepository(db).require(payload["lot_id"])
    if lot.status is LotStatus.PENDING_INSPECTION:
        inspection_service.start_inspection(db, lot_id=lot.id, actor_id=maker.id)
    inspection = inspection_service.record_inspection(
        db,
        lot_id=payload["lot_id"],
        sample_size=payload["sample_size"],
        defects_found=payload["defects_found"],
        observations=payload.get("observations"),
        actor_id=maker.id,
    )
    return inspection.reference


def _apply_production_request(db: Session, payload: dict, maker: User) -> str:
    request = production_service.create_request(
        db,
        station_id=payload["station_id"],
        part_id=payload["part_id"],
        quantity=payload["quantity"],
        priority=payload.get("priority", 3),
        notes=payload.get("notes"),
        actor_id=maker.id,
        submit_immediately=True,
    )
    return request.reference


APPLIERS: dict[ImportType, Callable[[Session, dict, User], str]] = {
    ImportType.RECEPTION: _apply_reception,
    ImportType.INSPECTION: _apply_inspection,
    ImportType.PRODUCTION_REQUEST: _apply_production_request,
}


def approve_import(
    db: Session, *, import_id: int, checker_id: int, comment: str | None = None
) -> DataImport:
    """Validate a batch and only then write it to the business tables."""
    batch = _require_pending(db, import_id)

    checker = UserRepository(db).optional(checker_id)
    if checker is None:
        raise NotFoundError(f"User {checker_id} not found")
    _assert_checker(batch, checker)

    if batch.valid_row_count == 0:
        raise ValidationError(
            f"Import {batch.reference} has no valid row to apply. Reject it instead."
        )

    maker = batch.maker
    applier = APPLIERS[batch.import_type]

    applied = 0
    for row in batch.rows:
        if row.status is not ImportRowStatus.PENDING:
            continue
        payload = json.loads(row.payload_json)
        try:
            # Applied through the normal services, so every business rule -
            # tolerance, sampling, workflow guards - still applies. Those
            # services know nothing about imports, so they stamp only the
            # operator; the provenance is attached afterwards.
            mark = audit_service.high_water_mark(db)
            row.result_reference = applier(db, payload, maker)
            audit_service.stamp_provenance(
                db,
                since_id=mark,
                maker=maker,
                checker=checker,
                decision=ValidationDecision.APPROVED.value,
                source_file=batch.source_filename,
                source_hash=batch.source_hash,
            )
            row.status = ImportRowStatus.APPLIED
            applied += 1
            audit_service.record(
                db,
                action=AuditAction.IMPORT_ROW_APPLIED,
                entity_type="import_row",
                entity_id=row.id,
                entity_reference=row.result_reference,
                actor=checker,
                maker=maker,
                checker=checker,
                source_file=batch.source_filename,
                source_hash=batch.source_hash,
                reason=(
                    f"Row {row.row_number} of {batch.reference} applied "
                    f"-> {row.result_reference}"
                ),
            )
        except Exception as exc:  # noqa: BLE001 - a bad row must not lose the batch
            row.status = ImportRowStatus.FAILED
            row.error_message = str(exc)

    now = datetime.now(timezone.utc)
    batch.status = ImportStatus.APPROVED
    batch.decision = ValidationDecision.APPROVED
    batch.decision_comment = comment
    batch.checker_id = checker.id
    batch.checker_reference = checker.employee_number
    batch.checker_role = checker.role_name
    batch.checker_service = checker.service
    batch.checked_at = now
    batch.applied_row_count = applied
    db.flush()

    audit_service.record(
        db,
        action=AuditAction.IMPORT_APPROVED,
        entity_type="data_import",
        entity_id=batch.id,
        entity_reference=batch.reference,
        actor=checker,
        maker=maker,
        checker=checker,
        decision=ValidationDecision.APPROVED.value,
        source_file=batch.source_filename,
        source_hash=batch.source_hash,
        status_before=ImportStatus.PENDING_REVIEW.value,
        status_after=batch.status.value,
        reason=(
            f"{checker.identity} approved {batch.reference} entered by {maker.identity}: "
            f"{applied}/{batch.valid_row_count} rows applied"
            + (f" - {comment}" if comment else "")
        ),
    )
    return batch


def reject_import(
    db: Session, *, import_id: int, checker_id: int, comment: str
) -> DataImport:
    """Refuse a batch. Nothing is applied; everything stays traceable."""
    if not comment or not comment.strip():
        raise ValidationError("Rejecting an import requires a comment")

    batch = _require_pending(db, import_id)

    checker = UserRepository(db).optional(checker_id)
    if checker is None:
        raise NotFoundError(f"User {checker_id} not found")
    _assert_checker(batch, checker)

    for row in batch.rows:
        if row.status is ImportRowStatus.PENDING:
            row.status = ImportRowStatus.REJECTED

    now = datetime.now(timezone.utc)
    batch.status = ImportStatus.REJECTED
    batch.decision = ValidationDecision.REJECTED
    batch.decision_comment = comment
    batch.checker_id = checker.id
    batch.checker_reference = checker.employee_number
    batch.checker_role = checker.role_name
    batch.checker_service = checker.service
    batch.checked_at = now
    batch.applied_row_count = 0
    db.flush()

    audit_service.record(
        db,
        action=AuditAction.IMPORT_REJECTED,
        entity_type="data_import",
        entity_id=batch.id,
        entity_reference=batch.reference,
        actor=checker,
        maker=batch.maker,
        checker=checker,
        decision=ValidationDecision.REJECTED.value,
        source_file=batch.source_filename,
        source_hash=batch.source_hash,
        status_before=ImportStatus.PENDING_REVIEW.value,
        status_after=batch.status.value,
        reason=(
            f"{checker.identity} rejected {batch.reference} entered by "
            f"{batch.maker.identity}: {comment}"
        ),
    )
    return batch


# ---------------------------------------------------------------------------
# Reads
# ---------------------------------------------------------------------------
def list_imports(
    db: Session, *, status: ImportStatus | None = None, limit: int = 100
) -> list[DataImport]:
    stmt = select(DataImport)
    if status is not None:
        stmt = stmt.where(DataImport.status == status)
    return list(
        db.execute(stmt.order_by(DataImport.id.desc()).limit(limit)).scalars().all()
    )


def get_import(db: Session, import_id: int) -> DataImport:
    batch = db.get(DataImport, import_id)
    if batch is None:
        raise NotFoundError(f"Import {import_id} not found")
    return batch


def eligible_checkers(db: Session, import_id: int) -> list[User]:
    """Active operators who may validate this batch (maker excluded)."""
    batch = get_import(db, import_id)
    allowed = CHECKER_ROLES[batch.import_type]
    return [
        user
        for user in UserRepository(db).all_with_roles()
        if user.is_active and user.id != batch.maker_id and user.role.name in allowed
    ]


def build_template(import_type: ImportType) -> bytes:
    """Generate an .xlsx template with the expected header and one example row."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = import_type.value[:31]

    headers = [column for column, _ in COLUMNS[import_type]]
    sheet.append(headers)

    header_fill = PatternFill("solid", start_color="0C1524")
    for index, (column, required) in enumerate(COLUMNS[import_type], start=1):
        cell = sheet.cell(row=1, column=index)
        cell.font = Font(bold=True, color="E8EFF9")
        cell.fill = header_fill
        sheet.column_dimensions[cell.column_letter].width = max(18, len(column) + 4)
        if required:
            cell.comment = None  # keep the file simple; required columns are documented

    for example in TEMPLATE_EXAMPLES[import_type]:
        sheet.append([example.get(column, "") for column, _ in COLUMNS[import_type]])

    sheet.freeze_panes = "A2"

    buffer = io.BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()
